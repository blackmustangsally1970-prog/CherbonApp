
from flask import (
    Flask, render_template, request, redirect, url_for,
    send_file, make_response, after_this_request, flash,
    jsonify, current_app, Response, session
)

from collections import defaultdict
from config import Config

# Database + Models
from extensions import db
from models import (
    BlockoutDate,
    BlockoutRange,
    Client,
    CourseEnrolment,
    CourseReference,
    CourseFormSubmission,
    DisclaimerState,
    DailyEvent,
    GeneralEnquirySubmission,
    GroupPricing,
    Horse,
    UpgradeItem,
    IncomingSubmission,
    Lesson,
    LessonBlockTag,
    LessonInvite,
    LessonTeacherTag,
    SmsLog,
    Teacher,
    TeacherBlock,
    TeacherGridOverride,
    TeacherHorse,
    TeacherSlot,
    TeacherTime,
    Term,
    Time,
    TrailRideSubmission,
    Users,
    Wedding,
    WeddingAssignment,
    WeddingStaff,
    WeddingStaffUnavailability,
    WeeklyEvent,
    Employee,
    EmployeeHours
)

# Core libs
import os
import io
import re
import json
import hashlib
import secrets
import string
import tempfile
import time
import subprocess
import unicodedata
import openpyxl
from datetime import date, datetime, time, timedelta
from functools import lru_cache, wraps
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

from markupsafe import Markup, escape
from sqlalchemy import func, text
from sqlalchemy.orm import joinedload

# Third‑party libs
import requests
import clicksend_client
from clicksend_client import SmsMessage
from clicksend_client.rest import ApiException

# Conditional imports
if os.environ.get("DISABLE_WEASYPRINT") != "1":
    from weasyprint import HTML
else:
    HTML = None

if os.environ.get("DISABLE_PLAYWRIGHT") != "1":
    from playwright.sync_api import sync_playwright
else:
    sync_playwright = None

# Security helpers
from werkzeug.security import generate_password_hash, check_password_hash

# Custom helpers
from helpers_jf import (
    extract_riders_from_submission,
    get_main_contact_fields,
    TRAIL_FORM_ID,
    GENERAL_ENQUIRY_FORM_ID,       # ⭐ already added
    parse_general_enquiry_payload  # ⭐ already added
)



BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_ROOT = os.path.join(BASE_DIR, "lesson_logs")
os.makedirs(LOG_ROOT, exist_ok=True)

WEDDING_SMS_ROOT = os.path.join(BASE_DIR, "weddingsms")
os.makedirs(WEDDING_SMS_ROOT, exist_ok=True)

CSV_ROOT = os.path.join(BASE_DIR, "csv")
os.makedirs(CSV_ROOT, exist_ok=True)


@lru_cache(maxsize=1)
def get_static_clients():
    return db.session.query(Client).order_by(Client.full_name).all()

@lru_cache(maxsize=1)
def get_static_horses():
    return db.session.query(Horse).order_by(Horse.horse).all()

@lru_cache(maxsize=1)
def get_static_times():
    return db.session.query(Time).order_by(Time.timerange).all()

@lru_cache(maxsize=1)
def get_static_teachers():
    return db.session.query(Teacher).order_by(Teacher.teacher).all()

@lru_cache(maxsize=1)
def get_static_teacher_time():
    return db.session.query(TeacherTime).order_by(
        TeacherTime.teacher_key,
        TeacherTime.weekday,
        TeacherTime.time
    ).all()

def get_incomplete_days(employee_id):
    sql = """
        SELECT work_date, sign_in, sign_out
        FROM employee_hours
        WHERE employee_id = %s
          AND work_date < CURRENT_DATE
          AND (sign_in IS NULL OR sign_out IS NULL)
        ORDER BY work_date DESC;
    """
    cur = mysql.connection.cursor()
    cur.execute(sql, (employee_id,))
    rows = cur.fetchall()
    cur.close()
    return rows


def smart_proper_name(name):
    if not name:
        return ""

    name = name.strip()
    parts = name.split()

    # Dutch particles that stay lowercase
    lowercase_particles = {
        "van", "der", "den", "de", "het", "ten", "ter", "op", "aan", "bij", "uit", "te"
    }

    fixed_parts = []
    for part in parts:
        p = part.lower()

        # Handle O' prefix
        if p.startswith("o'") and len(p) > 2:
            fixed_parts.append("O'" + p[2:].capitalize())
            continue

        # Handle D' prefix
        if p.startswith("d'") and len(p) > 2:
            fixed_parts.append("D'" + p[2:].capitalize())
            continue

        # Handle Mc prefix
        if p.startswith("mc") and len(p) > 2:
            fixed_parts.append("Mc" + p[2:].capitalize())
            continue

        # Handle Mac prefix
        if p.startswith("mac") and len(p) > 3:
            fixed_parts.append("Mac" + p[3:].capitalize())
            continue

        # Handle hyphens (Mary-Anne)
        if "-" in p:
            fixed_parts.append("-".join(s.capitalize() for s in p.split("-")))
            continue

        # Handle Dutch particles (always lowercase)
        if p in lowercase_particles:
            fixed_parts.append(p)
            continue

        # Default
        fixed_parts.append(p.capitalize())

    return " ".join(fixed_parts)


def fy_week1_monday(year):
    fy_start = date(year, 7, 1)
    return fy_start - timedelta(days=fy_start.weekday())

def build_fy_weeks(year):
    week1 = fy_week1_monday(year)
    fy_end = date(year + 1, 6, 30)

    weeks = []
    wk = 1
    current = week1

    while True:
        start = current
        end = current + timedelta(days=6)

        # If the next week would cross into July, STOP
        if end > fy_end:
            break

        weeks.append({
            "week_number": wk,
            "start": start,
            "end": end
        })

        wk += 1
        current += timedelta(days=7)

    return weeks

def build_daily_summary(selected_fy):
    from collections import defaultdict

    # Parse FY
    fy_start_year = int(selected_fy.split("-")[0])
    fy_end_year   = int(selected_fy.split("-")[1])

    start_date = datetime(fy_start_year, 7, 1).date()
    end_date   = datetime(fy_end_year, 6, 30).date()

    # Running totals
    weekly_running = 0
    ytd_total = 0

    # Weekly attendance counters
    weekly_y = 0
    weekly_c = 0
    weekly_n = 0

    daily_data = []

    # Loop through every day in FY
    current = start_date
    while current <= end_date:

        # Reset weekly totals on Sunday
        if current.weekday() == 6:  # Sunday
            weekly_running = 0
            weekly_y = 0
            weekly_c = 0
            weekly_n = 0

        # Sum payments for the day
        payment_sum = (
            db.session.query(func.sum(Lesson.payment))
            .filter(
                Lesson.lesson_date == current,
                Lesson.payment.isnot(None)
            )
            .scalar()
        ) or 0

        weekly_running += payment_sum

        # Attendance counts
        y_count = (
            db.session.query(func.count(Lesson.lesson_id))
            .filter(
                Lesson.lesson_date == current,
                Lesson.attendance == "Y"
            )
            .scalar()
        ) or 0

        c_count = (
            db.session.query(func.count(Lesson.lesson_id))
            .filter(
                Lesson.lesson_date == current,
                Lesson.attendance == "C"
            )
            .scalar()
        ) or 0

        n_count = (
            db.session.query(func.count(Lesson.lesson_id))
            .filter(
                Lesson.lesson_date == current,
                Lesson.attendance == "N"
            )
            .scalar()
        ) or 0

        # Add to weekly attendance totals
        weekly_y += y_count
        weekly_c += c_count
        weekly_n += n_count

        # Weekly + YTD totals only on Saturday or 30 June
        if current.weekday() == 5 or current == end_date:
            running_total_display = weekly_running
            ytd_total += weekly_running
            ytd_display = ytd_total

            weekly_y_display = weekly_y
            weekly_c_display = weekly_c
            weekly_n_display = weekly_n
        else:
            running_total_display = None
            ytd_display = None

            weekly_y_display = None
            weekly_c_display = None
            weekly_n_display = None

        # Append row
        daily_data.append({
            "date": current,
            "payment": payment_sum,
            "y": y_count,
            "c": c_count,
            "n": n_count,

            "weekly_y": weekly_y_display,
            "weekly_c": weekly_c_display,
            "weekly_n": weekly_n_display,

            "running_total_display": running_total_display,
            "ytd_display": ytd_display,

            "is_saturday": (current.weekday() == 5)
        })

        # Next day
        current += timedelta(days=1)

    return daily_data

def recalc_all_lessons():
    from sqlalchemy import distinct

    # Get all unique clients
    clients = db.session.query(distinct(Lesson.client)).all()
    clients = [c[0] for c in clients if c[0]]

    for client in clients:
        recalc_client_cascade(client)

    db.session.commit()


def recalc_client_cascade(client_name: str):
    lessons = (
        Lesson.query
        .filter_by(client=client_name)
        .order_by(
            db.func.date(Lesson.lesson_date).asc(),
            Lesson.lesson_id.asc()
        )
        .all()
    )

    first = True
    running_balance = 0

    for l in lessons:
        price = l.price_pl or 0
        payment = l.payment or 0
        adjust = l.adjust or 0
        att = (l.attendance or '').strip().upper()

        # FIRST LESSON → PRESERVE IMPORTED carry_fwd
        if first:
            carry = l.carry_fwd or 0
            first = False
        else:
            carry = running_balance

        balance = carry + payment + adjust

        if att in ['Y', 'N']:
            balance -= price

        l.carry_fwd = carry
        l.balance = balance

        running_balance = balance

    db.session.commit()



# -------------------------
#   Helper Functions
# -------------------------

def clean_name(name: str) -> str:
    if not name:
        return ""

    # Trim and collapse multiple spaces
    name = " ".join(name.split())

    # Insert spaces before capital letters (CamelCase fix)
    name = re.sub(r"(?<!^)(?=[A-Z])", " ", name)

    # Proper case
    name = name.title()

    return name


def strip_old_spans(text):
    if not text:
        return ""
    # Remove ANY real <span ...> or </span> tags
    text = re.sub(r"</?span[^>]*>", "", text, flags=re.IGNORECASE)
    # Remove ANY escaped &lt;span ...&gt; or &lt;/span&gt; tags
    text = re.sub(r"&lt;/?span[^&]*&gt;", "", text, flags=re.IGNORECASE)
    return text

def highlight(text, query):
    if not text:
        return ""
    if not query:
        # Just escape the raw text safely
        return Markup(escape(text))

    # 1) Strip any old spans (real or escaped)
    clean = strip_old_spans(text)

    # 2) Case-insensitive search on the raw text
    lower_clean = clean.lower()
    lower_q = query.lower()

    start = lower_clean.find(lower_q)
    if start == -1:
        # No match: just escape safely
        return Markup(escape(clean))

    end = start + len(query)

    # 3) Split into parts
    before = clean[:start]
    match = clean[start:end]
    after = clean[end:]

    # 4) Escape each part, but wrap the MATCH in a span
    before_safe = escape(before)
    match_safe = escape(match)
    after_safe = escape(after)

    return Markup(
        f"{before_safe}<span class='hl'>{match_safe}</span>{after_safe}"
    )

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def log_admin_action(action):
    user = session.get("username") or "system"
    uid  = session.get("user_id") or "?"
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open("admin_actions.log", "a") as f:
        f.write(f"[{ts}] by {user} (ID {uid}) | {action}\n")


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            user = Users.query.get(session.get("user_id"))
            if not user or user.role not in roles:
                return "Access denied", 403
            return f(*args, **kwargs)
        return wrapper
    return decorator

def match_existing_client(name, phone, email):
    q = Client.query

    # Normalise
    name = name.strip().lower() if name else None
    phone = phone.strip() if phone else None
    email = email.strip().lower() if email else None

    matches = []

    if name:
        matches += q.filter(db.func.lower(Client.full_name) == name).all()

    if phone:
        matches += q.filter(Client.mobile == phone).all()

    if email:
        matches += q.filter(db.func.lower(Client.email_primary) == email).all()

    # Remove duplicates using correct PK
    unique = {m.client_id: m for m in matches}.values()

    return list(unique)


def generate_financial_years(start_year=2022):
    from datetime import date

    today = date.today()

    # Determine current FY based on Australian FY (1 July – 30 June)
    if today.month >= 7:
        current_fy_start = today.year
    else:
        current_fy_start = today.year - 1

    # Build list from start_year up to next FY
    fy_list = []
    for y in range(start_year, current_fy_start + 2):
        fy_list.append(f"{y}-{y+1}")

    return fy_list

def get_sundays_for_financial_year(fy_string):
    start_year, end_year = map(int, fy_string.split("-"))

    fy_start = datetime(start_year, 7, 1)
    fy_end = datetime(end_year, 6, 30)

    # Find the first Sunday on or before FY start
    first_sunday = fy_start - timedelta(days=fy_start.weekday() + 1 if fy_start.weekday() != 6 else 0)

    sundays = []
    current = first_sunday

    while current <= fy_end:
        sundays.append(current.date())
        current += timedelta(days=7)

    return sundays





def get_week_window(sunday_date):
    from datetime import timedelta
    start = sunday_date
    end = sunday_date + timedelta(days=6)
    return start, end


def build_weekly_summary(sundays, selected_fy):
    from collections import defaultdict

    weekly_data = []
    running_total = 0

    # Map Python weekday() → our summary columns
    day_map = {
        6: "sun",
        0: "mon",
        1: "tue",
        2: "wed",
        3: "thu",
        4: "fri",
        5: "sat"
    }

    fy_start_year = int(selected_fy.split("-")[0])
    fy_end_year = int(selected_fy.split("-")[1])

    fy_start = datetime(fy_start_year, 7, 1).date()
    fy_end   = datetime(fy_end_year, 6, 30).date()

    for s in sundays:
        week_start, week_end = get_week_window(s)

        # Clamp to FY boundaries
        if week_start < fy_start:
            week_start = fy_start
        if week_end > fy_end:
            week_end = fy_end

        # ⭐ NEW: Calculate week ending (Saturday)
        week_ending = week_start + timedelta(days=6)


        # Attendance counts
        y_count = (
            db.session.query(func.count(Lesson.lesson_id))
            .filter(
                Lesson.lesson_date.between(week_start, week_end),
                Lesson.attendance == "Y"
            )
            .scalar()
        ) or 0

        n_count = (
            db.session.query(func.count(Lesson.lesson_id))
            .filter(
                Lesson.lesson_date.between(week_start, week_end),
                Lesson.attendance == "N"
            )
            .scalar()
        ) or 0

        c_count = (
            db.session.query(func.count(Lesson.lesson_id))
            .filter(
                Lesson.lesson_date.between(week_start, week_end),
                Lesson.attendance == "C"
            )
            .scalar()
        ) or 0

        # ⭐ DAILY BREAKDOWN (Sun → Sat)
        day_totals = defaultdict(float)

        lessons_in_week = Lesson.query.filter(
            Lesson.lesson_date.between(week_start, week_end),
            Lesson.payment.isnot(None)
        ).all()

        for l in lessons_in_week:
            dow = l.lesson_date.weekday()  # 0=Mon ... 6=Sun
            key = day_map.get(dow)
            if key:
                day_totals[key] += l.payment or 0

        # ⭐ Weekly + YTD totals
        weekly_total = sum(day_totals.values())
        running_total += weekly_total

        weekly_data.append({
            "week_start": week_start,
            "week_ending": week_ending,   # ⭐ NEW FIELD

            # Daily totals
            "sun": day_totals["sun"],
            "mon": day_totals["mon"],
            "tue": day_totals["tue"],
            "wed": day_totals["wed"],
            "thu": day_totals["thu"],
            "fri": day_totals["fri"],
            "sat": day_totals["sat"],

            # Totals
            "weekly_total": weekly_total,
            "ytd_total": running_total,

            # Attendance
            "y": y_count,
            "n": n_count,
            "c": c_count,

        })

    # ⭐ Mark highest and lowest attendance weeks
    if weekly_data:
        max_att = max(w["y"] for w in weekly_data)
        min_att = min(w["y"] for w in weekly_data)

        for w in weekly_data:
            w["is_max_attendance"] = (w["y"] == max_att)
            w["is_min_attendance"] = (w["y"] == min_att)

    return weekly_data


def log_lesson_changes(changes, user="system"):
    from datetime import datetime

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_date = datetime.now().strftime("%Y-%m-%d")
    log_path = os.path.join(LOG_ROOT, f"lesson_changes_{log_date}.log")

    # Separate lesson and client keys
    lesson_ids = [lid for lid in changes if isinstance(lid, int)]
    client_ids = [int(str(k).split("_")[1]) for k in changes if str(k).startswith("client_")]

    lessons = db.session.query(Lesson).filter(Lesson.lesson_id.in_(lesson_ids)).all()
    clients = db.session.query(Client).filter(Client.client_id.in_(client_ids)).all()

    # 🔑 keep both client name and lesson date
    lesson_lookup = {
        lesson.lesson_id: (lesson.client, lesson.lesson_date)
        for lesson in lessons
    }
    client_lookup = {client.client_id: client.full_name for client in clients}

    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] by {user}\n")
            for key in sorted(changes, key=lambda x: str(x)):
                if str(key).startswith("client_"):
                    cid = int(str(key).split("_")[1])
                    client = client_lookup.get(cid, f"Client {cid}")
                    lesson_date = ""   # no specific lesson date for client notes
                else:
                    client, lesson_date = lesson_lookup.get(key, (f"Lesson {key}", None))

                fields = changes[key]
                client_fmt = f"{client:<35}"
                date_fmt = lesson_date.strftime("%Y-%m-%d") if lesson_date else "—"

                parts = []
                if "payment" in fields:
                    old, new = fields["payment"]

                    # Normalize formatting
                    old_clean = (old or "").strip().replace("$", "")
                    new_clean = (new or "").strip().replace("$", "")

                    # Always log payment saves
                    parts.append(f"payment saved: '{new_clean}' (old was '{old_clean}')")
                if "attendance" in fields:
                    old, new = fields["attendance"]
                    if old != new:
                        parts.append(f"attendance: '{old}' -> '{new}'")
                if "horse" in fields:
                    old, new = fields["horse"]
                    if old != new:
                        parts.append(f"horse: '{old}' -> '{new}'")
                if "notes" in fields:
                    old, new = fields["notes"]
                    if (old or '').strip() != (new or '').strip():
                        parts.append(f"notes: '{old}' -> '{new}'")

                if parts:
                    # ✅ now includes lesson date
                    f.write(f"  {client_fmt} | {date_fmt} | " + " | ".join(parts) + "\n")
            f.write("\n")
    except Exception as e:
        print(f"Log write failed: {e}")

def log_disclaimer_processed(names):
    """
    Logs processed disclaimers to the main CherbonApp log folder
    using Brisbane time and a clean, consistent format.
    """

    try:
        # Brisbane timezone
        brisbane_tz = pytz.timezone("Australia/Brisbane")
        now_brisbane = datetime.now(brisbane_tz)
        timestamp = now_brisbane.strftime("%Y-%m-%d %H:%M:%S")

        # Correct log folder (the one your viewer actually reads)
        log_dir = "/var/log/cherbonapp"
        os.makedirs(log_dir, exist_ok=True)

        # Daily rotating file
        log_file = os.path.join(log_dir, f"disclaimers_{now_brisbane.strftime('%Y-%m-%d')}.log")

        # Write entries
        with open(log_file, "a", encoding="utf-8") as f:
            for name in names:
                f.write(f"{timestamp} — Processed disclaimer for: {name}\n")

        print(f"[DisclaimerLog] Logged {len(names)} disclaimers to {log_file}")

    except Exception as e:
        print(f"[DisclaimerLog][ERROR] {e}")

def parse_start(tframe):
    """
    Safely extract start time in minutes from 'HH:MM - HH:MM'.
    Returns a large number for invalid/missing times so they sort last.
    NEVER crashes.
    """
    try:
        if not tframe:
            return 99999
        tframe = tframe.replace("–", "-").replace("—", "-")
        start = tframe.split("-")[0].strip()
        if ":" not in start:
            return 99999
        h, m = start.split(":")
        return int(h) * 60 + int(m)
    except:
        return 99999

def group_blocks(rows):
    blocks = defaultdict(list)

    for r in rows:
        # Give Camp a stable time key
        time_key = r["time_frame"] or ("Camp" if r["lesson_type"] == "Camp" else "")
        key = (time_key, r["lesson_type"], r["group_priv"])
        blocks[key].append(r)

    grouped = []
    for (time, ltype, gp), riders in sorted(
        blocks.items(),
        key=lambda k: (
            -1 if k[0][0] in ["Camp", "", None] else parse_start(k[0][0])
        )
    ):
        grouped.append({
            "time": time,
            "lesson_type": ltype,
            "group_priv": gp,
            "riders": sorted(riders, key=lambda x: x["client_name"].lower())
        })

    return grouped


def log_wedding_sms(result, message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []

    for num in result.get("sent_numbers", []):
        lines.append(f"{timestamp} | SENT | {num} | {message}")

    for num in result.get("skipped_numbers", []):
        lines.append(f"{timestamp} | SKIPPED | {num} | {message}")

    for num in result.get("failed_numbers", []):
        lines.append(f"{timestamp} | FAILED | {num} | {message}")

    with open("wedding_sms_log.txt", "a", encoding="utf-8") as f:
        for line in lines:
            f.write(line + "\n")


def parse_dt(dt_str):
    return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M")

def normalize_name_for_lookup(name: str) -> str:
    """
    Normalize a client name for accent-insensitive lookup.
    - Keeps original spelling intact in DB/UI
    - Strips diacritics for search keys
    - Lowercases and removes spaces
    """
    if not name:
        return ""
    # Normalize to NFKD form (decompose accents)
    nfkd = unicodedata.normalize("NFKD", name)
    # Strip combining marks (accents/diacritics)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    # Lowercase and remove spaces for consistent lookup
    return stripped.lower().replace(" ", "")


def log_admin_action(action, user="system"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_date = datetime.now().strftime("%Y-%m-%d")
    log_path = os.path.join(LOG_ROOT, f"admin_actions_{log_date}.log")



    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] by {user} | {action}\n")
    except Exception as e:
        print(f"Admin log write failed: {e}")


def find_matching_clients(name):
    cleaned = name.strip().lower()
    return Clients.query.filter(
        db.func.lower(Clients.name) == cleaned
    ).all()

# ---------------------------------------------------------
# HELPERS (NO INDENT)
# ---------------------------------------------------------

def build_lesson_from_payload(l, client_id, client_name=""):
    lesson_type_raw = (l.get("type") or "").strip()
    lesson_type_norm = lesson_type_raw.lower()

    # Base fields from payload
    time_val   = l.get("time")
    date_val   = l.get("date")
    group_priv = l.get("grouppriv")
    price_val  = l.get("price")

    # --- CAMP mimics VOUCHER CR EXACTLY ---
    if lesson_type_norm in ("camp", "voucher cr", "voucher_cr", "vouchercr"):
        time_val   = ""          # Voucher CR uses empty time
        group_priv = "P"         # Voucher CR uses P
        attendance = "Pending"
        horse      = ""
        payment    = None

    else:
        attendance = "Pending"
        horse      = ""
        payment    = None

    return Lesson(
        lesson_date = date_val,
        time_frame  = time_val,
        lesson_type = lesson_type_raw,
        group_priv  = group_priv,
        price_pl    = price_val,
        client_id   = client_id,
        client      = client_name,
        horse       = horse,
        attendance  = attendance,
        payment     = payment,
    )



def handle_existing_client(data):
    client_id = data.get("existing_client_id")
    lessons   = data.get("lessons", [])

    if not client_id:
        return jsonify(success=False, error="No client selected")

    try:
        for l in lessons:
            # Camp + Voucher CR normalisation happens inside build_lesson_from_payload
            lesson = build_lesson_from_payload(l, client_id, client_name="")
            db.session.add(lesson)

        db.session.commit()
        return jsonify(success=True)

    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e))

def handle_new_client(data):
    name    = data.get("new_client_name")
    mobile  = data.get("mobile")
    lessons = data.get("lessons", [])

    if not name:
        return jsonify(success=False, error="Client name required")

    try:
        client = Client(
            full_name = name,
            mobile    = mobile,
            notes     = "",
        )
        db.session.add(client)
        db.session.flush()   # get client.client_id

        for l in lessons:
            # Camp + Voucher CR normalisation happens inside build_lesson_from_payload
            lesson = build_lesson_from_payload(l, client.client_id, client_name=name)
            db.session.add(lesson)

        db.session.commit()

        # Recalc after commit
        try:
            recalc_client_cascade(name)
        except Exception as e:
            print("Recalc failed:", e)

        return jsonify(success=True)

    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e))


def generate_unique_client_name(base_name):
    cleaned = base_name.strip()

    existing = Clients.query.filter(
        Clients.name.ilike(f"{cleaned}%")
    ).all()

    if not existing:
        return cleaned

    suffix = 2
    while True:
        candidate = f"{cleaned} ({suffix})"
        if not any(c.name == candidate for c in existing):
            return candidate
        suffix += 1

def clean_name(s):
    if not s:
        return ""
    # Collapse multiple spaces, strip ends, title case
    return " ".join(s.strip().split()).title()

def clean_mobile(s):
    if not s:
        return ""
    return s.replace(" ", "").strip()

def extract_number(value):
    """
    Extracts digits from a string like '63kg' or '178 cm' and returns an int.
    Returns None if no digits found.
    """
    if not value:
        return None
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    return int(digits) if digits else None

def compute_sort_order(day_of_week, timerange):
    # Day ordering for global weekly sort (Sunday → Saturday)
    day_index = {
        "Sunday": 1,
        "Monday": 2,
        "Tuesday": 3,
        "Wednesday": 4,
        "Thursday": 5,
        "Friday": 6,
        "Saturday": 7
    }.get(day_of_week, 99)

    # Extract start time from "HH:MM - HH:MM"
    try:
        start_time = timerange.split(" - ")[0]
        hours, minutes = map(int, start_time.split(":"))
        total_minutes = hours * 60 + minutes
    except:
        total_minutes = 9999  # fallback if bad data

    # Return a sortable integer
    # Example: Monday 07:00 → 1*10000 + 420 = 10420
    # We'll renumber these later to 1,2,3...
    return day_index * 10000 + total_minutes

def renumber_all_courses():
    courses = CourseReference.query.order_by(CourseReference.sort_order).all()
    counter = 1
    for c in courses:
        c.sort_order = counter
        counter += 1
    db.session.commit()

def get_active_term_weeks():
    term = Term.query.filter_by(active=True).first()
    if not term:
        return None

    weeks = []
    start = term.start_date
    for i in range(term.weeks):
        week_start = start + timedelta(days=i*7)
        week_end = week_start + timedelta(days=6)
        weeks.append((week_start, week_end))

    return weeks


def normalise_full_name(name: str) -> str:
    if not name:
        return ""

    import unicodedata
    import re

    # Strip accents
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))

    # Replace weird spaces
    name = name.replace("\xa0", " ")

    # Trim + collapse spaces
    name = " ".join(name.strip().split())

    # Split into parts (handles multi‑word names)
    parts = name.split()

    def fix_part(part):
        # Handle O', D', L', etc.
        if re.match(r"^[A-Za-z]'", part):
            return part[0].upper() + "'" + part[2:].capitalize()

        # Handle Mc / Mac prefixes
        if part.lower().startswith("mc") and len(part) > 2:
            return "Mc" + part[2:].capitalize()

        if part.lower().startswith("mac") and len(part) > 3:
            return "Mac" + part[3:].capitalize()

        # Handle hyphenated names (Smith-Jones)
        if "-" in part:
            return "-".join(p.capitalize() for p in part.split("-"))

        # Default proper case
        return part.capitalize()

    # Apply rules to each part
    cleaned = " ".join(fix_part(p) for p in parts)

    return cleaned

def levenshtein(a, b):
    # Simple pure‑Python Levenshtein distance
    if a == b:
        return 0
    if len(a) < len(b):
        return levenshtein(b, a)
    if len(b) == 0:
        return len(a)

    previous_row = range(len(b) + 1)
    for i, c1 in enumerate(a):
        current_row = [i + 1]
        for j, c2 in enumerate(b):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row

    return previous_row[-1]





# ------------------------------
# DISCLAIMER FORM FIELD MAP
# ------------------------------
HEIGHT_FIELDS = ["8", "14", "21", "28", "35", "39"]
WEIGHT_FIELDS = ["9", "15", "22", "29", "37", "38"]
NOTES_FIELDS  = ["11", "17", "25", "32", "60", "61"]

# ------------------------------
# INVITE FORM FIELD MAP
# ------------------------------
INVITE_FULLNAME_FIELDS = ["43", "46", "49", "52", "55", "58", "61", "64", "67", "70"]
INVITE_HEIGHT_FIELDS   = ["44", "47", "50", "53", "56", "59", "62", "65", "68", "71"]
INVITE_WEIGHT_FIELDS   = ["45", "48", "51", "54", "57", "60", "63", "66", "69", "72"]
INVITE_NOTES_FIELDS    = []   # none in this form



def normalize_name(s: str) -> str:
    """
    Normalize a name for matching:
    - strip accents
    - lowercase
    - collapse spaces
    - remove weird unicode spaces
    """
    if not s:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("\xa0", " ")
    return " ".join(s.strip().lower().split())


def parse_jotform_payload(payload, forced_submission_id=None, clients_cache=None, mode="full"):
    """
    mode:
      - "light": parse names/fields only. NO DB client load. NO matching.
      - "full": full matching behaviour (may use clients_cache or DB load).
    clients_cache:
      - Optional preloaded list of Client rows (or tuples) to avoid repeated DB scans.
    """
    import json
    import unicodedata
    from sqlalchemy.util._collections import immutabledict

    # Ensure payload is a real dict
    if isinstance(payload, immutabledict):
        payload = dict(payload)
    if not isinstance(payload, dict):
        try:
            payload = json.loads(payload)
        except Exception:
            print("ERROR: Could not decode payload:", type(payload))
            return []

    # Submission ID
    if forced_submission_id:
        submission_id = str(forced_submission_id)
    else:
        submission_id = str(
            payload.get("id")
            or payload.get("submission_id")
            or payload.get("form_id")
            or ""
        )

    answers = payload.get("answers", {}) or {}

    # ---- Email autodetect ----
    email = ""
    for key, item in answers.items():
        if item.get("type") == "control_email":
            ans = item.get("answer")
            if isinstance(ans, dict):
                email = ans.get("value") or ans.get("text") or ans.get("full") or ""
            else:
                email = ans or ""
            break

    if not email:
        for key, item in answers.items():
            if item.get("name", "").lower() == "email":
                ans = item.get("answer")
                if isinstance(ans, dict):
                    email = ans.get("value") or ans.get("text") or ans.get("full") or ""
                else:
                    email = ans or ""
                break

    email = email or ""

    # Detect form type
    form_id = str(payload.get("form_id") or "")
    is_invite_form = form_id == "253599154628066"

    # Invite token extraction
    invite_token = None
    field3 = answers.get("3")
    if field3:
        ans = field3.get("answer")
        if isinstance(ans, dict):
            invite_token = ans.get("text") or ans.get("value") or ans.get("full")
        else:
            invite_token = ans

    if not invite_token:
        for f in answers.values():
            if f.get("name") == "i_t":
                ans = f.get("answer")
                if isinstance(ans, dict):
                    invite_token = ans.get("text") or ans.get("value") or ans.get("full")
                else:
                    invite_token = ans
                break

    if not invite_token:
        direct = answers.get("i_t")
        if direct:
            ans = direct.get("answer")
            if isinstance(ans, dict):
                invite_token = ans.get("text") or ans.get("value") or ans.get("full")
            else:
                invite_token = ans

    invite_token = invite_token or ""

    # Age field detection
    age_fields = [
        key for key, item in answers.items()
        if item.get("type") in ("control_number", "control_dropdown")
        and "age" in item.get("text", "").lower()
    ]
    age_fields = sorted(age_fields, key=lambda x: int(x))

    # Global fields
    guardian = answers.get("86", {}).get("answer", "") or ""
    mobile = answers.get("87", {}).get("answer", {}).get("full", "") or ""
    email_fallback = answers.get("47", {}).get("answer", "") or ""
    disclaimer = answers.get("63", {}).get("answer", None)
    if email_fallback and not email:
        email = email_fallback

    # Detect rider fullname fields
    if is_invite_form:
        fullname_fields = INVITE_FULLNAME_FIELDS
    else:
        fullname_fields = [
            key for key, item in answers.items()
            if item.get("type") == "control_fullname"
            and item.get("text", "").lower().startswith("rider")
        ]
        fullname_fields = sorted(fullname_fields, key=lambda x: int(x))

    riders = []

    # Matching mode
    do_matching = (mode == "full")

    # ---------------------------------------------------------
    # CLIENT CACHE LOADING — FIXED (NO FULL TABLE SCAN)
    # ---------------------------------------------------------
    client_cache = None
    exact_lookup = None

    if do_matching:
        # If caller did NOT provide a cache, we DO NOT load all clients.
        # We only load clients when we have a rider name.
        client_cache = []
        exact_lookup = {}

    # ---------------------------------------------------------
    # MAIN RIDER LOOP
    # ---------------------------------------------------------
    for idx, fullname_key in enumerate(fullname_fields):
        item = answers.get(fullname_key)
        if not item:
            continue

        pretty = item.get("prettyFormat")
        if pretty:
            raw_name = pretty
        else:
            first = (item.get("answer", {}) or {}).get("first", "")
            last = (item.get("answer", {}) or {}).get("last", "")
            raw_name = f"{first} {last}"

        name_norm = normalize_name(raw_name)
        if not name_norm:
            continue

        # Age
        age_key = age_fields[idx] if idx < len(age_fields) else None
        age = answers.get(age_key, {}).get("answer") if age_key else None

        rider = {
            "name": name_norm,
            "age": age,
            "guardian": guardian,
            "mobile": mobile,
            "email": email,
            "disclaimer": disclaimer,
            "matches": [],
            "jotform_submission_id": submission_id,
            "invite_token": invite_token
        }

        # Height/weight/notes
        if is_invite_form:
            if idx < len(INVITE_HEIGHT_FIELDS):
                height_field = INVITE_HEIGHT_FIELDS[idx]
                weight_field = INVITE_WEIGHT_FIELDS[idx]
                notes_field = INVITE_NOTES_FIELDS[idx] if idx < len(INVITE_NOTES_FIELDS) else None
            else:
                height_field = weight_field = notes_field = None
        else:
            if idx < len(HEIGHT_FIELDS):
                height_field = HEIGHT_FIELDS[idx]
                weight_field = WEIGHT_FIELDS[idx]
                notes_field = NOTES_FIELDS[idx]
            else:
                height_field = weight_field = notes_field = None

        height_val = answers.get(height_field, {}).get("answer") if height_field else None
        weight_val = answers.get(weight_field, {}).get("answer") if weight_field else None
        notes_val = answers.get(notes_field, {}).get("answer") if notes_field else None

        rider["height_cm"] = extract_number(height_val)
        rider["weight_kg"] = extract_number(weight_val)
        rider["notes"] = notes_val or ""

        # ---------------------------------------------------------
        # MATCHING — FIXED (NO FULL TABLE SCAN)
        # ---------------------------------------------------------
        if do_matching:
            # Load only relevant clients ONCE per rider
            if clients_cache is None:
                compact = name_norm.replace(" ", "").replace("-", "")
                like_pattern = f"%{compact}%"

                clients_cache = db.session.query(
                    Client.client_id,
                    Client.full_name,
                    Client.mobile,
                    Client.email_primary,
                    Client.jotform_submission_id
                ).filter(
                    Client.full_name.isnot(None),
                    func.replace(
                        func.replace(func.lower(Client.full_name), " ", ""),
                        "-", ""
                    ).like(like_pattern)
                ).all()

            # Build lookup
            client_cache = []
            for c in clients_cache:
                full_name = getattr(c, "full_name", None)
                if not full_name:
                    continue
                norm = normalize_name(full_name)
                client_cache.append((c, norm, getattr(c, "jotform_submission_id", None)))

            exact_lookup = {norm: c for c, norm, _ in client_cache}

            # ---------------------------------------------------------
            # NEW MATCHING LOGIC — NAME → MOBILE → EMAIL
            # ---------------------------------------------------------
            matched_client = None

            # 1. Exact name match
            if name_norm in exact_lookup:
                matched_client = exact_lookup[name_norm]

            # 2. Mobile match (normalized)
            if not matched_client and mobile:
                mobile_norm = re.sub(r"\D", "", mobile)
                for c, norm, _ in client_cache:
                    c_mobile = re.sub(r"\D", "", (c.mobile or ""))
                    if c_mobile and c_mobile == mobile_norm:
                        matched_client = c
                        break

            # 3. Email match (lowercase)
            if not matched_client and email:
                email_norm = email.strip().lower()
                for c, norm, _ in client_cache:
                    if (c.email_primary or "").strip().lower() == email_norm:
                        matched_client = c
                        break

            # Store match
            if matched_client:
                rider["matches"].append(matched_client)


        riders.append(rider)

    return riders







def create_app():
    print(">>> JOTFORM KEY IN APP:", os.getenv("JOTFORM_API_KEY"))
    app = Flask(__name__)

    app.config["PROPAGATE_EXCEPTIONS"] = True
    app.config['SECRET_KEY'] = os.getenv("SECRET_KEY", "")
    app.config.from_object(Config())
    app.config['JOTFORM_API_KEY'] = os.getenv("JOTFORM_API_KEY", "")

    # Load ClickSend credentials
    app.config['CLICKSEND_USERNAME'] = (
        os.environ.get('CLICKSEND_USERNAME')
        or os.environ.get('APPSETTING_CLICKSEND_USERNAME')
        or ""
    )

    app.config['CLICKSEND_API_KEY'] = (
        os.environ.get('CLICKSEND_API_KEY')
        or os.environ.get('APPSETTING_CLICKSEND_API_KEY')
        or ""
    )



    app.config['SQLALCHEMY_ECHO'] = False
    db.init_app(app)

    # print("IncomingSubmission columns:", IncomingSubmission.__table__.columns.keys())  # <-- remove or wrap

    @app.route("/fetch_gift_vouchers")
    def fetch_gift_vouchers():
        print("FETCH GV ROUTE HIT — BEFORE API KEY CHECK")
        import requests
        from datetime import datetime
        from models import GiftVoucherSubmission
        from helpers_jf import GIFT_VOUCHER_FORM_ID, parse_gift_voucher_payload

        api_key = current_app.config.get("JOTFORM_API_KEY")
        if not api_key:
            flash("Missing JotForm API key", "danger")
            return redirect(url_for("gift_vouchers"))

        # BUILD URL FIRST
        url = f"https://api.jotform.com/form/{GIFT_VOUCHER_FORM_ID}/submissions?apiKey={api_key}"

        # DEBUG PRINT
        print("DEBUG FETCH URL:", url)


        response = requests.get(url)
        print("RAW RESPONSE:", response.text)
        data = response.json()

        if data.get("responseCode") != 200:
            flash("Failed to fetch from JotForm", "danger")
            return redirect(url_for("gift_vouchers"))

        submissions = data.get("content", [])
        new_count = 0

        for sub in submissions:
            submission_id = sub.get("id")
            created_at = sub.get("created_at")

            # Skip duplicates
            existing = GiftVoucherSubmission.query.filter_by(submission_id=submission_id).first()
            if existing:
                continue

            parsed = parse_gift_voucher_payload(sub)

            entry = GiftVoucherSubmission(
                submission_id=submission_id,
                created_at=datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S"),

                purchaser_name=parsed["purchaser_name"],
                recipient_name=parsed["recipient_name"],
                voucher_number=parsed["voucher_number"],
                amount_payable=parsed["amount_payable"],

                ignored=False,
                processed=False
            )

            db.session.add(entry)
            new_count += 1

        db.session.commit()

        flash(f"Imported {new_count} new gift vouchers.", "success")
        return redirect(url_for("gift_vouchers"))





    @app.route("/health")
    def health():
        return "OK", 200


    # ---------------------------------------------------------
    # HELPERS (ALL INSIDE create_app)
    # ---------------------------------------------------------

    def send_sms_clicksend(to_number, message, sender_number):
        configuration = clicksend_client.Configuration()
        configuration.username = app.config['CLICKSEND_USERNAME']
        configuration.password = app.config['CLICKSEND_API_KEY']

        api_instance = clicksend_client.SMSApi(
            clicksend_client.ApiClient(configuration)
        )

        sms_message = SmsMessage(
            source="python",
            body=message,
            to=to_number,
            _from=sender_number
        )

        sms_messages = clicksend_client.SmsMessageCollection(
            messages=[sms_message]
        )

        try:
            response = api_instance.sms_send_post(sms_messages)
            print("ClickSend API response:", response)

            try:
                http_code = int(getattr(response, "http_code", 0))
            except:
                http_code = 0

            if http_code == 200:
                return True

            print("ClickSend returned non-200:", response)
            return False

        except ApiException as e:
            print("ClickSend SMS failed:", e)
            return False


    def load_wedding_unsubscribes():
        base = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(base, "unsubscribe", "wedding_unsubscribe.txt")

        if not os.path.exists(path):
            return set()

        with open(path, "r") as f:
            return {line.strip() for line in f if line.strip()}

    def generate_invite_token(lesson_date_str, lesson_time_str):
        """
        Generates a unique invite token. If a collision occurs,
        regenerates a new random suffix until unique.
        """
        import secrets
        import string

        date_comp = (lesson_date_str or "").replace("-", "")
        time_comp = (lesson_time_str or "").replace(":", "")[:4]

        while True:
            rand = ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
            token = f"TR-{date_comp}-{time_comp}-{rand}"

            # Check DB for existing token
            exists = db.session.query(
                db.session.query(LessonInvite.id).filter_by(invite_token=token).exists()
            ).scalar()

            if not exists:
                return token
            # else: collision → loop again

    def teacher_times_map():
        rows = get_static_teacher_time()   # <— cached rows instead of DB query

        out = {}
        for r in rows:
            teacher_key = r.teacher_key
            weekday = str(r.weekday)
            time_val = r.time

            if teacher_key not in out:
                out[teacher_key] = {}
            if weekday not in out[teacher_key]:
                out[teacher_key][weekday] = []

            out[teacher_key][weekday].append(time_val)

        return out

    def match_submission_to_invite(sub):
        payload = json.loads(sub.raw_payload)
        riders = parse_jotform_payload(payload, forced_submission_id=sub.id)
        if not riders:
            return None

        token = (
            riders[0].get("invite_token") or
            riders[0].get("i_t") or
            riders[0].get("token")
        )

        if not token:
            return None

        invite = db.session.query(LessonInvite).filter(
                LessonInvite.token == token,
                LessonInvite.status != "completed"
        ).first()

        return invite

    def get_or_create_client_from_rider(rider, invite_mobile, submission_id):
        name = normalise_full_name(rider["name"])

        # Match by NAME ONLY
        client = db.session.query(Client).filter(
            func.lower(Client.full_name) == name.lower()
        ).first()

        if client:
            return client

        # Create new client
        client = Client(
            full_name=name,
            mobile=invite_mobile or "",
            email_primary=rider.get("email") or "",
            guardian_name=rider.get("guardian") or "",
            jotform_submission_id=submission_id
        )

        db.session.add(client)
        return client


    def create_lesson_from_invite(invite, primary_client):
        clean_time = invite.time_frame
        start, end = clean_time.split(" - ")

        lesson = Lesson(
            client_id=primary_client.client_id,
            lesson_date=invite.lesson_date,
            start_time=start,
            end_time=end,
            lesson_type=invite.lesson_type,
            group_priv=invite.group_priv,
            attendance="Pending",
            payment=None
        )

        db.session.add(lesson)
        db.session.commit()
        return lesson

    def attach_rider_to_lesson(lesson, invite, client, rider):
        # Attach rider-specific notes to the existing lesson
        note = rider.get("notes")
        if note:
            existing = lesson.notes or ""
            lesson.notes = existing + f"\n{client.full_name}: {note}"


    # Utilities
    def normalize_name_for_lookup(name):
        return re.sub(r'[^a-z]', '', (name or '').lower())

    def norm(s):
        return re.sub(r'\s+', '', (s or '').strip().lower())

    def norm_timerange_key(timerange):
        return re.sub(r'[\s:-]+', '', (timerange or '').strip())

    def to_proper_case(s):
        return ' '.join(word.capitalize() for word in (s or '').split())

    def parse_money(v):
        if v is None:
            return None
        s = str(v).strip()
        if s == '':
            return None
        s = re.sub(r'[\$,]', '', s)
        try:
            return float(s)
        except Exception:
            return None

    def parse_selected_date():
        raw = request.values.get('date') or request.values.get('selected_date') or ''
        try:
            return datetime.strptime(raw, '%Y-%m-%d').date(), raw
        except Exception:
            return None, raw

    # Register helpers with Jinja
    app.jinja_env.globals.update(
        parse_money=parse_money,
        to_proper_case=to_proper_case,
        norm=norm,
        norm_timerange_key=norm_timerange_key,
        normalize_name_for_lookup=normalize_name_for_lookup
    )

    def finalize_invite_and_submission(invite, sub, lesson):
        invite.lesson_id = lesson.lesson_id
        invite.status = "completed"
        sub.processed = True
        db.session.commit()


    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("user", "").strip()
            password = request.form.get("password", "").strip()

            user = Users.query.filter_by(username=username, active=True).first()
            if user and check_password_hash(user.password_hash, password):
                session["user_id"] = user.user_id
                session["username"] = user.username
                return redirect(url_for("dashboard"))

            return render_template("login.html", error="Invalid username or password")

        return render_template("login.html")

    @app.route("/superlookup", methods=["GET", "POST"])
    def superlookup():
        print(">>> SUPERLOOKUP ROUTE HIT")
        q = request.form.get("q", "").strip()
        results = []

        if q:
            q_lower = q.lower()
            matches = Client.query.filter(
                db.or_(
                    func.coalesce(Client.full_name, '').ilike(f"%{q_lower}%"),
                    func.coalesce(Client.guardian_name, '').ilike(f"%{q_lower}%"),
                    func.coalesce(Client.email_primary, '').ilike(f"%{q_lower}%"),
                    func.coalesce(Client.mobile, '').ilike(f"%{q_lower}%")
                )
            ).all()

            for c in matches:
                print("RAW:", c.full_name)
                print("HIGHLIGHTED:", highlight(c.full_name, q))

                results.append({
                    "name": highlight(c.full_name, q),
                    "mobile": highlight(c.mobile, q),
                    "guardian": highlight(c.guardian_name, q),
                    "email": highlight(c.email_primary, q)
                })

        return render_template("superlookup.html", q=q, results=results)

    @app.route("/logout")
    def logout():
        session.clear()
        return redirect(url_for("login"))

    @app.route("/admin_reset_password/<username>/<newpassword>")
    def admin_reset_password(username, newpassword):
        user = Users.query.filter_by(username=username).first()
        if not user:
            return f"User '{username}' does not exist"

        user.password_hash = generate_password_hash(newpassword)
        db.session.commit()
        return f"Password for '{username}' reset successfully"



    @app.route("/admin_create_user/<username>/<password>/<role>")
    @login_required
    @role_required("admin", "management")
    def admin_create_user(username, password, role):
        existing = Users.query.filter_by(username=username).first()
        if existing:
            return f"User '{username}' already exists"

        pw_hash = generate_password_hash(password)
        user = Users(username=username, password_hash=pw_hash, active=True, role=role)
        db.session.add(user)
        db.session.commit()

        return f"User '{username}' created with role '{role}'"


    @app.route("/gift_voucher_edit/<int:id>", methods=["GET", "POST"])
    def gift_voucher_edit(id):
        from models import GiftVoucherSubmission
        from datetime import datetime

        v = GiftVoucherSubmission.query.get_or_404(id)

        if request.method == "POST":
            # Clean purchaser + recipient names
            purchaser = smart_proper_name(request.form.get("purchaser_name", v.purchaser_name))
            recipient = smart_proper_name(request.form.get("recipient_name", v.recipient_name))

            # Clean voucher number (trim only)
            voucher_no = request.form.get("voucher_number", v.voucher_number).strip()

            # Clean amount (trim only)
            amount = request.form.get("amount_payable", v.amount_payable).strip()

            # Save cleaned values
            v.purchaser_name = purchaser
            v.recipient_name = recipient
            v.voucher_number = voucher_no
            v.amount_payable = amount

            db.session.commit()
            flash("Voucher updated successfully.", "success")
            return redirect(url_for("gift_vouchers"))

        return render_template("gift_voucher_edit.html", v=v)

    @app.route('/pull_courses_from_jotform')
    def pull_courses_from_jotform():
        import requests
        import datetime
        import json
        from flask import request

        API_KEY = app.config['JOTFORM_API_KEY']
        FORM_ID = "212936006493860"

        url = f"https://api.jotform.com/form/{FORM_ID}/submissions?apiKey={API_KEY}"
        r = requests.get(url).json()

        if "content" not in r:
            return "JotForm API error: no content"

        pulled = 0

        for sub in r["content"]:
            # ⭐ ALWAYS normalise JotForm ID to string
            sub_id = str(sub.get("id")).strip()

            # ---------------------------------------------------------
            # DEDUPE CHECK — skip if already in DB
            # ---------------------------------------------------------
            existing = CourseFormSubmission.query.filter_by(jotform_id=sub_id).first()
            if existing:
                continue

            answers = sub.get("answers", {})

            # Map unique names → answers
            mapped = {}
            for qid, data in answers.items():
                uname = data.get("name")
                if uname:
                    mapped[uname] = data.get("answer")

            # Rider name
            rider_first = mapped.get("riderName", {}).get("first", "")
            rider_last  = mapped.get("riderName", {}).get("last", "")
            raw_name = f"{rider_first} {rider_last}"
            rider_full = clean_name(raw_name)

            # Course + FT/W
            courseno = mapped.get("courseNo", "")
            ftor = mapped.get("ft", "")

            # Horses
            horse_1 = mapped.get("horse_1", "")
            horse_2 = mapped.get("horse_2", "")
            horse_3 = mapped.get("horse_3", "")

            # Notes
            notes = mapped.get("anythingWe", "")

            # Term info
            term_year = mapped.get("year", None)
            term_number = mapped.get("term", None)

            # Skip empty junk submissions
            if not rider_full and not courseno:
                continue

            entry = CourseFormSubmission(
                jotform_id=sub_id,   # ⭐ critical for dedupe
                rider_name=rider_full,
                courseno=courseno,
                ftor=ftor,
                horse_1=horse_1,
                horse_2=horse_2,
                horse_3=horse_3,
                notes=notes,
                term_year=term_year,
                term_number=term_number,
                submitted_at=datetime.datetime.utcnow()
            )

            db.session.add(entry)
            pulled += 1

        db.session.commit()

        # ---------------------------------------------------------
        # PRESERVE FILTER (year + term) AFTER PULL
        # ---------------------------------------------------------
        selected_year = request.args.get("year", datetime.datetime.now().year)
        selected_term = request.args.get("term", 1)

        rows = CourseFormSubmission.query.filter_by(
            term_year=int(selected_year),
            term_number=int(selected_term)
        ).order_by(CourseFormSubmission.id.desc()).all()

        years = sorted({r.term_year for r in CourseFormSubmission.query.all() if r.term_year})

        return render_template(
            'course_form_results.html',
            rows=rows,
            years=years,
            selected_year=int(selected_year),
            selected_term=int(selected_term)
        )


    @app.route('/update_course_submission/<int:id>', methods=['POST'])
    def update_course_submission(id):
        sub = CourseFormSubmission.query.get(id)
        if not sub:
            return "Submission not found"

        new_course = request.form.get('courseno')
        sub.courseno = new_course

        db.session.commit()

        return redirect(url_for('course_form_results', 
                                year=sub.term_year, 
                                term=sub.term_number))


    @app.route('/edit_course_submission/<int:id>')
    def edit_course_submission(id):
        sub = CourseFormSubmission.query.get(id)
        if not sub:
            return "Submission not found"

        # Day ordering for correct sorting
        day_order = {
            "Sunday": 1,
            "Monday": 2,
            "Tuesday": 3,
            "Wednesday": 4,
            "Thursday": 5,
            "Friday": 6,
            "Saturday": 7
        }

        # Fetch all courses
        courses = CourseReference.query.all()

        # Sort manually in Python
        def sort_key(c):
            # Extract start time from timerange "HH:MM - HH:MM"
            try:
                start = c.timerange.split(" - ")[0]
                h, m = map(int, start.split(":"))
                minutes = h * 60 + m
            except:
                minutes = 9999

            return (day_order.get(c.day_of_week, 99), minutes)

        courses = sorted(courses, key=sort_key)

        return render_template(
            'edit_course_submission.html',
            sub=sub,
            courses=courses
        )

    @app.route('/delete_course_submission/<int:id>')
    def delete_course_submission(id):
        sub = CourseFormSubmission.query.get(id)
        if not sub:
            return "Submission not found"

        db.session.delete(sub)
        db.session.commit()

        return redirect(url_for('course_form_results'))


    @app.route('/course_form_results')
    def course_form_results():
        import datetime

        # -----------------------------
        # 1. Determine available years
        # -----------------------------
        years = db.session.query(CourseFormSubmission.term_year)\
                          .distinct()\
                          .order_by(CourseFormSubmission.term_year.desc())\
                          .all()
        years = [y[0] for y in years if y[0] is not None]

        # If no data yet, default to current year
        current_year = datetime.datetime.now().year
        if not years:
            years = [current_year]

        # -----------------------------
        # 2. Read filters from GET
        # -----------------------------
        selected_year = request.args.get('year', type=int)
        selected_term = request.args.get('term', type=int)

        # Default year
        if not selected_year:
            selected_year = years[0]

        # Default term (1–4)
        if not selected_term:
            month = datetime.datetime.now().month
            selected_term = ((month - 1) // 3) + 1  # simple quarter logic

        # -----------------------------
        # 3. Query filtered submissions
        # -----------------------------
        rows = CourseFormSubmission.query\
            .filter(CourseFormSubmission.term_year == selected_year)\
            .filter(CourseFormSubmission.term_number == selected_term)\
            .order_by(CourseFormSubmission.id.desc())\
            .all()

        # -----------------------------
        # 4. Render template
        # -----------------------------
        return render_template(
            'course_form_results.html',
            rows=rows,
            years=years,
            selected_year=selected_year,
            selected_term=selected_term
        )




    @app.route("/daily-summary", methods=["GET", "POST"])
    def daily_summary():
        # Determine selected FY
        selected_fy = request.values.get("fy")

        # Default FY if none selected
        if not selected_fy:
            today = date.today()
            if today.month >= 7:
                selected_fy = f"{today.year}-{today.year+1}"
            else:
                selected_fy = f"{today.year-1}-{today.year}"

        # Build FY list starting from 2022
        fy_years = list(range(2022, date.today().year + 2))

        # Handle SAVE (POST)
        if request.method == "POST":
            for key, value in request.form.items():
                if key.startswith("field1_"):
                    date_str = key.replace("field1_", "")
                    d = date.fromisoformat(date_str)

                    row = DailyEvent.query.filter_by(date=d, fy=selected_fy).first()
                    if not row:
                        row = DailyEvent(date=d, fy=selected_fy)
                        db.session.add(row)

                    row.field1 = value or None

                if key.startswith("field2_"):
                    date_str = key.replace("field2_", "")
                    d = date.fromisoformat(date_str)

                    row = DailyEvent.query.filter_by(date=d, fy=selected_fy).first()
                    if not row:
                        row = DailyEvent(date=d, fy=selected_fy)
                        db.session.add(row)

                    row.field2 = value or None

            db.session.commit()
            return redirect(url_for("daily_summary", fy=selected_fy))

        # GET → Build daily summary data
        daily_data = build_daily_summary(selected_fy)

        return render_template(
            "daily_summary.html",
            selected_fy=selected_fy,
            fy_years=fy_years,
            daily_data=daily_data
        )


    @app.route('/approve_course_submission/<int:id>')
    def approve_course_submission(id):
        import datetime

        sub = CourseFormSubmission.query.get(id)
        if not sub:
            return "Submission not found"

        # 1. Update status
        sub.status = "approved"

        # 2. Insert into real enrolment table
        enrol = CourseEnrolment(
            rider_name=sub.rider_name,
            course_code=sub.courseno,
            term_year=sub.term_year,
            term_number=sub.term_number,
            created_at=datetime.datetime.utcnow()
        )

        db.session.add(enrol)
        db.session.commit()

        return redirect(url_for('course_form_results'))



    @app.route('/courses_menu')
    def courses_menu():
        return render_template('courses_menu.html')

    @app.route("/admin_delete_user/<username>")
    @login_required
    @role_required("admin", "management")
    def admin_delete_user(username):
        user = Users.query.filter_by(username=username).first()
        if not user:
            return f"User '{username}' does not exist"

        db.session.delete(user)
        db.session.commit()
        return f"User '{username}' deleted"

    @app.route("/print/lessons/<date>")
    def print_lessons(date):
        from datetime import datetime
        from collections import defaultdict

        # -----------------------------
        # PARSE DATE
        # -----------------------------
        lesson_date = datetime.strptime(date, "%Y-%m-%d").date()
        dow = lesson_date.strftime("%A")
        pretty_date = lesson_date.strftime("%d %B %Y")

        # -----------------------------
        # FETCH ALL LESSON ROWS
        # -----------------------------
        lessons = (
            Lesson.query
            .filter_by(lesson_date=lesson_date)
            .order_by(Lesson.time_frame)
            .all()
        )

        # -----------------------------
        # FILTER OUT NON-LESSON TYPES
        # -----------------------------
        SAFE_TYPES = ("Payment", "Voucher CR", "Camp")
        lessons = [l for l in lessons if l.lesson_type not in SAFE_TYPES]

        # -----------------------------
        # BUILD CLEAN DATA DICTIONARIES
        # -----------------------------
        data = []
        for l in lessons:
            client = Client.query.filter_by(full_name=l.client).first()

            # SAFE DISCLAIMER
            raw = str(getattr(client, "disclaimer", "")).strip()
            disclaimer_val = int(raw) if raw.isdigit() else 0

            # SAFE BALANCE
            raw_balance = str(getattr(l, "balance", "")).strip()
            if raw_balance.lower() == "none" or raw_balance == "":
                balance = 0.0
            else:
                try:
                    balance = float(raw_balance)
                except:
                    balance = 0.0

            # CLEAN TIME FRAME
            tf = (l.time_frame or "").strip().replace("–", "-").replace("—", "-")

            # SKIP ANYTHING WITH NO TIME
            if not tf or "-" not in tf:
                continue

            # SAFE PAYMENT
            raw_payment = getattr(l, "payment", 0)
            try:
                payment = float(raw_payment) if raw_payment not in (None, "", "None") else 0.0
            except:
                payment = 0.0

            # SAFE PRICE
            raw_price = getattr(l, "price_pl", 0)
            try:
                price = float(raw_price) if raw_price not in (None, "", "None") else 0.0
            except:
                price = 0.0

            data.append({
                "time_frame": tf,
                "lesson_type": (l.lesson_type or "").strip(),
                "group_priv": (l.group_priv or "").strip().upper(),

                "client_name": l.client or "",
                "freq": getattr(l, "freq", "") or "",
                "att": getattr(l, "attendance", False),

                "payment": payment,
                "price": price,
                "balance": balance,

                "age": getattr(client, "age", "") or "",
                "guardian": getattr(client, "guardian_name", "") or "",
                "mobile": getattr(client, "mobile", "") or "",
                "weight": getattr(client, "weight_kg", "") or "",
                "height": getattr(client, "height_cm", "") or "",
                "notes": getattr(client, "notes", "") or "",
                "disclaimer": disclaimer_val,

                "horse": getattr(l, "horse", "") or "",
            })

        # -----------------------------
        # PRIVATE LESSON INDEXING
        # -----------------------------
        private_counter = defaultdict(int)
        for d in data:
            if d["group_priv"] == "P":
                key = (d["time_frame"], d["lesson_type"])
                private_counter[key] += 1
                d["priv_index"] = private_counter[key]
            else:
                d["priv_index"] = 0

        # -----------------------------
        # SAFE TIME PARSER
        # -----------------------------
        def parse_start(tf):
            try:
                start = tf.split("-")[0].strip()
                h, m = start.split(":")
                return int(h) * 60 + int(m)
            except:
                return 99999

        # -----------------------------
        # SPLIT ARENA VS OTHERS
        # -----------------------------
        arena = [d for d in data if d["lesson_type"].lower().startswith("arena")]
        others = [d for d in data if not d["lesson_type"].lower().startswith("arena")]

        def safe_sort_key(tf):
                if not tf:  # no time (e.g. Camp)
                        return 999999  # push to bottom, or -1 if you want top
                return parse_start(tf)


        arena.sort(key=lambda x: safe_sort_key(x["time_frame"]))
        others.sort(key=lambda x: safe_sort_key(x["time_frame"]))

        # -----------------------------
        # GROUP BLOCKS SAFELY
        # -----------------------------
        def group_blocks(rows):
                blocks = defaultdict(list)
                for r in rows:
                        key = (r["time_frame"], r["lesson_type"], r["group_priv"], r["priv_index"])
                        blocks[key].append(r)

                grouped = []
                for (time, ltype, gp, priv_i), riders in sorted(
                        blocks.items(),
                        key=lambda k: parse_start(k[0][0])
                ):
                        grouped.append({
                                "time": time,
                                "lesson_type": ltype,
                                "group_priv": gp,
                                "riders": sorted(riders, key=lambda x: x["client_name"].lower())
                        })
                return grouped

        arena_blocks = group_blocks(arena)
        other_blocks = group_blocks(others)

        # -----------------------------
        # RENDER TEMPLATE
        # -----------------------------
        return render_template(
            "print_lessons.html",
            dow=dow,
            pretty_date=pretty_date,
            arena_blocks=arena_blocks,
            other_blocks=other_blocks,
        )


    @app.route("/print/horses/<date>")
    def print_horses(date):
        from datetime import datetime
        from collections import defaultdict

        # -----------------------------
        # PARSE DATE
        # -----------------------------
        lesson_date = datetime.strptime(date, "%Y-%m-%d").date()
        pretty_date = lesson_date.strftime("%d %B %Y")

        # -----------------------------
        # FETCH ALL LESSON ROWS
        # -----------------------------
        lessons = (
            Lesson.query
            .filter_by(lesson_date=lesson_date)
            .order_by(Lesson.time_frame)
            .all()
        )

        # -----------------------------
        # FILTER OUT NON-LESSON TYPES
        # -----------------------------
        SAFE_TYPES = ("Payment", "Voucher CR", "Camp")
        lessons = [l for l in lessons if l.lesson_type not in SAFE_TYPES]

        # -----------------------------
        # BUILD CLEAN DATA
        # -----------------------------
        data = []
        for l in lessons:
            client = Client.query.filter_by(full_name=l.client).first()

            tf = (l.time_frame or "").strip().replace("–", "-").replace("—", "-")
            if not tf or "-" not in tf:
                continue

            data.append({
                "time_frame": tf,
                "lesson_type": (l.lesson_type or "").strip(),
                "group_priv": (l.group_priv or "").strip().upper(),
                "client_name": l.client or "",
                "att": getattr(l, "attendance", False),
                "teacher": getattr(l, "teacher", "") or "",
                "horse": getattr(l, "horse", "") or "",
            })

        # -----------------------------
        # SAFE TIME PARSER
        # -----------------------------
        def parse_start(tf):
            try:
                start = tf.split("-")[0].strip()
                h, m = start.split(":")
                return int(h) * 60 + int(m)
            except:
                return 99999

        # -----------------------------
        # GROUP BY TIME + LESSON TYPE + PRIV
        # -----------------------------
        blocks = defaultdict(list)
        for d in data:
            key = (d["time_frame"], d["lesson_type"], d["group_priv"])
            blocks[key].append(d)

        grouped = []
        for (time, ltype, gp), riders in sorted(
            blocks.items(),
            key=lambda k: parse_start(k[0][0])
        ):
            grouped.append({
                "time": time,
                "lesson_type": ltype,
                "group_priv": gp,
                "riders": sorted(riders, key=lambda x: x["horse"].lower())
            })

        # -----------------------------
        # RENDER TEMPLATE
        # -----------------------------
        return render_template(
            "horse_print.html",
            date=pretty_date,
            grouped=grouped
        )



    @app.post("/recalculate_all")
    def recalculate_all():
        try:
            recalc_all_lessons()
            return jsonify(success=True, message="All lesson balances recalculated.")
        except Exception as e:
            return jsonify(success=False, error=str(e))


    @app.route('/terms')
    def terms():
        all_terms = Term.query.order_by(Term.year, Term.term_number).all()
        return render_template('terms.html', terms=all_terms)

    @app.route('/terms/add', methods=['POST'])
    def add_term():
        year = request.form['year']
        term_number = request.form['term_number']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        week_pattern = request.form['week_pattern']

        t = Term(
            year=year,
            term_number=term_number,
            start_date=start_date,
            end_date=end_date,
            week_pattern=week_pattern,
            weeks=10
        )
        db.session.add(t)
        db.session.commit()

        return redirect(url_for('terms'))

    @app.route('/terms/activate/<int:id>')
    def activate_term(id):
        Term.query.update({Term.active: False})
        t = Term.query.get(id)
        t.active = True
        db.session.commit()
        return redirect(url_for('terms'))

    @app.route('/terms/delete/<int:id>')
    def delete_term(id):
        t = Term.query.get(id)
        db.session.delete(t)
        db.session.commit()
        return redirect(url_for('terms'))




    @app.post("/update_lesson_field")
    def update_lesson_field():
        data = request.get_json()
        lesson_id = data.get("lesson_id")
        field = data.get("field")
        value = data.get("value")

        allowed = {
            "attendance",
            "carry_fwd",
            "payment",
            "price_pl",
            "adjust",
            "lesson_notes"      # ⭐ NEW
        }
        if field not in allowed:
            return jsonify(success=False, error="Invalid field")

        lesson = Lesson.query.get(lesson_id)
        if not lesson:
            return jsonify(success=False, error=f"Lesson {lesson_id} not found")

        numeric_fields = {"carry_fwd", "payment", "price_pl", "adjust"}
        if field in numeric_fields:
            try:
                value = float(value) if value not in (None, "", "None") else 0
            except:
                return jsonify(success=False, error="Invalid number")

        setattr(lesson, field, value)

        # no recalculation here
        db.session.commit()

        return jsonify(success=True)


    # ---------------------------------------------------------
    # ROUTES (ALL INSIDE create_app)
    # ---------------------------------------------------------

    @app.post("/update_client_field")
    def update_client_field():
        data = request.get_json()
        client_id = data.get("client_id")
        field = data.get("field")
        value = data.get("value")

        # Allowed editable fields (EXCLUDES client name + disclaimer)
        allowed = {
            "guardian_name",
            "mobile",
            "age",
            "weight_kg",
            "height_cm",
            "email_primary",
            "email_secondary",
            "guardian2_name",
            "mobile2",
            "email_guardian2",
            "ndis_number",
            "ndis_code",
            "notes",
            "notes2"
        }

        if field not in allowed:
            return jsonify(success=False, error="Invalid field")

        client = Client.query.get(client_id)
        if not client:
            return jsonify(success=False, error="Client not found")

        # Convert numeric fields safely
        numeric_fields = {"age", "weight_kg", "height_cm"}
        if field in numeric_fields:
            try:
                value = int(value) if value else None
            except:
                return jsonify(success=False, error="Invalid number")

        setattr(client, field, value)
        db.session.commit()

        return jsonify(success=True)





    @app.route('/notifications/process/<int:sub_id>')
    def process_submission_route(sub_id):
        sub = db.session.query(IncomingSubmission).get_or_404(sub_id)
        result = process_submission(sub)
        flash(result, "success")
        return redirect(url_for('notifications'))

    @app.route("/summaries", methods=["GET"])
    def summaries_page():
        fy_list = generate_financial_years(start_year=2022)
        selected_fy = request.args.get("fy")

        sundays = None
        weekly_data = None

        if selected_fy:
            sundays = get_sundays_for_financial_year(selected_fy)
            weekly_data = build_weekly_summary(sundays, selected_fy)

        return render_template(
            "summaries.html",
            fy_list=fy_list,
            selected_fy=selected_fy,
            sundays=sundays,
            weekly_data=weekly_data
        )

    @app.route('/invite_submissions')
    def invite_submissions():
        INVITE_FORM_ID = "253599154628066"

        subs = (
            db.session.query(IncomingSubmission)
            .filter(IncomingSubmission.form_id == INVITE_FORM_ID)
            .order_by(IncomingSubmission.received_at.desc())
            .all()
        )

        return render_template('invite_submissions.html', subs=subs)

    @app.route('/pending_lessons')
    def pending_lessons():
        # Get all invites that are not completed
        invites = (
            db.session.query(LessonInvite)
            .filter(LessonInvite.status != "completed")
            .order_by(LessonInvite.id.desc())
            .all()
        )

        # Get all unprocessed submissions
        submissions = (
            db.session.query(IncomingSubmission)
            .filter_by(processed=False)
            .all()
        )

        # Build lookup: invite_token → submission model
        submission_lookup = {}
        for sub in submissions:

            # Ignore submissions from the disclaimer form
            # Your SMS invite form ID = 253599154628066
            if sub.form_id != "253599154628066":
                continue

            try:
                riders = parse_jotform_payload(sub.raw_payload, forced_submission_id=sub.id)
                for rider in riders:
                    token = rider.get("invite_token")
                    if token:
                        submission_lookup[token] = sub
            except Exception:
                continue

        # Build pending list for template
        pending = []
        for inv in invites:
            sub = submission_lookup.get(inv.token)
            # If we found a matching submission, update invite status
            if sub and inv.status in ["waiting", "form_received"]:
                inv.status = "process"
                db.session.commit()

            parsed = None
            if sub:
                try:
                    riders = parse_jotform_payload(sub.raw_payload, forced_submission_id=sub.id)
                    parsed = riders[0] if riders else None
                except Exception:
                    parsed = None

            pending.append({
                "invite": inv,
                "submission": parsed,   # template‑safe dict
                "raw": sub              # optional model reference
            })

        return render_template("pending_lessons.html", pending=pending)

    @app.route("/equestrian_sms")
    def equestrian_sms():
        return render_template("equestrian_sms.html")

    @app.route("/equestrian/sms/send", methods=["POST"])
    def equestrian_sms_send():
        data = request.get_json()

        message = data.get("message", "")
        numbers = data.get("numbers", [])

        sender = app.config.get("EQUESTRIAN_SENDER", None)

        # No unsubscribe list for equestrian (yet)
        filtered = numbers[:]   # copy all numbers
        skipped = []

        sent_numbers = []
        failed_numbers = []

        for num in filtered:
            ok = send_sms_clicksend(num, message, sender)
            if ok:
                sent_numbers.append(num)
            else:
                failed_numbers.append(num)

        return jsonify({
            "ok": True,
            "sent": len(sent_numbers),
            "skipped": len(skipped),
            "failed": len(failed_numbers),
            "sent_numbers": sent_numbers,
            "skipped_numbers": skipped,
            "failed_numbers": failed_numbers,
            "status": (
                f"Sent {len(sent_numbers)} messages. "
                f"Failed {len(failed_numbers)}."
            )
        })


    @app.route("/equestrian_bulk_sms")
    def equestrian_bulk_sms():
        return render_template("equestrian_bulk_sms.html")

    @app.route("/equestrian/bulk_sms/send", methods=["POST"])
    def equestrian_bulk_sms_send():
        data = request.get_json()

        template = data.get("template", "").strip()
        rows = data.get("rows", [])
        fields = data.get("fields", [])

        sender = app.config.get("EQUESTRIAN_SENDER")

        sent_numbers = []
        failed_numbers = []
        skipped_numbers = []

        if not template:
            return jsonify({
                "ok": False,
                "status": "Template is empty.",
                "sent": 0,
                "failed": 0,
                "skipped": 0,
                "sent_numbers": [],
                "failed_numbers": [],
                "skipped_numbers": []
            })

        if not rows:
            return jsonify({
                "ok": False,
                "status": "No rows received from client.",
                "sent": 0,
                "failed": 0,
                "skipped": 0,
                "sent_numbers": [],
                "failed_numbers": [],
                "skipped_numbers": []
            })

        if "mobile" not in fields:
            return jsonify({
                "ok": False,
                "status": "CSV does not contain 'mobile' column.",
                "sent": 0,
                "failed": 0,
                "skipped": len(rows),
                "sent_numbers": [],
                "failed_numbers": [],
                "skipped_numbers": [r.get("mobile", "") for r in rows]
            })

        for row in rows:
            mobile = (row.get("mobile") or "").strip()
            if not mobile:
                skipped_numbers.append(mobile)
                continue

            personalised = template
            for field in fields:
                placeholder = f"{{{field}}}"
                if placeholder in personalised:
                    personalised = personalised.replace(placeholder, row.get(field, "") or "")

            ok = send_sms_clicksend(mobile, personalised, sender)
            if ok:
                sent_numbers.append(mobile)
            else:
                failed_numbers.append(mobile)

        status = (
            f"Sent {len(sent_numbers)} messages. "
            f"Failed {len(failed_numbers)}. "
            f"Skipped {len(skipped_numbers)}."
        )

        return jsonify({
            "ok": True,
            "status": status,
            "sent": len(sent_numbers),
            "failed": len(failed_numbers),
            "skipped": len(skipped_numbers),
            "sent_numbers": sent_numbers,
            "failed_numbers": failed_numbers,
            "skipped_numbers": skipped_numbers
        })

    def archive_wedding_sms_numbers(numbers):
        from datetime import datetime

        pending_path = os.path.join(WEDDING_SMS_ROOT, "wedding_sms_pending.txt")
        archive_path = os.path.join(
            WEDDING_SMS_ROOT,
            f"wedding_sms_sent_{datetime.now().strftime('%Y-%m-%d')}.txt"
        )


        # Append all numbers to archive
        with open(archive_path, "a", encoding="utf-8") as f:
            for num in numbers:
                f.write(num + "\n")

        # Clear pending file
        with open(pending_path, "w", encoding="utf-8") as f:
            f.write("")

    @app.get("/wedding/sms/pending")
    def wedding_sms_pending():
        path = os.path.join(WEDDING_SMS_ROOT, "wedding_sms_pending.txt")


        try:
            with open(path, "r", encoding="utf-8") as f:
                numbers = [line.strip() for line in f.readlines() if line.strip()]
        except FileNotFoundError:
            numbers = []

        return {"numbers": numbers}


    @app.route('/wedding/sms/send', methods=['POST'])
    def wedding_sms_send():
        data = request.get_json()

        message = data.get("message", "")
        numbers = data.get("numbers", [])

        sender = app.config['WEDDING_SENDER']

        # Load unsubscribe list
        base = os.path.dirname(os.path.abspath(__file__))
        unsub_path = os.path.join(base, "unsubscribe", "wedding_unsubscribe.txt")

        unsub = set()
        if os.path.exists(unsub_path):
            with open(unsub_path, "r") as f:
                unsub = {line.strip() for line in f if line.strip()}

        # Filter numbers
        filtered = []
        skipped = []

        for num in numbers:
            if num in unsub:
                skipped.append(num)
            else:
                filtered.append(num)

        # Send SMS to filtered numbers
        sent_numbers = []
        failed_numbers = []

        for num in filtered:
            ok = send_sms_clicksend(num, message, sender)
            if ok:
                sent_numbers.append(num)
            else:
                failed_numbers.append(num)

        # ---------------------------------------------------------
        # ARCHIVE ALL NUMBERS (Option A)
        # ---------------------------------------------------------
        archive_wedding_sms_numbers(numbers)

        return jsonify({
            "ok": True,
            "sent": len(sent_numbers),
            "skipped": len(skipped),
            "failed": len(failed_numbers),
            "sent_numbers": sent_numbers,
            "skipped_numbers": skipped,
            "failed_numbers": failed_numbers,
            "status": (
                f"Sent {len(sent_numbers)} messages. "
                f"Skipped {len(skipped)} unsubscribed numbers. "
                f"Failed {len(failed_numbers)}."
            )
        })



    @app.post("/recalculate_client_cascade")
    def recalculate_client_cascade_route():
        client = request.form.get("client")

        try:
            recalc_client_cascade(client)
            flash(f"Cascading recalculation completed for {client}.", "success")
        except Exception as e:
            flash(f"Error during cascading recalculation: {e}", "danger")

        return redirect(url_for("client_view", client_filter=client), code=303)


    @app.route('/api/teacher_times.json')
    def api_all_teacher_times():
        return {"teacher_times": teacher_times_map()}



    @app.route('/')
    def index():
        return redirect(url_for("login"))

    @app.route('/dashboard')
    @login_required
    @role_required("admin", "management")
    def dashboard():
        return render_template("index.html")


    @app.post('/api/add_client')
    def api_add_client():
        data = request.json

        new_client = Client(
            full_name=data.get("full_name"),
            weight_kg=data.get("weight_kg"),
            height_cm=data.get("height_cm"),
            mobile=data.get("mobile"),
            email_primary=data.get("email_primary"),
            notes=data.get("notes"),
            age=data.get("age"),
        )

        db.session.add(new_client)
        db.session.commit()

        return {
            "status": "ok",
            "client_id": new_client.client_id
        }


    @app.get('/api/test')
    def api_test():
        print("TEST ROUTE HIT")
        return {"status": "ok"}


    @app.route("/pdf/<date>")
    def pdf_for_date(date):
        # 1. Render the HTML using the existing lessons_by_date logic
        with app.test_request_context(f"/lessons_by_date?date={date}"):
            html = lessons_by_date()

        # 2. Generate PDF using Chromium
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.set_content(html)
            pdf_bytes = page.pdf(format="A4", landscape=True)
            browser.close()

        # 3. Return PDF to browser
        resp = make_response(pdf_bytes)
        resp.headers["Content-Type"] = "application/pdf"
        resp.headers["Content-Disposition"] = f"inline; filename=lessons_{date}.pdf"
        return resp


    def build_lessons_context(selected_date, selected_date_str):
        """
        Takes a date + its string form and returns the full context
        needed by lessons_by_date.html.
        """
        grouped_lessons = {}
        horse_list = [to_proper_case(h.horse) for h in get_static_horses()]
        horse_schedule = defaultdict(list)
        client_horse_history = defaultdict(list)
        times = [t.timerange for t in db.session.query(Time).order_by(Time.timerange).all()]
        teacher_names = [t.teacher for t in get_static_teachers()]
        block_tag_lookup = {}
        invoice_clients = []
        if selected_date:
            lesson_rows = (
                db.session.query(Lesson)
                .filter(
                    Lesson.lesson_date == selected_date,
                    Lesson.lesson_type.notin_(["Payment", "Voucher CR"])
                )
                .order_by(Lesson.time_frame)
                .all()
            )
        else:
            lesson_rows = []

        # Load per-lesson T1–T5 overrides
        override_rows = (
            db.session.query(LessonTeacherTag)
            .filter_by(lesson_date=selected_date)
            .all()
        )

        lesson_tag_overrides = {
            row.lesson_id: {
                "T1": row.t1,
                "T2": row.t2,
                "T3": row.t3,
                "T4": row.t4,
                "T5": row.t5,
            }
            for row in override_rows
        }


        # Populate horse_schedule from lesson_rows
        for lesson in lesson_rows:
            horse_name = to_proper_case(lesson.horse)
            att = (lesson.attendance or '').strip().upper()
            time_frame = (lesson.time_frame or '').strip()

            if horse_name and time_frame and att != 'C':
                horse_schedule[horse_name].append(time_frame)

        # Build time lookup
        time_lookup = {norm(t.timerange): t for t in db.session.query(Time).all()}

        # Build client lookup
        client_lookup = {}
        clients = db.session.query(Client).all()

        for c in clients:
            full = c.full_name or ''
            key = normalize_name_for_lookup(full)
            client_lookup[key] = c
            client_lookup[(full or '').lower().replace(' ', '')] = c

        # Group lessons
        grouped = defaultdict(list)
        for lesson in lesson_rows:
            time_key = norm(lesson.time_frame)
            time_obj = time_lookup.get(time_key)
            client_key = normalize_name_for_lookup(lesson.client or '')
            client_obj = client_lookup.get(client_key)

            if not client_obj:
                alt_key = (lesson.client or '').lower().replace(' ', '')
                client_obj = client_lookup.get(alt_key)

            timerange_display = time_obj.timerange if time_obj else (lesson.time_frame or '—')
            group_key = (timerange_display, lesson.lesson_type or '', lesson.group_priv or '')
            grouped[group_key].append((lesson, client_obj))

        # Sorting helper
        def time_sort_key(timerange):
            try:
                return datetime.strptime(timerange.split('-')[0].strip(), '%H:%M').time()
            except Exception:
                return time.min

        sorted_keys = sorted(
            grouped.keys(),
            key=lambda k: (0 if (k[1] or '').lower() == 'arena' else 1, time_sort_key(k[0]))
        )

        grouped_lessons = {
            k: sorted(grouped[k], key=lambda pair: ((pair[1].full_name if pair[1] else pair[0].client) or '').lower())
            for k in sorted_keys
        }

        # Build block_tag_lookup from DB
        block_tag_lookup = {}
        rows = db.session.query(LessonBlockTag).filter_by(lesson_date=selected_date).all()
        for r in rows:
            key = norm_timerange_key(r.time_range)
            tags = [t.strip() for t in (r.teacher_tags or '').split(',') if t.strip()]
            block_tag_lookup[key] = tags

        # Apply fallback tags based on lesson type
        for (timerange, lesson_type, group_priv), lesson_group in grouped_lessons.items():
            block_key = norm_timerange_key(timerange)
            if block_key not in block_tag_lookup:
                if (lesson_type or '').lower() == 'arena':
                    block_tag_lookup[block_key] = ['T1']
                else:
                    block_tag_lookup[block_key] = ['T2']

        # Merge lesson-level overrides with block defaults
        merged_tags = {}

        for (timerange, lesson_type, group_priv), lesson_group in grouped_lessons.items():
            for lesson, client_obj in lesson_group:
                lid = lesson.lesson_id

                # 1. Lesson-level override exists → highest priority
                if lid in lesson_tag_overrides:
                    merged_tags[lid] = [
                        tag for tag, val in lesson_tag_overrides[lid].items() if val
                    ]
                    continue

                # 2. Block-level tags
                block_key = norm_timerange_key(timerange)
                if block_key in block_tag_lookup:
                    merged_tags[lid] = block_tag_lookup[block_key]
                    continue

                # 3. Fallback defaults (Arena → T1, others → T2)
                if (lesson.lesson_type or '').lower() == 'arena':
                    merged_tags[lid] = ['T1']
                else:
                    merged_tags[lid] = ['T2']


        # Build teacher_horse_usage
        teacher_horse_usage = {}
        teacher_rows = db.session.query(TeacherHorse).filter(
            TeacherHorse.date == selected_date
        ).all()

        for th in teacher_rows:
            bk = (th.block_key or "").strip()
            if len(bk) == 8 and bk.isdigit():
                start, end = bk[:4], bk[4:]
                time_slot = f"{start[:2]}:{start[2:]} - {end[:2]}:{end[2:]}"
            else:
                time_slot = bk or "??:??"

            for raw_h in (th.horse1, th.horse2):
                if raw_h:
                    name = to_proper_case(str(raw_h).strip())
                    teacher_horse_usage.setdefault(name, []).append(time_slot)

        # Deduplicate + sort
        for h, usage_times in teacher_horse_usage.items():
            teacher_horse_usage[h] = sorted(set(usage_times))

        # Build slot_map
        slot_rows = (
            db.session.query(TeacherSlot)
            .filter_by(lesson_date=selected_date)
            .all()
        )
        slot_map = {s.slot_number: s.teacher_name for s in slot_rows}


        return {
            "grouped_lessons": grouped_lessons,
            "horse_list": horse_list,
            "horse_schedule": horse_schedule,
            "client_horse_history": client_horse_history,
            "times": times,
            "teacher_names": teacher_names,
            "block_tag_lookup": block_tag_lookup,
            "invoice_clients": invoice_clients,
            "lesson_rows": lesson_rows,
            "time_lookup": time_lookup,
            "client_lookup": client_lookup,
            "clients": clients,
            "teacher_horse_usage": teacher_horse_usage,
            "slot_map": slot_map,
            "merged_tags": merged_tags,             
        }





    @app.route('/lessons_by_date', methods=['GET', 'POST'])
    def lessons_by_date():
        db.session.rollback()
        selected_date, selected_date_str = parse_selected_date()

        # If no date was provided, default to Brisbane "today"
        if not selected_date:
            selected_date = datetime.now(ZoneInfo("Australia/Brisbane")).date()
            selected_date_str = selected_date.strftime('%Y-%m-%d')

        # Brisbane weekday
        weekday_int = datetime.now(ZoneInfo("Australia/Brisbane")).weekday()

        # Required for date inputs
        selected_date_iso = selected_date.strftime("%Y-%m-%d")

        # DEBUG
        print("=== DEBUG lessons_by_date ENTRY ===")
        print("method:", request.method)
        print("request.values:", dict(request.values))
        raw = request.values.get('date') or request.values.get('selected_date') or ''
        print("raw date (string):", repr(raw))
        try:
            parsed = datetime.strptime(raw, '%Y-%m-%d').date() if raw else None
            print("parsed date:", parsed, "type:", type(parsed))
        except Exception as e:
            parsed = None
            print("parse error:", e)
        rows = db.session.query(Lesson).filter(Lesson.lesson_date == parsed).limit(5).all() if parsed else []
        print("rows found (count):", len(rows))
        for r in rows:
            print("  sample:", r.lesson_id, getattr(r, 'client', None), getattr(r, 'lesson_date', None))
        print("=== END DEBUG ===")

        # ⭐ BUILD CONTEXT (we will patch inside this function) ⭐
        ctx = build_lessons_context(selected_date, selected_date_str)

        # Load grid overrides
        overrides = TeacherGridOverride.query.filter_by(
            override_date=selected_date
        ).all()

        grid_overrides = {
            (o.time_label, o.teacher_index): o.state
            for o in overrides
        }

        # Load teacher blocks
        teacher_blocks = TeacherBlock.query.filter_by(date=selected_date_str).all()

        # Detach ORM objects
        db.session.expunge_all()

        # Normalize dates
        for block in teacher_blocks:
            if isinstance(block.date, str):
                try:
                    block.date = datetime.strptime(block.date, "%Y-%m-%d").date()
                except:
                    block.date = None

        teacher_blocks_json = [
            {
                "id": tb.id,
                "date": tb.date,
                "block_key": tb.block_key,
                "horse": tb.horse,
                "teacher_name": tb.teacher_name,
                "notes": tb.notes
            }
            for tb in teacher_blocks
        ]

        # Unpack context
        grouped_lessons = ctx["grouped_lessons"]
        horse_list = ctx["horse_list"]
        horse_schedule = ctx["horse_schedule"]
        client_horse_history = ctx["client_horse_history"]
        times = ctx["times"]
        teacher_names = ctx["teacher_names"]
        block_tag_lookup = ctx["block_tag_lookup"]
        invoice_clients = ctx["invoice_clients"]
        lesson_rows = ctx["lesson_rows"]
        time_lookup = ctx["time_lookup"]
        client_lookup = ctx["client_lookup"]
        clients = ctx["clients"]
        teacher_horse_usage = ctx["teacher_horse_usage"]
        slot_map = ctx["slot_map"]


        return render_template(
            'lessons_by_date.html',
            grouped_lessons=grouped_lessons,
            selected_date=selected_date_str,
            selected_date_iso=selected_date_iso,
            weekday_int=weekday_int,
            horse_list=horse_list,
            horse_schedule=horse_schedule,
            teacher_names=teacher_names,
            block_tag_lookup=block_tag_lookup,
            client_horse_history=client_horse_history,
            invoice_clients=invoice_clients,
            clients=clients,
            teacher_times=teacher_times_map(),
            teacher_horse_usage=teacher_horse_usage,
            times=times,
            get_static_teachers=get_static_teachers,
            client_lookup=client_lookup,
            slot_map=slot_map,
            merged_tags=ctx["merged_tags"],
            norm_timerange_key=norm_timerange_key,
            grid_overrides=grid_overrides,
            teacher_blocks=teacher_blocks_json,
        )


    def detect_conflicts(riders):
        conflicts = []
        for rider in riders:
            if rider.get("matches"):
                conflicts.append({
                    "rider": rider,
                    "matches": rider["matches"]
                })
        return conflicts


    def create_lesson_row(invite, client_name):
        """
        Creates one Lesson row for a single client using fields from lesson_invites.
        """
        clean_time = invite.time_frame

        lesson = Lesson(
            lesson_date=invite.lesson_date,
            time_frame=clean_time,
            lesson_type=invite.lesson_type,
            group_priv=invite.group_priv,
            price_pl=invite.cost_per_person,
            client=client_name,
            horse=""
        )

        db.session.add(lesson)
        db.session.flush()

        return lesson

    def soft_match_client_for_invite(rider_name: str):
        """
        Soft‑match a rider name to existing clients.
        Returns:
            ("exact", client)
            ("none", None)
            ("ambiguous", [clients])
        """
        if not rider_name:
            return ("none", None)

        name_norm = (rider_name or "").strip().lower()
        compact = name_norm.replace(" ", "").replace("-", "")

        # ---------------------------------------------------------
        # 1. Exact match (case‑insensitive)
        # ---------------------------------------------------------
        exact = db.session.query(Client).filter(
            func.lower(func.trim(Client.full_name)) == name_norm
        ).all()

        if len(exact) == 1:
            return ("exact", exact[0])
        if len(exact) > 1:
            return ("ambiguous", exact)

        # ---------------------------------------------------------
        # 2. Soft match using compact form (SQL LIKE)
        #    Avoid full-table scan by filtering in SQL
        # ---------------------------------------------------------
        like_pattern = f"%{compact}%"

        candidates = db.session.query(Client).filter(
            Client.full_name.isnot(None),
            func.replace(
                func.replace(func.lower(Client.full_name), " ", ""),
                "-", ""
            ).like(like_pattern)
        ).all()

        if not candidates:
            return ("none", None)
        if len(candidates) == 1:
            return ("exact", candidates[0])
        return ("ambiguous", candidates)

    def process_submission(sub):
        """
        Robust replacement for the old process_submission.
        """

        # ---------------------------------------------------------
        # TESTING MODE (set True while developing)
        # ---------------------------------------------------------
        TESTING_MODE = True

        # ---------------------------------------------------------
        # MATCH INVITE FIRST (we need invite.status for guardrail)
        # ---------------------------------------------------------
        invite = match_submission_to_invite(sub)
        if not invite:
            return "No matching invite"

        submission_id = str(sub.id)


        # ---------------------------------------------------------
        # PRODUCTION GUARDRAIL:
        # Only skip if the invite is already completed.
        # ---------------------------------------------------------
        if not TESTING_MODE:
            if invite.status == "completed":
                sub.processed = True
                db.session.commit()
                return f"Invite already completed for submission {submission_id}"
        # In TESTING_MODE, we allow full reprocessing.

        # ---------------------------------------------------------
        # PARSE RIDERS
        # ---------------------------------------------------------
        try:
            riders = parse_jotform_payload(
                sub.raw_payload,
                forced_submission_id=sub.id
            )
        except Exception as e:
            print("RIDER PARSE ERROR:", e)
            return f"Invalid payload: {e}"

        if not riders:
            return "No riders found"

        # ---------------------------------------------------------
        # CONFLICT DETECTION
        # ---------------------------------------------------------
        conflicts = detect_conflicts(riders) if 'detect_conflicts' in globals() else None
        if conflicts:
            session['conflicts'] = conflicts
            session['sub_id_pending_conflict'] = sub.id
            return "conflict"

        # ---------------------------------------------------------
        # PREPARE INVITE MOBILE
        # ---------------------------------------------------------
        def convert_mobile_for_client(m):
            if not m:
                return ""
            digits = re.sub(r'\D+', '', str(m))
            if digits.startswith("61") and len(digits) > 2:
                return "0" + digits[2:]
            return digits

        invite_mobile = convert_mobile_for_client(invite.mobile)

        # ---------------------------------------------------------
        # CREATE / ATTACH LESSONS
        # ---------------------------------------------------------
        created_lessons = []
        errors = []

        clean_time = invite.time_frame

        for rider in riders:
            try:
                # ---------------------------------------------------------
                # STEP 3: Replace placeholder client + update pending lesson
                # ---------------------------------------------------------
                token = rider.get("invite_token")
                invite_for_rider = LessonInvite.query.filter_by(token=token).first()

                if invite_for_rider:
                    # 1. Fetch the pending lesson created at send_invite
                    pending_lesson = Lesson.query.get(invite_for_rider.lesson_id)

                    # 2. Fetch the placeholder client
                    placeholder_client = Client.query.filter_by(full_name=pending_lesson.client).first()

                    # 3. Resolve the REAL client using soft-match logic
                    rider_name = rider.get("full_name") or rider.get("name") or ""
                    match_type, match_data = soft_match_client_for_invite(rider_name)

                    if match_type == "exact":
                        real_client = match_data

                    elif match_type == "none":
                        # Create new client with ONLY name + height/weight
                        real_client = Client(
                            full_name=rider_name,
                            height_cm=rider.get("height_cm"),
                            weight_kg=rider.get("weight_kg"),
                            guardian_name="",
                            mobile="",
                            notes=""
                        )
                        db.session.add(real_client)
                        db.session.flush()

                    else:  # ambiguous
                        sub.needs_client_match = True
                        db.session.commit()
                        return "needs_client_match"

                    # 4. Update the pending lesson with REAL client details
                    pending_lesson.client = real_client.full_name
                    pending_lesson.height_cm = rider.get("height_cm")
                    pending_lesson.weight_kg = rider.get("weight_kg")
                    pending_lesson.horse = ""  # horse not assigned yet
                    pending_lesson.attendance = "Pending"
                    pending_lesson.payment = None

                    # 5. Mark invite as completed for this rider
                    invite_for_rider.status = "completed"

                    # 6. Delete placeholder client
                    db.session.delete(placeholder_client)

                    # 7. Add updated lesson to created list
                    created_lessons.append(pending_lesson)

                    # 8. Continue to next rider (no new lesson created here)
                    continue

                # ---------------------------------------------------------
                # STEP 4: Multi‑rider lesson cloning (rider #2+)
                # ---------------------------------------------------------
                if created_lessons:
                    # This is NOT the first rider — create a cloned lesson row

                    # Resolve client again (soft-match already done above)
                    rider_name = rider.get("full_name") or rider.get("name") or ""
                    match_type, match_data = soft_match_client_for_invite(rider_name)

                    if match_type == "exact":
                        real_client = match_data
                    elif match_type == "none":
                        real_client = Client(
                            full_name=rider_name,
                            height_cm=rider.get("height_cm"),
                            weight_kg=rider.get("weight_kg"),
                            guardian_name="",
                            mobile="",
                            notes=""
                        )
                        db.session.add(real_client)
                        db.session.flush()
                    else:
                        sub.needs_client_match = True
                        db.session.commit()
                        return "needs_client_match"

                    # Clone the first lesson's structure
                    base_lesson = created_lessons[0]

                    new_lesson = Lesson(
                        lesson_date=base_lesson.lesson_date,
                        time_frame=base_lesson.time_frame,
                        lesson_type=base_lesson.lesson_type,
                        group_priv=base_lesson.group_priv,
                        price_pl=base_lesson.price_pl,
                        client=real_client.full_name,
                        height_cm=rider.get("height_cm"),
                        weight_kg=rider.get("weight_kg"),
                        horse="",
                        attendance="Pending",
                        payment=None
                    )

                    db.session.add(new_lesson)
                    db.session.flush()

                    created_lessons.append(new_lesson)
                    continue



            except Exception as e:
                print("RIDER ERROR:", rider, e)   # 👈 add this  
                errors.append(str(e))
                continue

        # ---------------------------------------------------------
        # FINALIZE
        # ---------------------------------------------------------
        if created_lessons:
            # Always link invite to the FIRST lesson created/updated
            first_lesson = created_lessons[0]
            invite.lesson_id = first_lesson.lesson_id

            # Mark invite as completed
            invite.status = "completed"

            # Mark submission as processed
            sub.processed = True

        else:
            # Safety fallback — should never happen now
            sub.processed = True

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"DB commit failed: {e}"

        if errors:
            return f"Processed {len(created_lessons)} riders into lessons; errors: {len(errors)}"
        return f"Processed {len(created_lessons)} riders into lessons"

    @app.route('/client/<int:client_id>/statement')
    def client_statement_pdf(client_id):
        try:
            client = Client.query.get_or_404(client_id)

            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')

            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            lessons = (
                Lesson.query
                .filter(
                    Lesson.client == client.full_name,
                    Lesson.lesson_date >= start_date,
                    Lesson.lesson_date <= end_date
                )
                .order_by(
                    db.func.date(Lesson.lesson_date).asc(),
                    Lesson.lesson_id.asc()
                )
                .all()
            )

            # ---- BUILD ROWS USING TRUE SYSTEM LOGIC ----
            running_balance = None
            rows = []

            for l in lessons:
                price = l.price_pl or 0
                payment = l.payment or 0
                adjust = l.adjust or 0
                att = (l.attendance or '').strip().upper()

                # FIRST LESSON → PRESERVE IMPORTED carry_fwd
                if running_balance is None:
                    carry = l.carry_fwd or 0
                else:
                    carry = running_balance

                # BASE BALANCE
                balance = carry + payment + adjust

                # APPLY CHARGE ONLY IF ATTENDED
                charge = price if att in ['Y', 'N'] else 0
                balance -= charge

                # RUNNING BALANCE
                running_balance = balance

                rows.append({
                    "lesson": l,
                    "payment": payment,
                    "charge": charge,
                    "running_balance": balance
                })

            # ---- SUMMARY TOTALS ----
            lesson_count = len(rows)
            total_payments = sum(r["payment"] for r in rows)
            total_charges = sum(r["charge"] for r in rows)
            final_balance = rows[-1]["running_balance"] if rows else 0

            # GST = 11% of payments
            gst_amount = total_payments * 0.11

            html = render_template(
                'client_statement_pdf.html',
                client=client,
                rows=rows,
                start_date=start_date,
                end_date=end_date,
                lesson_count=lesson_count,
                total_payments=total_payments,
                total_charges=total_charges,
                final_balance=final_balance,
                gst_amount=gst_amount
            )

            pdf = HTML(string=html).write_pdf()
            return send_file(
                io.BytesIO(pdf),
                mimetype='application/pdf',
                download_name=f"{client.full_name}_statement.pdf"
            )

        except Exception as e:
            import traceback, sys
            print(">>> PDF ROUTE ERROR:", e, file=sys.stderr)
            traceback.print_exc()
            raise

    @app.route('/lessons_by_date_pdf')
    def lessons_by_date_pdf():
        db.session.rollback()

        # Parse date from query string
        selected_date, selected_date_str = parse_selected_date()

        if not selected_date:
            selected_date = datetime.now(ZoneInfo("Australia/Brisbane")).date()
            selected_date_str = selected_date.strftime('%Y-%m-%d')

        # Build the SAME context as the HTML route
        ctx = build_lessons_context(selected_date, selected_date_str)

        # Render the SAME template into HTML
        html = render_template(
            'lessons_by_date.html',
            grouped_lessons=ctx["grouped_lessons"],
            selected_date=selected_date_str,
            weekday_int=selected_date.weekday(),
            horse_list=ctx["horse_list"],
            horse_schedule=ctx["horse_schedule"],
            teacher_names=ctx["teacher_names"],
            block_tag_lookup=ctx["block_tag_lookup"],
            client_horse_history=ctx["client_horse_history"],
            invoice_clients=ctx["invoice_clients"],
            clients=ctx["clients"],
            teacher_times=teacher_times_map(),
            teacher_horse_usage=ctx["teacher_horse_usage"],
            times=ctx["times"],
            get_static_teachers=get_static_teachers,
            client_lookup=ctx["client_lookup"],
            slot_map=ctx["slot_map"]
        )

        # Generate PDF
        pdf = weasyprint.HTML(string=html).write_pdf()

        # Return PDF response
        return Response(
            pdf,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'inline; filename=lessons_{selected_date_str}.pdf'
            }
        )


    @app.route('/notifications/invite_conflicts')
    def invite_conflicts():
        rows = (
            db.session.query(IncomingSubmission)
            .filter_by(needs_client_match=True, processed=False)
            .order_by(IncomingSubmission.id.asc())
            .all()
        )
        return render_template('invite_conflict_queue.html', rows=rows)


    @app.route('/notifications/invite_conflict/<int:submission_id>/<int:rider_index>', methods=['POST'])
    def finalize_invite_conflict(submission_id, rider_index):
        choice = request.form.get("choice")
        client_id = request.form.get("client_id")

        row = db.session.query(IncomingSubmission).get_or_404(submission_id)

        riders = parse_jotform_payload(
            row.raw_payload,
            forced_submission_id=row.id
        )
        rider = riders[rider_index - 1]


        rider_name = rider.get("full_name") or rider.get("name") or ""

        # ---------------------------------------------------------
        # SAFE FIELD EXTRACTION (required for overwrite + update)
        # ---------------------------------------------------------
        def safe_int(value):
            if value is None:
                return None
            if isinstance(value, int):
                return value
            value = str(value).strip()
            return int(value) if value.isdigit() else None

        def safe_text(value):
            if not value:
                return None
            value = str(value).strip()
            return value if value not in ("", "N/A") else None

        age = safe_int(rider.get("age"))
        disclaimer = safe_int(
            rider.get("disclaimer") or
            rider.get("id") or
            rider.get("disclaimer_number") or
            rider.get("disc_no")
        )
        height_cm = safe_int(rider.get("height_cm"))
        weight_kg = safe_int(rider.get("weight_kg"))
        notes = safe_text(rider.get("notes"))
        guardian = safe_text(rider.get("guardian"))
        mobile = safe_text(rider.get("mobile"))
        email = safe_text(rider.get("email"))



        # ---------------------------------------------------------
        # USE EXISTING (NEVER CREATE DUPLICATES)
        # ---------------------------------------------------------
        if choice == "use_existing" and client_id:
            client = db.session.query(Client).get(int(client_id))
            if not client:
                return "Client not found", 400

        # ---------------------------------------------------------
        # RESOLVE CLIENT BASED ON CHOICE
        # ---------------------------------------------------------
        elif choice == "use_existing" and client_id:
            client = db.session.query(Client).get(int(client_id))


        elif choice == "new":
            client = Client(
                full_name=rider_name,
                height_cm=rider.get("height_cm"),
                weight_kg=rider.get("weight_kg"),
                guardian_name="",
                mobile="",
                notes=""
            )
            db.session.add(client)
            db.session.flush()

        elif choice == "new_same_name":
            base = rider_name
            counter = 2
            while True:
                candidate = f"{base} ({counter})"
                exists = db.session.query(Client).filter_by(full_name=candidate).first()
                if not exists:
                    break
                counter += 1

            client = Client(
                full_name=clean_name(candidate),
                height_cm=rider.get("height_cm"),
                weight_kg=rider.get("weight_kg"),
                guardian_name="",
                mobile="",
                notes=""
            )
            db.session.add(client)
            db.session.flush()

        elif choice == "ignore":
            row.processed = True
            row.ignored = True
            db.session.commit()
            return redirect(url_for('invite_conflicts'))

        else:
            return "Invalid choice", 400

        # ---------------------------------------------------------
        # STORE RESOLVED CLIENT ID ON RIDER OBJECT
        # ---------------------------------------------------------
        rider["resolved_client_id"] = client.client_id

        # ---------------------------------------------------------
        # CHECK IF ALL RIDERS ARE NOW RESOLVED
        # ---------------------------------------------------------
        all_resolved = True
        for r in riders:
            if not r.get("resolved_client_id"):
                all_resolved = False
                break

        if not all_resolved:
            db.session.commit()
            return redirect(url_for('invite_conflicts'))

        # ---------------------------------------------------------
        # ALL RIDERS RESOLVED → CLEAR FLAG AND REPROCESS
        # ---------------------------------------------------------
        row.needs_client_match = False
        db.session.commit()

        return redirect(url_for('process_all_pending'))



    # ---------------- ROUTES ---------------- #

    @app.route('/client/change_name/<int:client_id>', methods=['POST'])
    def change_client_name(client_id):
        new_name = request.form.get('new_name', '').strip()
        if not new_name:
            flash("New name cannot be empty.", "danger")
            return redirect(url_for('client_view', client=client_id))

        client = db.session.query(Client).get(client_id)
        if not client:
            flash("Client not found.", "danger")
            return redirect(url_for('client_view'))

        old_name = client.full_name

        # Update client record
        client.full_name = new_name

        # Update all lessons referencing the old name
        db.session.query(Lesson).filter(
            func.lower(func.trim(Lesson.client)) == func.lower(func.trim(old_name))
        ).update(
            {Lesson.client: new_name},
            synchronize_session=False
        )

        db.session.commit()

        flash(f"Name updated from '{old_name}' to '{new_name}'.", "success")
        return redirect(url_for('client_view', client=client_id))


    @app.route("/export_clients_xlsx")
    @login_required
    @role_required("admin")
    def export_clients_xlsx():
        from openpyxl import Workbook
        from flask import send_file
        import io

        # Correct model: Client (not Clients)
        clients = Client.query.order_by(Client.full_name.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"

        ws["A1"] = "Client"
        ws["B1"] = "Disclaimer Number"

        row = 2
        for c in clients:
            ws[f"A{row}"] = c.full_name
            ws[f"B{row}"] = c.disclaimer
            row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="clients_alpha.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route('/client_history/<client_name>')
    def client_history(client_name):

        today = datetime.now(ZoneInfo("Australia/Brisbane")).date()

        rows = (
            db.session.query(Lesson)
            .filter(
                Lesson.client == client_name,
                Lesson.lesson_date < today
            )
            .order_by(Lesson.lesson_date.desc(), Lesson.lesson_id.desc())
            .limit(50)   # fetch more so we can extract 10 unique
            .all()
        )

        # Extract horse names, strip blanks
        horses = [(r.horse or "").strip() for r in rows]
        clean = [h for h in horses if h]

        # Preserve order but remove duplicates
        unique = []
        seen = set()
        for h in clean:
            if h not in seen:
                unique.append(h)
                seen.add(h)

        # Limit to last 10 unique
        result = unique[:10]

        return jsonify(result)

    
    @app.route('/send_invite', methods=['POST'])
    def send_invite():
        print("DEBUG AUTH:", os.getenv("CLICKSEND_USERNAME"), os.getenv("CLICKSEND_API_KEY"))
        print("SMS INVITE POST:", dict(request.form))
        # ------------------------------
        # 1. Extract posted fields
        # ------------------------------
        lesson_date_str = (
            request.form.get("invite_date")
            or request.form.get("lesson_date")
            or request.form.get("date")
        )

        invite_time_frame = (request.form.get("invite_time_frame") or "").strip()
        invite_lesson_type = (
            request.form.get("invite_lesson_type")
            or request.form.get("lesson_type")
            or ""
        ).strip()

        invite_group_priv = (
            request.form.get("invite_group_priv")
            or request.form.get("group_priv")
            or ""
        ).strip()

        mobile_raw = (
            request.form.get("invite_mobile")
            or request.form.get("mobile")
            or request.form.get("client_phone")
            or ""
        )

        riders_requested_raw = (
            request.form.get("invite_riders")
            or request.form.get("riders_requested")
            or "1"
        )

        cost_per_person_raw = (
            request.form.get("invite_cost")
            or request.form.get("cost_per_person")
            or "0"
        )

        teacher_id_raw = request.form.get("invite_teacher_id")
        try:
            teacher_id = int(teacher_id_raw)
        except:
            teacher_id = None

        # Canonical time frame
        lesson_time_str = invite_time_frame

        # ------------------------------
        # 2. Validate required fields
        # ------------------------------
        digits = re.sub(r'\D+', '', mobile_raw or "")
        if digits.startswith("0"):
            digits = digits[1:]
        mobile_clean = "+61" + digits if digits else ""

        if not lesson_date_str or not lesson_time_str or not mobile_clean:
            flash("Missing lesson date, time, or mobile number for invite.", "danger")
            return redirect(url_for('lessons_by_date', date=lesson_date_str or date.today().isoformat()))

        try:
            riders_requested = int(riders_requested_raw)
        except:
            riders_requested = 1

        try:
            cost_per_person = float(cost_per_person_raw)
        except:
            cost_per_person = 0.0

        # ------------------------------
        # 3. Generate token
        # ------------------------------
        token = generate_invite_token(lesson_date_str, lesson_time_str)

        # ------------------------------
        # 4. CREATE PLACEHOLDER CLIENT
        # ------------------------------
        placeholder_client = Client(
            full_name=f"[Pending {token}]",
            mobile=mobile_clean,
            notes="Pending SMS invite"
        )
        db.session.add(placeholder_client)
        db.session.flush()

        # ------------------------------
        # 5. CREATE PENDING LESSON
        # ------------------------------
        lesson = Lesson(
            lesson_date=datetime.strptime(lesson_date_str, "%Y-%m-%d").date(),
            time_frame=lesson_time_str,
            lesson_type=invite_lesson_type,
            group_priv=invite_group_priv,
            price_pl=cost_per_person,
            client=placeholder_client.full_name,
            horse=""
        )
        db.session.add(lesson)
        db.session.flush()

        # ------------------------------
        # 6. CREATE LESSON INVITE
        # ------------------------------
        invite = LessonInvite(
            lesson_id=lesson.lesson_id,
            token=token,
            mobile=mobile_clean,
            riders_requested=riders_requested,
            cost_per_person=cost_per_person,
            time_frame=lesson_time_str,
            lesson_type=invite_lesson_type,
            group_priv=invite_group_priv,
            status="awaiting_form",
            lesson_date=datetime.strptime(lesson_date_str, "%Y-%m-%d").date()
        )

        db.session.add(invite)
        db.session.commit()

        # ------------------------------
        # 7. Build SMS + send
        # ------------------------------
        base_url = "https://form.jotform.com/253599154628066"
        params = urlencode({
            "i_t": token,
            "r_r": riders_requested
        })
        jotform_url = f"{base_url}?{params}"

        from dateutil import parser
        try:
            date_obj = parser.parse(lesson_date_str)
            formatted_date = date_obj.strftime("%-d %b")
        except:
            formatted_date = lesson_date_str

        m = re.search(r'(\d{1,2}:\d{2})', lesson_time_str or "")
        start_24 = m.group(1) if m else lesson_time_str.split("-")[0].strip()

        try:
            t_obj = datetime.strptime(start_24, "%H:%M")
            hour_12 = t_obj.strftime("%I").lstrip("0")
            minute = t_obj.strftime("%M")
            ampm = t_obj.strftime("%p").lower()
            start_pretty = f"{hour_12}.{minute}{ampm}"
        except:
            start_pretty = start_24

        rider_word = "rider" if riders_requested == 1 else "riders"

        sms_text = (
            f"Hi! Please confirm your lesson on {formatted_date} at {start_pretty}. "
            f"{riders_requested} {rider_word}, ${cost_per_person:.0f} ea. "
            f"{jotform_url}"
        )

        sender = app.config.get("EQUESTRIAN_SENDER", "")

        ok = send_sms_clicksend(mobile_clean, sms_text, sender)
        flash("SMS invite sent successfully." if ok else "Failed to send SMS invite.", 
              "success" if ok else "danger")

        return redirect(url_for('lessons_by_date', date=lesson_date_str))



    # ---------------- ROUTES ---------------- #

    @app.route('/delete_client', methods=['POST'])
    def delete_client():
        client_id = request.form.get('client_id')
        client = Client.query.get(client_id)

        if client:
            db.session.delete(client)
            db.session.commit()

        return redirect(url_for('client_view'))


    @app.post("/save_booking")
    def save_booking():
        data = request.get_json()

        booking_type = data.get("booking_type")

        if booking_type == "existing":
            return handle_existing_client(data)

        if booking_type == "new":
            return handle_new_client(data)

        return jsonify(success=False, error="Invalid booking type")





    @app.route("/gift_vouchers")
    def gift_vouchers():
        from models import GiftVoucherSubmission

        unprocessed = GiftVoucherSubmission.query.filter_by(ignored=False, processed=False).order_by(GiftVoucherSubmission.created_at.desc()).all()
        processed = GiftVoucherSubmission.query.filter_by(processed=True).order_by(GiftVoucherSubmission.processed_at.desc()).all()
        ignored = GiftVoucherSubmission.query.filter_by(ignored=True).order_by(GiftVoucherSubmission.created_at.desc()).all()

        return render_template(
            "gift_vouchers.html",
            unprocessed=unprocessed,
            processed=processed,
            ignored=ignored
        )



    @app.route("/gift_voucher_ignore/<int:id>")
    def gift_voucher_ignore(id):
        from models import GiftVoucherSubmission

        v = GiftVoucherSubmission.query.get_or_404(id)
        v.ignored = True
        db.session.commit()

        flash("Voucher marked as ignored.", "warning")
        return redirect(url_for("gift_vouchers"))


    @app.route("/gift_voucher_unignore/<int:id>")
    def gift_voucher_unignore(id):
        from models import GiftVoucherSubmission

        v = GiftVoucherSubmission.query.get_or_404(id)
        v.ignored = False
        db.session.commit()

        flash("Voucher unignored.", "success")
        return redirect(url_for("gift_vouchers"))

    @app.route("/gift_voucher_process/<int:id>")
    def gift_voucher_process(id):
        from datetime import datetime, date
        from models import GiftVoucherSubmission, Client, Lesson

        v = GiftVoucherSubmission.query.get_or_404(id)

        # Already processed?
        if v.processed:
            flash("This voucher has already been processed.", "info")
            return redirect(url_for("gift_vouchers"))

        # 1. Find or create client (recipient)
        client = Client.query.filter_by(full_name=v.recipient_name).first()
        if not client:
            client = Client(full_name=v.recipient_name)
            db.session.add(client)
            db.session.commit()

        # Convert "$170" → 170.0
        amount = float(v.amount_payable.replace("$", "").strip())

        # 2. Create the Voucher CR lesson entry
        lesson = Lesson(
            lesson_date=date.today(),
            time_frame="",
            block_key="",
            client=v.recipient_name,
            horse="",
            adjust=0,
            carry_fwd=0,
            payment=amount,      # voucher credit
            price_pl=0,          # NOT a lesson price
            attendance="",
            balance=0,
            lesson_notes="",
            lesson_type="Voucher CR",
            group_priv="",
            blockends=None,
            lesson_no=None,
            freq="S",
            voucher_number=v.voucher_number
        )

        db.session.add(lesson)

        # 3. Mark submission processed
        v.processed = True
        v.processed_at = datetime.now()

        db.session.commit()

        flash("Gift voucher processed and credit added.", "success")
        return redirect(url_for("gift_vouchers"))


    @app.route("/gift_voucher_view/<int:id>")
    def gift_voucher_view(id):
        from models import GiftVoucherSubmission

        v = GiftVoucherSubmission.query.get_or_404(id)
        return render_template("gift_voucher_view.html", v=v)




    @app.route('/wedding/sms', methods=['GET'])
    def wedding_sms():
        return render_template('wedding_sms.html')

    @app.route("/purge_jotform_invites")
    def purge_jotform_invites():
        import requests

        API_KEY = os.getenv("JOTFORM_API_KEY", "")
        FORM_ID = "253599154628066"

        # Fetch all submissions
        url = (
            f"https://api.jotform.com/form/{FORM_ID}/submissions"
            f"?apiKey={API_KEY}&orderby=created_at&direction=DESC&limit=1000"
        )
        data = requests.get(url).json()
        subs = data.get("content", [])

        count = 0
        for s in subs:
            sid = s.get("id")
            if sid:
                del_url = f"https://api.jotform.com/submission/{sid}?apiKey={API_KEY}"
                requests.delete(del_url)
                count += 1

        return f"Deleted {count} JotForm submissions."


    @app.route('/u/<code>')
    def short_redirect(code):
        if code in short_links:
            return redirect(short_links[code])
        return "Invalid or expired link", 404


    @app.route('/debug', methods=['GET', 'POST'])
    def debug_page():
        selected_client = request.args.get('client') or request.form.get('client')

        # Query all clients
        q = db.session.query(Client).order_by(Client.full_name.asc()).all()

        clients = []
        for row in q:
            if isinstance(row, (list, tuple)):
                v = row[0]
            else:
                v = getattr(row, 'full_name', None) if hasattr(row, 'full_name') else str(row)

            if v:
                v = str(v).strip()
                if v:
                    clients.append(v)

        # Deduplicate
        seen = set()
        clients = [c for c in clients if not (c in seen or seen.add(c))]

        print("DEBUG clients payload (count):", len(clients), "sample:", clients[:50])

        return render_template('debug.html', clients=clients, selected_client=selected_client)


    @app.route('/notifications')
    def notifications():
        """
        Fast notifications listing:
        - Limits rows rendered (prevents timeouts)
        - Parses payload in LIGHT mode (no client-table load / no matching)
        """
        page = request.args.get("page", default=1, type=int)
        per_page = request.args.get("per_page", default=50, type=int)
        per_page = max(10, min(per_page, 200))  # clamp

        rows = (
            db.session.query(IncomingSubmission)
            .filter_by(processed=False, ignored=False)
            .order_by(IncomingSubmission.received_at.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        for r in rows:
            try:
                riders = parse_jotform_payload(
                    r.raw_payload,
                    forced_submission_id=r.id,        # ✅
                    clients_cache=None,
                    mode="light"
                )

                names = []
                for rider in riders:
                    n = rider.get("name")
                    if n:
                        names.append(n)

                r.display_names = ", ".join(names) if names else "Invite Submission"

            except Exception as e:
                print("ERROR extracting names:", e)
                r.display_names = "(unknown)"

        return render_template(
            'notifications.html',
            rows=rows,
            page=page,
            per_page=per_page
        )

        
    @app.route('/notifications/fetch')
    def fetch_jotform_submissions():
        import requests
        import json
        import hashlib
        from datetime import datetime, timedelta

        API_KEY = os.getenv("JOTFORM_API_KEY", "")
        FORM_ID = "211021514885045"   # Disclaimer & Indemnity form

        # ---------------------------------------------------------
        # 0. LOAD GLOBAL MAX DISCLAIMER NUMBER
        # ---------------------------------------------------------
        state = DisclaimerState.query.first()
        if not state:
            state = DisclaimerState(max_disclaimer_number=0)
            db.session.add(state)
            db.session.commit()

        current_max_disclaimer = state.max_disclaimer_number or 0

        # ---------------------------------------------------------
        # 1. DYNAMIC CUTOFF — ignore anything older than 30 days
        # ---------------------------------------------------------
        CUTOFF = datetime.utcnow() - timedelta(days=30)

        # ---------------------------------------------------------
        # 2. CHECKPOINT — latest submission (ANY state)
        # ---------------------------------------------------------
        latest_any = (
            db.session.query(IncomingSubmission)
            .filter(IncomingSubmission.form_id == FORM_ID)
            .order_by(IncomingSubmission.received_at.desc())
            .first()
        )
        latest_ts = latest_any.received_at if latest_any else None

        # ---------------------------------------------------------
        # 3. PAGINATION — fetch ALL pages from JotForm
        # ---------------------------------------------------------
        submissions = []
        offset = 0
        limit = 1000  # JotForm max per page

        while True:
            url = (
                f"https://api.jotform.com/form/{FORM_ID}/submissions"
                f"?apiKey={API_KEY}&offset={offset}&limit={limit}"
            )
            r = requests.get(url)
            if r.status_code != 200:
                print("ERROR: JotForm API failed:", r.text)
                break

            page = r.json().get("content", [])
            if not page:
                break

            submissions.extend(page)
            offset += limit

        print(f"Fetched {len(submissions)} submissions from JotForm.")

        inserted = 0
        new_global_max = current_max_disclaimer

        # ---------------------------------------------------------
        # 4. PROCESS EACH SUBMISSION
        # ---------------------------------------------------------
        for sub in submissions:
            submission_id = str(sub.get("id"))

            # PRIMARY DEDUPE — submission_id
            existing = (
                db.session.query(IncomingSubmission)
                .filter_by(submission_id=submission_id)
                .first()
            )
            if existing:
                continue

            # EXTRACT ALL DISCLAIMER NUMBERS FROM PAYLOAD
            answers = sub.get("answers", {}) or {}
            disclaimer_numbers = []

            for key, val in answers.items():
                text = (val.get("text") or "").lower()
                if "disclaimer" in text:
                    ans = val.get("answer")
                    if isinstance(ans, list):
                        disclaimer_numbers.extend(ans)
                    elif ans is not None:
                        disclaimer_numbers.append(ans)

            # NORMALISE TO INTS AND FIND MAX
            numeric_disclaimers = []
            for dn in disclaimer_numbers:
                try:
                    numeric_disclaimers.append(int(str(dn).strip()))
                except Exception:
                    continue

            if numeric_disclaimers:
                max_in_submission = max(numeric_disclaimers)
            else:
                max_in_submission = None

            # HARD RULE:
            # If this submission's max disclaimer number is
            # <= global max, it is OLD and must be ignored.
            if max_in_submission is not None:
                if max_in_submission <= current_max_disclaimer:
                    continue
                # Track the highest we've seen this run
                if max_in_submission > new_global_max:
                    new_global_max = max_in_submission

            # SECONDARY DEDUPE — hash of payload
            payload_str = json.dumps(sub, sort_keys=True)
            payload_hash = hashlib.sha256(payload_str.encode()).hexdigest()

            # TIMESTAMP
            submission_created = sub.get("created_at") or sub.get("updated_at")

            try:
                if submission_created:
                    # JotForm sends ISO8601 strings, not epoch ints
                    submission_dt = datetime.fromisoformat(
                        submission_created.replace("Z", "")
                    )
                else:
                    submission_dt = datetime.utcnow()
            except Exception as e:
                print("TIMESTAMP ERROR:", e, "RAW VALUE:", submission_created)
                submission_dt = datetime.utcnow()



            # CUTOFF
            if submission_dt < CUTOFF:
                continue

            # INSERT NEW ROW
            row = IncomingSubmission(
                submission_id=submission_id,
                form_id=FORM_ID,
                raw_payload=sub,
                processed=False,
                unique_hash=payload_hash,
                received_at=submission_dt,
                jotform_id=submission_id
            )
            db.session.add(row)
            inserted += 1

        # ---------------------------------------------------------
        # 5. UPDATE GLOBAL MAX DISCLAIMER NUMBER
        # ---------------------------------------------------------
        if new_global_max > current_max_disclaimer:
            state.max_disclaimer_number = new_global_max
            db.session.commit()
            print(f"Updated max disclaimer number to {new_global_max}")
        else:
            db.session.commit()

        print(f"Inserted {inserted} new submissions.")

        return redirect(url_for('notifications'))


    @app.route('/notifications/<int:webhook_id>')
    def process_notification(webhook_id):
        submission = db.session.query(IncomingSubmission).get_or_404(webhook_id)
        riders = parse_jotform_payload(
            submission.raw_payload,
            forced_submission_id=submission.id
        )

        # NEW: filter valid riders (skip incomplete)
        valid_riders = [r for r in riders if not r.get("incomplete")]

        # NEW: if all riders incomplete → auto-ignore
        if not valid_riders:
            submission.processed = True
            submission.ignored = True
            submission.processed_at = datetime.utcnow()
            db.session.commit()
            return redirect(url_for('notifications'))

        # NEW: conflict detection only on valid riders
        for i, rider in enumerate(valid_riders, start=1):
            if rider.get("matches"):
                return redirect(url_for(
                    'resolve_conflict',
                    submission_id=submission.id,
                    rider_index=i
                ))

        # No conflicts → show normal processing screen
        return render_template(
            'process_notification.html',
            submission=submission,
            clients=valid_riders
        )





    @app.route('/notifications/conflict/<int:submission_id>/<int:rider_index>', methods=['GET'])
    def resolve_conflict(submission_id, rider_index):
        row = db.session.query(IncomingSubmission).get_or_404(submission_id)
        riders = parse_jotform_payload(row.raw_payload, forced_submission_id=row.id)

        # rider_index is 1‑based
        rider = riders[rider_index - 1]
        print("RIDER DEBUG:", rider)
        print("RIDER KEYS:", rider.keys())

        # Raw matches from parse_jotform_payload (SQLAlchemy tuples)
        raw_matches = rider.get("matches", [])

        # Convert SQLAlchemy row tuples → dicts
        match_dicts = [
            {
                "client_id": m[0],
                "full_name": m[1],
                "mobile": m[2],
                "email": m[3],
                "jotform_submission_id": m[4]
            }
            for m in raw_matches
        ]

        # Extract client IDs
        match_ids = [m["client_id"] for m in match_dicts]

        # Load REAL Client objects
        matches = Client.query.filter(Client.client_id.in_(match_ids)).all()

        # ⭐ Attach last lesson date
        for m in matches:
            last = (
                Lesson.query
                .filter_by(client=m.full_name)
                .order_by(Lesson.lesson_date.desc())
                .first()
            )
            m.last_lesson_date = last.lesson_date if last else None

        return render_template(
            'conflict_resolution.html',
            submission=row,
            rider=rider,
            matches=matches,
            rider_index=rider_index
        )


    @app.route('/notifications/conflict/<int:submission_id>/<int:rider_index>', methods=['POST'])
    def finalize_conflict(submission_id, rider_index):
        choice = request.form.get("choice")
        client_id = request.form.get("client_id")

        row = db.session.query(IncomingSubmission).get_or_404(submission_id)

        # Parse riders once
        riders = parse_jotform_payload(
            row.raw_payload,
            forced_submission_id=row.id
        )
        rider = riders[rider_index - 1]

        # Skip incomplete riders
        if rider.get("incomplete"):
            row.processed = True
            row.processed_at = datetime.utcnow()
            db.session.commit()
            return redirect(url_for('process_all_pending'))

        # ---------------------------------------------------------
        # IGNORE OPTION — HARD STOP (NO FALLTHROUGH)
        # ---------------------------------------------------------
        if choice == "ignore":
                row.processed = True
                row.ignored = True
                row.processed_at = datetime.utcnow()
                db.session.commit()
                return redirect(url_for('process_all_pending'))

        # Preload clients
        all_clients = db.session.query(Client).all()
        clients_by_id = {c.client_id: c for c in all_clients}
        existing = clients_by_id.get(int(client_id)) if client_id else None

        # ---------------------------------------------------------
        # SAFE EXTRACTION (prevents NameErrors + bad values)
        # ---------------------------------------------------------
        def safe_int(v):
            if v is None:
                return None
            if isinstance(v, int):
                return v
            v = str(v).strip()
            return int(v) if v.isdigit() else None

        def safe_text(v):
            if not v:
                return None
            v = str(v).strip()
            return v if v not in ("", "N/A") else None

        raw_name = rider.get("name")
        name = clean_name(raw_name)
        age = safe_int(rider.get("age"))
        guardian = safe_text(rider.get("guardian"))
        mobile = clean_mobile(rider.get("mobile"))
        email = safe_text(rider.get("email"))
        disclaimer = safe_int(
            rider.get("disclaimer") or
            rider.get("id") or
            rider.get("disclaimer_number") or
            rider.get("disc_no")
        )
        height_cm = safe_int(rider.get("height_cm"))
        weight_kg = safe_int(rider.get("weight_kg"))
        notes = safe_text(rider.get("notes"))

        jotform_id = str(row.form_id)

        # ---------------------------------------------------------
        # USE EXISTING (SAFE MERGE)
        # ---------------------------------------------------------
        if choice == "use_existing" and existing:

            if guardian:
                existing.guardian_name = guardian

            if age is not None:
                existing.age = age

            if mobile:
                existing.mobile = mobile

            if email:
                existing.email_primary = email

            if disclaimer is not None:
                existing.disclaimer = disclaimer

            if height_cm is not None:
                existing.height_cm = height_cm

            if weight_kg is not None:
                existing.weight_kg = weight_kg

            if notes is not None:
                existing.notes = notes

            existing.jotform_submission_id = jotform_id

            # ⭐ NEW: mark submission processed
            row.processed = True
            row.processed_at = datetime.utcnow()

            db.session.commit()
            return redirect(url_for('process_all_pending'))


        # ---------------------------------------------------------
        # OVERWRITE EXISTING (FULL REPLACEMENT)
        # ---------------------------------------------------------
        if choice == "overwrite" and existing:

            existing.full_name = clean_name(name)
            existing.age = age
            existing.guardian_name = guardian
            existing.mobile = mobile
            existing.email_primary = email
            existing.disclaimer = disclaimer

            existing.height_cm = height_cm
            existing.weight_kg = weight_kg
            existing.notes = notes

            existing.jotform_submission_id = jotform_id

            # ⭐ NEW: mark submission processed
            row.processed = True
            row.processed_at = datetime.utcnow()

            db.session.commit()
            return redirect(url_for('process_all_pending'))


        # ---------------------------------------------------------
        # CREATE NEW CLIENT
        # ---------------------------------------------------------
        if choice == "new":
            new_client = Client(
                full_name=clean_name(name),
                age=age,
                guardian_name=guardian,
                mobile=mobile,
                email_primary=email,
                disclaimer=disclaimer,
                height_cm=height_cm,
                weight_kg=weight_kg,
                notes=notes,
                invoice_required=False,
                jotform_submission_id=jotform_id
            )
            db.session.add(new_client)

            # ⭐ NEW: mark submission processed
            row.processed = True
            row.processed_at = datetime.utcnow()

            db.session.commit()
            return redirect(url_for('process_all_pending'))


        # ---------------------------------------------------------
        # CREATE NEW CLIENT (SAME NAME)
        # ---------------------------------------------------------
        if choice == "new_same_name":
            base = clean_name(name)
            counter = 2

            while True:
                candidate = f"{base} ({counter})"
                exists = db.session.query(Client).filter_by(full_name=candidate).first()
                if not exists:
                    break
                counter += 1

            new_client = Client(
                full_name=clean_name(candidate),
                age=age,
                guardian_name=guardian,
                mobile=mobile,
                email_primary=email,
                disclaimer=disclaimer,
                height_cm=height_cm,
                weight_kg=weight_kg,
                notes=notes,
                invoice_required=False,
                jotform_submission_id=jotform_id
            )
            db.session.add(new_client)

            # ⭐ NEW: mark submission processed
            row.processed = True
            row.processed_at = datetime.utcnow()

            db.session.commit()
            return redirect(url_for('process_all_pending'))

        # ---------------------------------------------------------
        # FALLBACK
        # ---------------------------------------------------------
        db.session.commit()
        return redirect(url_for('process_all_pending'))

    @app.route("/pricing_setup")
    def pricing_setup():
        pricing_rows = GroupPricing.query.order_by(GroupPricing.group_priv).all()
        return render_template("pricing_setup.html", pricing=pricing_rows)

    @app.route("/update_pricing_field", methods=["POST"])
    def update_pricing_field():
        data = request.get_json()
        pid = data.get("id")
        field = data.get("field")
        value = data.get("value")

        try:
            row = GroupPricing.query.get(pid)
            if not row:
                return {"success": False, "error": "Pricing row not found"}

            if hasattr(row, field):
                setattr(row, field, float(value))
            else:
                return {"success": False, "error": "Invalid field"}

            db.session.commit()
            return {"success": True}

        except Exception as e:
            return {"success": False, "error": str(e)}


    @app.route('/notifications/<int:webhook_id>', methods=['POST'])
    def finalize_notification(webhook_id):
        row = db.session.query(IncomingSubmission).get_or_404(webhook_id)

        # PHASE 3: Parse riders ONCE
        riders = parse_jotform_payload(
            row.raw_payload,
            forced_submission_id=row.id
        )

        jotform_id = str(row.form_id)

        # PHASE 3: Skip incomplete riders
        valid_riders = [r for r in riders if not r.get("incomplete")]

        # If all riders incomplete → auto-ignore
        if not valid_riders:
            row.processed = True
            row.ignored = True
            row.processed_at = datetime.utcnow()
            db.session.commit()
            return redirect(url_for('notifications'))

        # PHASE 3: Preload clients ONCE
        all_clients = db.session.query(Client).all()
        clients_by_id = {c.id: c for c in all_clients}

        for i, rider in enumerate(valid_riders, start=1):
            choice = request.form.get(f"client_choice_{i}")

            # Safety skip
            if rider.get("incomplete"):
                continue

            # Extract fields
            name = request.form.get(f"name_{i}") or rider.get("name") or "Unknown"
            age = request.form.get(f"age_{i}") or rider.get("age")
            guardian = request.form.get(f"guardian_{i}") or rider.get("guardian")

            raw_mobile = request.form.get(f"mobile_{i}") or rider.get("mobile")
            mobile = clean_mobile(raw_mobile) if raw_mobile else None

            email = request.form.get(f"email_{i}") or rider.get("email")

            raw_disclaimer = request.form.get(f"disclaimer_{i}") or rider.get("disclaimer")
            try:
                disclaimer = int(raw_disclaimer) if raw_disclaimer else None
            except ValueError:
                disclaimer = None

            height_cm = extract_number(request.form.get(f"height_{i}") or rider.get("height_cm"))
            weight_kg = extract_number(request.form.get(f"weight_{i}") or rider.get("weight_kg"))
            notes = request.form.get(f"notes_{i}") or rider.get("notes") or ""

            # --- CREATE NEW ---
            if choice == "new":
                new_client = Client(
                    full_name=name,
                    guardian_name=guardian,
                    age=age,
                    mobile=mobile,
                    email_primary=email,
                    disclaimer=disclaimer,
                    height_cm=height_cm,
                    weight_kg=weight_kg,
                    notes=notes,
                    invoice_required=False,
                    jotform_submission_id=jotform_id
                )
                db.session.add(new_client)
                db.session.flush()
                log_submission_link("CREATE_NEW", new_client, jotform_id)
                continue

            # --- CREATE NEW (SAME NAME) ---
            if choice == "new_same_name":
                final_name = generate_unique_client_name(name)
                new_client = Client(
                    full_name=final_name,
                    guardian_name=guardian,
                    age=age,
                    mobile=mobile,
                    email_primary=email,
                    disclaimer=disclaimer,
                    height_cm=height_cm,
                    weight_kg=weight_kg,
                    notes=notes,
                    invoice_required=False,
                    jotform_submission_id=jotform_id
                )
                db.session.add(new_client)
                db.session.flush()
                log_submission_link("CREATE_NEW_SAME_NAME", new_client, jotform_id)
                continue

            # --- USE EXISTING ---
            if choice and choice.startswith("existing_"):
                existing_id = int(choice.replace("existing_", ""))
                client = clients_by_id.get(existing_id)
                if client:
                    client.jotform_submission_id = jotform_id
                    log_submission_link("USE_EXISTING", client, jotform_id)
                continue

            # --- OVERWRITE EXISTING ---
            if choice and choice.startswith("overwrite_"):
                existing_id = int(choice.replace("overwrite_", ""))
                client = clients_by_id.get(existing_id)
                if client:
                    client.full_name = name
                    client.guardian_name = guardian
                    client.age = age
                    client.mobile = mobile
                    client.email_primary = email
                    client.disclaimer = disclaimer
                    client.height_cm = height_cm
                    client.weight_kg = weight_kg
                    client.notes = notes
                    client.jotform_submission_id = jotform_id
                    log_submission_link("OVERWRITE_EXISTING", client, jotform_id)
                continue

        # SAFE LOGGING
        riders_for_log = parse_jotform_payload(
            row.raw_payload,
            forced_submission_id=row.id
        )
        safe_names = [r.get("name") for r in riders_for_log if r.get("name")]
        if safe_names:
            log_disclaimer_processed(safe_names)
        else:
            log_disclaimer_processed(["(invite submission)"])

        row.processed = True
        row.processed_at = datetime.utcnow()
        db.session.commit()

        return redirect(url_for('notifications'))

    @app.route('/notifications/process_all')
    def process_all_pending():
        from datetime import datetime

        # Get the next unprocessed + not ignored submission
        next_row = (
            db.session.query(IncomingSubmission)
            .filter_by(processed=False, ignored=False)
            .first()
        )

        if not next_row:
            return redirect(url_for('notifications'))

        # -----------------------------------------
        # PHASE 3: Parse riders ONCE
        # -----------------------------------------
        riders = parse_jotform_payload(
            next_row.raw_payload,
            forced_submission_id=next_row.id        # ✅
        )

        # -----------------------------------------
        # PHASE 3: Skip incomplete riders
        # -----------------------------------------
        valid_riders = [r for r in riders if not r.get("incomplete")]

        # If all riders incomplete → auto-ignore
        if not valid_riders:
            next_row.processed = True
            next_row.ignored = True
            next_row.processed_at = datetime.utcnow()
            db.session.commit()
            return redirect(url_for('process_all_pending'))

        # -----------------------------------------
        # PHASE 3: Preload clients ONCE
        # -----------------------------------------
        all_clients = db.session.query(Client).all()
        clients_by_name = {
            normalise_full_name(c.full_name): c
            for c in all_clients
        }

        # -----------------------------------------
        # PHASE 3: Conflict detection (fast)
        # -----------------------------------------
        for idx, rider in enumerate(valid_riders, start=1):
            matches = rider.get("matches", [])
            if matches:
                return redirect(url_for(
                    'finalize_conflict',
                    submission_id=next_row.id,
                    rider_index=idx
                ))

        # -----------------------------------------
        # PHASE 3: No conflicts → fast create
        # -----------------------------------------
        for rider in valid_riders:
            name = smart_proper_name(rider["name"])
            age = int(rider["age"]) if rider["age"] else None
            guardian = rider["guardian"]
            mobile = clean_mobile(rider["mobile"])
            email = rider["email"]
            disclaimer = int(rider["disclaimer"]) if rider["disclaimer"] else None
            jotform_id = rider["jotform_submission_id"]

            new_client = Client(
                full_name=name,
                age=age,
                guardian_name=guardian,
                mobile=mobile,
                email_primary=email,
                disclaimer=disclaimer,
                height_cm=rider.get("height_cm"),
                weight_kg=rider.get("weight_kg"),
                notes=rider.get("notes"),
                jotform_submission_id=jotform_id
            )
            db.session.add(new_client)

        # -----------------------------------------
        # Mark processed
        # -----------------------------------------
        names = [r["name"] for r in valid_riders]
        log_disclaimer_processed(names)

        next_row.processed = True
        next_row.processed_at = datetime.utcnow()
        db.session.commit()

        return redirect(url_for('process_all_pending'))


    @app.route('/admin/cleanup_incoming_submissions')
    def cleanup_incoming_submissions():
        """
        Deletes processed incoming submissions older than 40 days.
        Safe to run — unprocessed submissions are never touched.
        """
        cutoff = datetime.utcnow() - timedelta(days=40)

        old = (
            db.session.query(IncomingSubmission)
            .filter(
                IncomingSubmission.processed == True,
                IncomingSubmission.received_at < cutoff
            )
        )

        count = old.count()
        old.delete()
        db.session.commit()

        flash(f"Deleted {count} processed submissions older than 40 days.", "success")
        return redirect(url_for('debug_page'))



    @app.route("/fetch_invite_submissions")
    def fetch_invite_submissions():
        import requests
        import json
        from datetime import datetime

        INVITE_FORM_ID = "253599154628066"
        API_KEY = os.getenv("JOTFORM_API_KEY", "")

        url = f"https://api.jotform.com/form/{INVITE_FORM_ID}/submissions?apiKey={API_KEY}"
        r = requests.get(url)
        data = r.json()

        submissions = data.get("content", [])
        count = 0

        latest = (
            db.session.query(IncomingSubmission)
            .filter(IncomingSubmission.form_id == INVITE_FORM_ID)
            .order_by(IncomingSubmission.received_at.desc())
            .first()
        )
        latest_ts = latest.received_at if latest else None

        for s in submissions:
            submission_id = str(s.get("id"))
            form_id = s.get("form_id")

            # PRIMARY DEDUPE: submission_id
            existing = (
                db.session.query(IncomingSubmission)
                .filter_by(submission_id=submission_id)
                .first()
            )
            if existing:
                continue

            # Compute hash (secondary dedupe only)
            payload_str = json.dumps(s, sort_keys=True)
            payload_hash = hashlib.sha256(payload_str.encode()).hexdigest()

            created_ts = s.get("created_at") or s.get("updated_at")
            try:
                submission_dt = datetime.utcfromtimestamp(int(created_ts))
            except Exception:
                submission_dt = datetime.utcnow()

            if latest_ts and submission_dt <= latest_ts:
                continue

            new_sub = IncomingSubmission(
                submission_id=submission_id,
                form_id=form_id,
                raw_payload=json.dumps(s),
                received_at=submission_dt,
                processed=False,
                needs_client_match=False,
                unique_hash=payload_hash
            )

            db.session.add(new_sub)
            db.session.commit()

            invite = match_submission_to_invite(new_sub)
            if invite:
                invite.submission_id = new_sub.id
                invite.status = "form_received"
                db.session.commit()

            count += 1

        return render_template("fetch_invite_submissions.html", count=count)


    @app.route("/process_all_invites")
    def process_all_invites():
        # Only submissions that are not processed
        pending = (
            db.session.query(IncomingSubmission)
            .filter_by(processed=False)
            .all()
        )

        count = 0

        for sub in pending:
            # Try to match to an invite
            invite = match_submission_to_invite(sub)
            if not invite:
                continue

            # Only process invites that are awaiting_form
            if invite.status != "process":
                continue

            # Mark invite as form_received
            invite.status = "form_received"
            db.session.commit()

            # Process the submission (creates lesson, attaches riders, etc.)
            process_submission(sub)
            count += 1

        return f"Processed {count} invite submissions."

    @app.route("/clear_processed_invites")
    def clear_processed_invites():
        processed = (
            db.session.query(IncomingSubmission)
            .filter_by(processed=True)
            .all()
        )

        count = len(processed)

        for sub in processed:
            db.session.delete(sub)

        db.session.commit()

        return f"Cleared {count} processed invite submissions."


    @app.route("/send_negative_balance_sms", methods=["POST"])
    def send_negative_balance_sms():
        client_name = request.form.get("client_name")
        guardian = request.form.get("guardian") or ""
        mobile = request.form.get("mobile")
        balance = request.form.get("balance")
        tc_balance = request.form.get("tc_balance") or ""
        test_mode = request.form.get("test_mode") == "true"

        # ⭐ NEW — get the EXACT preview text the user saw
        message = request.form.get("final_message")

        # Test mode: redirect to admin number
        if test_mode:
            to_number = current_app.config.get("SMS_ADMIN_NUMBER", "<YOUR_ADMIN_MOBILE>")
        else:
            to_number = mobile

        sender_id = "+61417704671"

        payload = {
            "messages": [
                {
                    "source": "python",
                    "from": sender_id,
                    "body": message,
                    "to": to_number
                }
            ]
        }

        username = current_app.config.get("CLICKSEND_USERNAME", "")
        api_key = current_app.config.get("CLICKSEND_API_KEY", "")

        try:
            response = requests.post(
                "https://rest.clicksend.com/v3/sms/send",
                json=payload,
                auth=(username, api_key),
                timeout=10
            )
            ok = response.status_code == 200
        except Exception as e:
            ok = False
            response_text = str(e)
        else:
            response_text = response.text

        # Log to file
        log_line = (
            f"{datetime.now(ZoneInfo('Australia/Brisbane')).isoformat()} | "
            f"{'TEST MODE' if test_mode else 'LIVE'} | "
            f"{client_name} | {mobile} | balance={balance} | tc_balance={tc_balance} | "
            f"status={'OK' if ok else 'ERROR'} | {response_text}\n"
        )
        os.makedirs("logs", exist_ok=True)
        with open("logs/sms_log.txt", "a", encoding="utf-8") as f:
            f.write(log_line)

        # Save SMS to database if sent OK
        if ok:
            client = Client.query.filter_by(full_name=client_name).first()
            if client:
                entry = SmsLog(
                    client_id=client.client_id,
                    guardian=guardian,
                    mobile=mobile,
                    message_body=message,
                    sent_at=datetime.now(ZoneInfo("Australia/Brisbane"))
                )
                db.session.add(entry)
                db.session.commit()

            return jsonify({"status": "ok"})
        else:
            return jsonify({"status": "error", "detail": response_text}), 500

    @app.route("/minus_balances")
    def minus_balances_page():
        # Ensure balances are fresh
        recalc_all_lessons()

        clients = db.session.query(Client).order_by(Client.full_name).all()
        final_output = []

        for c in clients:
            latest = (
                db.session.query(Lesson)
                .filter(Lesson.client == c.full_name)
                .filter(Lesson.lesson_date <= date.today())
                .order_by(Lesson.lesson_date.desc())
                .first()
            )

            if not latest or latest.balance is None or latest.balance >= 0:
                continue

            history = (
                db.session.query(Lesson)
                .filter(Lesson.client == c.full_name)
                .filter(Lesson.lesson_date <= date.today())
                .order_by(Lesson.lesson_date.desc())
                .all()
            )

            trimmed = []
            for h in history:
                trimmed.append(h)
                if h.balance == 0:
                    break

            # ⭐ NEW: Load last SMS sent to this client
            last_sms = (
                db.session.query(SmsLog)
                .filter_by(client_id=c.client_id)
                .order_by(SmsLog.sent_at.desc())
                .first()
            )

            final_output.append({
                "client": c.full_name,
                "guardian": c.guardian_name or "",
                "mobile": c.mobile,
                "current_balance": latest.balance,
                "lessons": trimmed,

                # ⭐ NEW FIELDS
                "last_sms": last_sms.message_body if last_sms else None,
                "last_sms_time": last_sms.sent_at if last_sms else None
            })
        
        return render_template("minus_balances.html", rows=final_output, date=date)



    @app.route("/manage_teachers")
    def manage_teachers_page():
        rows = db.session.query(Teacher).order_by(Teacher.teacher).all()
        return render_template("manage_teachers.html", teachers=rows)

    @app.route("/teachers", methods=["POST"])
    def add_teacher():
        name = request.form.get("teacher", "").strip()
        if name:
            db.session.add(Teacher(teacher=name))
            db.session.commit()
        return redirect(url_for("manage_teachers_page"))

    @app.route("/teachers/delete/<int:tid>", methods=["POST"])
    def delete_teacher(tid):
        row = db.session.query(Teacher).get(tid)
        if row:
            db.session.delete(row)
            db.session.commit()
        return redirect(url_for("manage_teachers_page"))


    @app.route("/manage_horses")
    def manage_horses_page():
        rows = db.session.query(Horse).order_by(Horse.orderpdk).all()
        return render_template("manage_horses.html", horses=rows)

    @app.route("/horses", methods=["POST"])
    def add_horse():
        horse = request.form.get("horse", "").strip()
        orderpdk = request.form.get("orderpdk", "").strip()
        age = request.form.get("age")
        sex = request.form.get("sex", "").strip()

        db.session.add(Horse(horse=horse, orderpdk=orderpdk, age=age, sex=sex))
        db.session.commit()
        return redirect(url_for("manage_horses_page"))



    @app.route('/course_reference')
    def course_reference():
        courses = CourseReference.query.order_by(CourseReference.sort_order).all()
        times = Time.query.order_by(Time.timerange).all()

        pricing_rows = GroupPricing.query.all()
        pricing = {p.group_priv: p for p in pricing_rows}

        return render_template(
            'course_reference.html',
            courses=courses,
            times=times,
            pricing=pricing
        )


    @app.route('/update_course_field', methods=['POST'])
    def update_course_field():
        data = request.get_json()
        cid = data.get('id')
        field = data.get('field')
        value = data.get('value')

        print("UPDATE:", field, value)

        course = CourseReference.query.get(cid)
        if not course:
            return jsonify(success=False, error="Course not found")

        allowed = {
            "course_code",
            "display_label",
            "day_of_week",
            "timerange",
            "lesson_type",
            "group_priv",
            "active"
        }

        if field not in allowed:
            return jsonify(success=False, error="Invalid field")

        if field == "course_code":
            existing = CourseReference.query.filter_by(course_code=value).first()
            if existing and existing.id != course.id:
                return jsonify(success=False, error="Course code already exists")

        try:
            if field == "active":
                setattr(course, field, bool(int(value)))
            else:
                setattr(course, field, value)

            db.session.flush()

            course.sort_order = compute_sort_order(
                course.day_of_week,
                course.timerange
            )

            db.session.commit()
            renumber_all_courses()
            return jsonify(success=True)

        except Exception as e:
            db.session.rollback()
            return jsonify(success=False, error=str(e))



    @app.route('/add_course_reference')
    def add_course_reference():
        new_course = CourseReference(
            course_code="NEW",
            display_label="New Course",
            day_of_week="Monday",
            timerange="07:00 - 08:00",
            lesson_type="Arena",
            group_priv="CC",
            active=True
        )

        new_course.sort_order = compute_sort_order(
            new_course.day_of_week,
            new_course.timerange
        )

        db.session.add(new_course)
        db.session.commit()
        renumber_all_courses()
        return redirect(url_for('course_reference'))


    @app.route('/delete_course_reference', methods=['POST'])
    def delete_course_reference():
        course_id = request.json.get('id')
        course = CourseReference.query.get(course_id)
        if course:
            db.session.delete(course)
            db.session.commit()
            renumber_all_courses()
            return jsonify(success=True)
        return jsonify(success=False, error="Course not found")

    @app.route('/fix_sort_orders', methods=['POST'])
    def fix_sort_orders():
        courses = CourseReference.query.all()
        for c in courses:
            c.sort_order = compute_sort_order(c.day_of_week, c.timerange)
        db.session.commit()
        renumber_all_courses()
        return ("", 204)


    @app.route("/horses/delete/<int:hid>", methods=["POST"])
    def delete_horse(hid):
        row = db.session.query(Horse).get(hid)
        if row:
            db.session.delete(row)
            db.session.commit()
        return redirect(url_for("manage_horses_page"))

    @app.route("/horses/update/<int:hid>", methods=["POST"])
    def update_horse_inline(hid):
        row = db.session.query(Horse).get(hid)
        if row:
            row.horse = request.form.get("horse")
            row.orderpdk = request.form.get("orderpdk")
            row.age = request.form.get("age")
            row.sex = request.form.get("sex")
            db.session.commit()
        return redirect(url_for("manage_horses_page"))

    @app.route("/horses/edit/<int:hid>")
    def edit_horse_page(hid):
        row = db.session.query(Horse).get(hid)
        return render_template("edit_horse.html", horse=row)

    @app.route("/horses/edit/<int:hid>", methods=["POST"])
    def edit_horse_save(hid):
        row = db.session.query(Horse).get(hid)
        if row:
            row.horse = request.form.get("horse")
            row.orderpdk = request.form.get("orderpdk")
            row.age = request.form.get("age")
            row.sex = request.form.get("sex")
            db.session.commit()
        return redirect(url_for("manage_horses_page"))


    @app.route("/manage_blockout_dates_page")
    def manage_blockout_dates_page():
        blockouts = BlockoutDate.query.order_by(BlockoutDate.block_date).all()
        times = Time.query.order_by(Time.timerange).all()
        return render_template(
            "blockout_dates.html",
            blockouts=blockouts,
            times=times
        )

    @app.route("/manage_blockout_ranges")
    def manage_blockout_ranges_page():
        return render_template("blockout_ranges.html")

    @app.route('/blockout_dates', methods=['GET','POST'])
    def manage_blockout_dates():
        if request.method == 'POST':
            raw_date = request.form.get('block_date')
            reason = request.form.get('reason','').strip()
            block_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
            db.session.add(BlockoutDate(block_date=block_date, reason=reason))
            db.session.commit()
            flash(f"Blockout added: {block_date} ({reason})", "success")
        rows = db.session.query(BlockoutDate).order_by(BlockoutDate.block_date).all()
        return render_template('blockout_dates.html', blockouts=rows)

    @app.route('/blockout_dates/delete/<int:row_id>', methods=['POST'])
    def delete_blockout_date(row_id):
        row = db.session.query(BlockoutDate).get(row_id)
        if row:
            db.session.delete(row)
            db.session.commit()
            flash("Blockout removed.", "success")
        return redirect(url_for('manage_blockout_dates'))

    @app.route('/blockout_ranges', methods=['GET','POST'])
    def manage_blockout_ranges():
        if request.method == 'POST':
            start_raw = request.form.get('start_date')
            end_raw = request.form.get('end_date')
            reason = request.form.get('reason','').strip()
            start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_raw, '%Y-%m-%d').date()
            db.session.add(BlockoutRange(start_date=start_date, end_date=end_date, reason=reason))
            db.session.commit()
            flash(f"Blockout range added: {start_date} → {end_date} ({reason})", "success")
        rows = db.session.query(BlockoutRange).order_by(BlockoutRange.start_date).all()
        return render_template('blockout_ranges.html', blockouts=rows)

    @app.route('/blockout_ranges/delete/<int:row_id>', methods=['POST'])
    def delete_blockout_range(row_id):
        row = db.session.query(BlockoutRange).get(row_id)
        if row:
            db.session.delete(row)
            db.session.commit()
            flash("Blockout range removed.", "success")
        return redirect(url_for('manage_blockout_ranges'))

    @app.route("/api/check_blockouts", methods=["POST"])
    def api_check_blockouts():
        from datetime import datetime, timedelta, date

        data = request.get_json() or {}

        start_date_str = data.get("start_date")
        freq = (data.get("freq") or "S").strip()  # "S" (single), "W" (weekly), "F" (fortnightly)

        if not start_date_str:
            return {"blocked": False, "blocked_dates": []}

        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_of_year = date(start_date.year, 12, 24)

        # Build candidate dates using same pattern as your lesson recurrence
        if freq == "S":
            dates = [start_date]
        elif freq == "F":
            dates = []
            d = start_date
            while d <= end_of_year:
                dates.append(d)
                d += timedelta(days=14)
        elif freq == "W":
            dates = []
            d = start_date
            while d <= end_of_year:
                dates.append(d)
                d += timedelta(days=7)
        else:
            dates = [start_date]

        # Load blockouts from your tables
        block_dates = {
            b.block_date: (b.reason or "")
            for b in db.session.query(BlockoutDate).all()
        }
        block_ranges = list(db.session.query(BlockoutRange).all())

        def is_blocked(d):
            if d in block_dates:
                return True, (block_dates[d] or "Single‑day blockout")
            for r in block_ranges:
                if r.start_date <= d <= r.end_date:
                    reason = r.reason or f"Range {r.start_date} → {r.end_date}"
                    return True, reason
            return False, ""

        blocked_list = []
        for d in dates:
            blocked, reason = is_blocked(d)
            if blocked:
                blocked_list.append({
                    "date": d.strftime("%Y-%m-%d"),
                    "reason": reason,
                })

        return {
            "blocked": bool(blocked_list),
            "blocked_dates": blocked_list,
        }


    @app.route("/new_lesson", methods=["POST"])
    def new_lesson():
        import re
        from datetime import datetime, timedelta, date
        print("DEBUG FORM:", dict(request.form))

        lesson_date_str = request.form.get("date")
        lesson_id       = request.form.get("lesson_id")
        start_raw       = request.form.get("start", "")
        end_raw         = request.form.get("end", "")
        price_pl        = request.form.get("price_pl", "$0.00")
        lesson_type     = request.form.get("lesson_type", "Arena")
        group_priv      = request.form.get("group_priv", "")
        freq            = request.form.get("recurrence_type", "").strip()
        if freq not in ["S", "W", "F"]:
            freq = "S"

        respect_blockouts = request.form.get("respect_blockouts", "yes")

        time_frame_raw = request.form.get("time_frame")

        if time_frame_raw:
            try:
                start_raw, end_raw = [s.strip() for s in time_frame_raw.split("-")]
            except Exception:
                return "Invalid time_frame format", 400

        client_mode   = request.form.get("client_mode", "existing")
        client_id_raw = request.form.get("client_id")
        client_name   = request.form.get("client_name", "").strip()
        client_phone  = request.form.get("client_phone", "").strip()
        horse         = request.form.get("horse", "").strip()

        weight_kg = request.form.get("weight_kg")
        height_cm = request.form.get("height_cm")
        notes     = request.form.get("notes")

        print("DEBUG FORM:", dict(request.form))

        def _parse_times(start_label, end_form):
            if end_form and str(end_form).strip():
                start_token = (str(start_label).split('-')[0]).strip() if start_label else ''
                return start_token, str(end_form).strip()

            m = re.match(r'^\s*(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})', str(start_label or ''))
            if m:
                return m.group(1), m.group(2)

            return (str(start_label or '').strip(), str(end_form or '').strip())

        def _clean_time_range(time_range):
            parts = [p.strip() for p in str(time_range or "").split('-') if p.strip()]
            if len(parts) >= 2:
                return f"{parts[0]} - {parts[-1]}"
            return str(time_range or "").strip()

        # ---------------------------------------------------------
        # EDIT EXISTING LESSON
        # ---------------------------------------------------------
        if lesson_id:
            db.session.rollback()

            lesson = Lesson.query.get(int(lesson_id))
            if not lesson:
                return redirect(url_for("lessons_by_date", date=lesson_date_str))

            end_effective   = end_raw.strip()
            start_effective = start_raw.strip()

            if not end_effective:
                existing_tf = lesson.time_frame or ""
                m = re.match(r'^\s*(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})', existing_tf)
                if m:
                    if not start_effective:
                        start_effective = m.group(1)
                    end_effective = m.group(2)

            start_token, end_token = _parse_times(start_effective or lesson.time_frame, end_effective)
            time_range = _clean_time_range(f"{start_token} - {end_token}")
            time_frame = time_range

            existing_range = db.session.query(Time).filter_by(timerange=time_range).first()
            if not existing_range:
                db.session.add(Time(timerange=time_range))
                db.session.commit()

            if client_mode == "new":
                # Only requirement: a name
                if client_name.strip():
                    new_client = Client(
                        full_name = client_name.strip(),
                        mobile    = (client_phone or "").strip() or None,
                        weight_kg = weight_kg or None,
                        height_cm = height_cm or None,
                        notes     = notes or "",
                        age       = request.form.get("age") or None
                    )
                    db.session.add(new_client)
                    db.session.flush()
                    client_id = new_client.client_id
                else:
                    # No name = cannot create client
                    client_id = None

            lesson.lesson_date = datetime.strptime(lesson_date_str, "%Y-%m-%d").date() if lesson_date_str else lesson.lesson_date
            lesson.time_frame  = time_frame
            lesson.price_pl    = parse_money(price_pl)
            lesson.lesson_type = lesson_type
            lesson.group_priv  = group_priv
            if freq:
                lesson.freq = freq
            lesson.horse       = horse

            db.session.commit()

            # ⭐ LOG: EDIT
            log_admin_action(
                f"Lesson updated for {lesson.client} on {lesson.lesson_date} "
                f"({lesson.time_frame}, {lesson.lesson_type} {lesson.group_priv})"
            )

            try:
                if freq in ["W", "F"] and group_priv not in ["J", "CT", "CC", "D"]:
                    future_lessons = Lesson.query.filter(
                        Lesson.client == lesson.client,
                        Lesson.lesson_date > lesson.lesson_date
                    ).order_by(Lesson.lesson_date).all()

                    for fl in future_lessons:
                        fl.horse = horse

                    db.session.commit()
            except Exception as e:
                print("[DEBUG] Cascade error:", e)

            return redirect(url_for("lessons_by_date", date=lesson_date_str))

        # ---------------------------------------------------------
        # CREATE NEW LESSON(S)
        # ---------------------------------------------------------
        start_token, end_token = _parse_times(start_raw, end_raw)
        time_range = _clean_time_range(f"{start_token} - {end_token}")

        if lesson_type not in ["Payment", "Voucher CR", "Camp"]:
            if not start_token or not end_token or "-" not in time_range:
                print("[DEBUG] Missing or invalid time range for NEW lesson, rejecting.")
                return redirect(url_for("lessons_by_date", date=lesson_date_str))
        else:
            start_token = ""
            end_token = ""
            time_range = ""
            time_frame = ""

        time_frame = time_range

        existing_range = db.session.query(Time).filter_by(timerange=time_range).first()
        if not existing_range:
            db.session.add(Time(timerange=time_range))
            db.session.commit()

        if client_mode == "new":
            # Only requirement: a name
            if client_name.strip():
                new_client = Client(
                    full_name = client_name.strip(),
                    mobile    = (client_phone or "").strip() or None,
                    weight_kg = weight_kg or None,
                    height_cm = height_cm or None,
                    notes     = notes or "",
                    age       = request.form.get("age") or None
                )
                db.session.add(new_client)
                db.session.flush()
                client_id = new_client.client_id
            else:
                client_id = None
        else:
            client_id = int(client_id_raw) if client_id_raw and str(client_id_raw).isdigit() else None

        if client_id:
            client_obj = Client.query.get(client_id)
            canonical_name = client_obj.full_name if client_obj else client_name
        else:
            canonical_name = client_name

        block_dates  = {b.block_date for b in db.session.query(BlockoutDate).all()}
        block_ranges = [(r.start_date, r.end_date) for r in db.session.query(BlockoutRange).all()]

        def is_blocked(d):
            if d in block_dates:
                return True
            for start_d, end_d in block_ranges:
                if start_d <= d <= end_d:
                    return True
            return False

        start_date = datetime.strptime(lesson_date_str, "%Y-%m-%d").date()
        end_of_year = date(start_date.year, 12, 24)

        if freq == "S":
            dates = [start_date]
        elif freq == "F":
            dates = []
            d = start_date
            while d <= end_of_year:
                dates.append(d)
                d += timedelta(days=14)
        elif freq == "W":
            dates = []
            d = start_date
            while d <= end_of_year:
                dates.append(d)
                d += timedelta(days=7)
        else:
            dates = [start_date]

        added = 0
        for d in dates:
            blocked = is_blocked(d)
            print(f"[DEBUG] candidate={d} blocked={blocked} respect_blockouts={respect_blockouts}")

            if respect_blockouts == "yes" and blocked:
                # ⭐ LOG: SKIPPED BLOCKOUT
                log_admin_action(
                    f"Lesson skipped for {canonical_name} on {d} due to blockout"
                )
                continue

            payment_raw = request.form.get("payment", "").replace("$", "").strip()
            try:
                payment_val = float(payment_raw) if payment_raw else 0
            except:
                payment_val = 0

            lesson = Lesson(
                lesson_date=d,
                time_frame=time_frame,
                price_pl=parse_money(price_pl),
                payment=payment_val,
                lesson_type=lesson_type,
                group_priv=group_priv,
                freq=freq,
                client=canonical_name,
                horse=horse,
            )
            db.session.add(lesson)
            added += 1

            # ⭐ LOG: NEW LESSON
            log_admin_action(
                f"Lesson created for {canonical_name} on {d} "
                f"({time_frame}, {lesson_type} {group_priv})"
            )

        db.session.commit()
        print(f"[DEBUG] commit done, total lessons added={added}")

        if request.headers.get("X-Requested-With") == "fetch":
            return "OK", 200

        return redirect(url_for("lessons_by_date", date=lesson_date_str))


    @app.route("/save_all_lessons", methods=["POST"])
    def save_all_lessons():
        data = request.get_json()
        print("RAW JSON RECEIVED:", data)

        date = data.get("date")
        lessons = data.get("lessons", [])
        teacher_blocks = data.get("teacher_blocks", [])

        if not date:
            return {"status": "error", "message": "Missing date"}, 400

        try:
            # Always start clean

            # --------------------------------------------------------
            # SAVE LESSONS
            # --------------------------------------------------------
            for item in lessons:
                lesson_id = item.get("lesson_id")
                if not lesson_id:
                    continue  # safety: never create lessons

                lesson = Lesson.query.get(lesson_id)
                if not lesson:
                    continue

                # ---------------------------
                # SANITIZE MONEY FIELDS
                # ---------------------------
                import re
                payment_raw = re.sub(r"[^\d.\-]", "", item.get("payment") or "")
                price_raw   = re.sub(r"[^\d.\-]", "", item.get("price_pl") or "")

                # PAYMENT (FIX ZERO BUG)
                if payment_raw is not None and payment_raw.strip() != "":
                    try:
                        lesson.payment = float(payment_raw)
                    except ValueError:
                        pass
                else:
                    lesson.payment = 0.0

                # PRICE PER LESSON (FIX ZERO BUG)
                if price_raw is not None and price_raw.strip() != "":
                    try:
                        lesson.price_pl = float(price_raw)
                    except ValueError:
                        pass
                else:
                    lesson.price_pl = 0.0

                # ---------------------------
                # NORMAL FIELDS
                # ---------------------------
                horse_val = item.get("horse")
                print("DEBUG HORSE RECEIVED:", repr(horse_val), "for lesson", lesson_id)

                if horse_val in ("", None, "None"):
                    horse_val = ""

                lesson.horse = horse_val
                lesson.attendance = (item.get("attendance") or "")

                gp_val = item.get("group_priv")
                if gp_val not in ("", None, "None"):
                    lesson.group_priv = gp_val

                lt_val = item.get("lesson_type")
                if lt_val not in ("", None, "None"):
                    lesson.lesson_type = lt_val

                teacher_val = item.get("teacher")
                if teacher_val not in ("", None, "None"):
                    lesson.teacher = teacher_val

                # ---------------------------
                # CLIENT NOTES
                # ---------------------------
                notes_val = item.get("notes")
                if notes_val not in ("", None, "None"):
                    if lesson.client:
                        client_obj = Client.query.filter_by(full_name=lesson.client).first()
                        if client_obj:
                            client_obj.notes = notes_val

                # ---------------------------
                # TIME REQUIREMENTS (Payment / Voucher CR EXEMPT)
                # ---------------------------
                lt_val = item.get("lesson_type")

                # ---------------------------
                # TIME REQUIREMENTS (Payment / Voucher CR / Camp EXEMPT)
                # ---------------------------
                lt_val = item.get("lesson_type")

                if lt_val in ["Payment", "Voucher CR", "Camp"]:
                    lesson.start = ""
                    lesson.end = ""
                else:
                    lesson.start = item.get("start")
                    lesson.end   = item.get("end")

                # ---------------------------
                # CARRY-FORWARD ENGINE
                # ---------------------------
                if lesson.freq in ("W", "F") and lesson.horse not in ("", None, "None"):

                    future_lessons = (
                        Lesson.query
                        .filter(
                            Lesson.client == lesson.client,
                            Lesson.lesson_date > lesson.lesson_date
                        )
                        .order_by(Lesson.lesson_date)
                        .all()
                    )

                    step = 7 if lesson.freq == "W" else 14

                    for fl in future_lessons:
                        delta = (fl.lesson_date - lesson.lesson_date).days

                        if delta % step == 0:
                            old_att = fl.attendance
                            fl.horse = lesson.horse
                            fl.attendance = old_att

            # --------------------------------------------------------
            # SAVE TEACHER BLOCKS
            # --------------------------------------------------------
            incoming_ids = set()

            for tb in teacher_blocks:
                tb_id = tb.get("id")
                block_key = tb.get("block_key") or ""
                horse = tb.get("horse") or ""
                teacher_name = tb.get("teacher_name") or ""
                notes = tb.get("notes") or ""

                if tb_id:
                    incoming_ids.add(int(tb_id))
                    obj = TeacherBlock.query.get(int(tb_id))
                    if obj:
                        obj.block_key = block_key
                        obj.horse = horse
                        obj.teacher_name = teacher_name
                        obj.notes = notes
                else:
                    new_tb = TeacherBlock(
                        block_key=block_key,
                        horse=horse,
                        teacher_name=teacher_name,
                        notes=notes,
                        date=date
                    )
                    db.session.add(new_tb)
                    db.session.flush()
                    incoming_ids.add(new_tb.id)

            existing = TeacherBlock.query.filter_by(date=date).all()
            for obj in existing:
                if obj.id not in incoming_ids:
                    db.session.delete(obj)

            # --------------------------------------------------------
            # FINAL COMMIT
            # --------------------------------------------------------

            db.session.commit()
            print("✅ COMMIT OK")
            return {"status": "ok"}, 200

        except Exception as e:
            db.session.rollback()
            print("🔥 ERROR:", e)
            return {"status": "error", "message": str(e)}, 500





    @app.route('/delete_client_lessons/<int:client_id>', methods=['POST'])
    def delete_client_lessons(client_id):
        mode = (request.form.get('mode') or '').strip()
        group_priv = (request.form.get('group_priv') or '').strip()
        cutoff_str = (request.form.get('cutoff_date') or '').strip()

        # Default to today if no date provided
        cutoff_date = date.today()
        if cutoff_str:
            try:
                cutoff_date = datetime.strptime(cutoff_str, "%Y-%m-%d").date()
            except ValueError:
                flash("Invalid cutoff date format, using today instead.", "warning")

        client = db.session.query(Client).get(client_id)
        if not client:
            flash("Client not found.", "danger")
            return redirect(url_for('debug_page'))

        if mode == "all":
            db.session.query(Lesson).filter(
                func.lower(func.trim(Lesson.client)) == func.lower(func.trim(client.full_name)),
                Lesson.lesson_date >= cutoff_date
            ).delete(synchronize_session=False)

        elif mode == "group_priv" and group_priv:
            db.session.query(Lesson).filter(
                func.lower(func.trim(Lesson.client)) == func.lower(func.trim(client.full_name)),
                Lesson.lesson_date >= cutoff_date,
                func.lower(func.trim(Lesson.group_priv)) == group_priv.lower()
            ).delete(synchronize_session=False)

        else:
            flash("Invalid delete request.", "danger")
            return redirect(url_for('client_view', client=client.full_name))

        db.session.commit()
        flash(f"Delete executed for {client.full_name} from {cutoff_date}", "success")
        log_admin_action(f"Delete executed for {client.full_name} from {cutoff_date}", user="system")
        return redirect(url_for('client_view', client=client.full_name))





    @app.route('/admin/logs')
    def view_admin_logs():
        log_dir = LOG_ROOT
        logs = []
        try:
            # Collect all admin log files
            for fname in sorted(os.listdir(log_dir), reverse=True):
                if fname.startswith("admin_actions_") and fname.endswith(".log"):
                    path = os.path.join(log_dir, fname)
                    with open(path, "r", encoding="utf-8") as f:
                        logs.append((fname, f.read()))
        except Exception as e:
            logs.append(("Error", f"Failed to read logs: {e}"))
        return render_template("admin_logs.html", logs=logs)

    @app.route('/admin/disclaimer_logs')
    @login_required
    def view_disclaimer_logs():
        log_dir = "/var/log/cherbonapp"
        logs = []

        try:
            for fname in sorted(os.listdir(log_dir), reverse=True):
                if fname.startswith("disclaimers_") and fname.endswith(".log"):
                    path = os.path.join(log_dir, fname)
                    with open(path, "r", encoding="utf-8") as f:
                        logs.append((fname, f.read()))
        except Exception as e:
            logs.append(("Error", f"Failed to read disclaimer logs: {e}"))

        return render_template("admin_logs.html", logs=logs)




    @app.route('/admin/reindex', methods=['POST'])
    def reindex_tables():
        try:
            tables = ['lessons', 'clients', 'teacher_time']
            for table in tables:
                db.session.execute(text(f'REINDEX TABLE {table};'))
            db.session.commit()
            flash('Reindexing completed successfully.', 'success')
            log_admin_action(f"Reindexed tables: {', '.join(tables)}", user="system")
        except Exception as e:
            db.session.rollback()
            flash(f'Reindexing failed: {str(e)}', 'danger')
            log_admin_action(f"Reindex failed: {str(e)}", user="system")
        return redirect(url_for('debug_page'))





    @app.route('/trailride_enquiries', methods=['GET'])
    def trailride_enquiries():
        page = request.args.get('page', 1, type=int)

        pagination = (TrailRideSubmission.query
                      .filter_by(processed=False, ignored=False)
                      .order_by(TrailRideSubmission.received_at.desc())
                      .paginate(page=page, per_page=20, error_out=False))

        display_rows = []

        # Filters
        filter_name = request.args.get('name', '').strip().lower()
        filter_phone = request.args.get('phone', '').strip()
        filter_email = request.args.get('email', '').strip().lower()
        filter_match = request.args.get('match', '')

        needs_commit = False

        for e in pagination.items:
            payload = e.raw_payload
            riders = extract_riders_from_submission(payload)
            contact = get_main_contact_fields(payload)

            # NORMALISE PHONE + EMAIL
            phone = contact.get("phone")
            email = contact.get("email")

            if isinstance(phone, dict):
                phone = phone.get("full") or phone.get("value") or phone.get("text") or ""
            if isinstance(email, dict):
                email = email.get("value") or email.get("text") or email.get("full") or ""

            if phone:
                phone = str(phone).replace("-", "").replace(" ", "").strip()

            phone = phone or ""
            email = email or ""

            contact["phone"] = phone
            contact["email"] = email

            # SAFE MAIN NAME EXTRACTION
            if riders and isinstance(riders, list) and len(riders) > 0 and riders[0].get("name"):
                main_name = riders[0]["name"]
            else:
                main_name = "(no riders)"

            rider_count = len(riders)

            # MATCHING LOGIC
            matches = match_existing_client(
                name=main_name,
                phone=phone,
                email=email
            )

            has_match = len(matches) > 0
            multiple_matches = len(matches) > 1

            if e.needs_client_match != multiple_matches:
                e.needs_client_match = multiple_matches
                needs_commit = True

            # FILTERS
            if filter_name and filter_name not in main_name.lower():
                continue
            if filter_phone and filter_phone not in phone:
                continue
            if filter_email and filter_email not in email.lower():
                continue
            if filter_match == "yes" and not has_match:
                continue
            if filter_match == "no" and has_match:
                continue
            if filter_match == "multi" and not multiple_matches:
                continue

            display_rows.append({
                "id": e.id,
                "main_name": main_name,
                "phone": phone,
                "email": email,
                "rider_count": rider_count,
                "created_at": e.received_at.strftime("%d %b %Y %I:%M %p"),
                "riders": riders,
                "has_match": has_match,
                "multiple_matches": multiple_matches,
                "match_count": len(matches)
            })

        if needs_commit:
            db.session.commit()

        # Ensure newest first after filtering
        display_rows.sort(
            key=lambda x: datetime.strptime(x["created_at"], "%d %b %Y %I:%M %p"),
            reverse=True
        )

        return render_template(
            'trailride_enquiries.html',
            enquiries=display_rows,
            pagination=pagination,
            times = Time.query.order_by(Time.timerange).all()
        )


    @app.route('/trailride_enquiries/process', methods=['POST'])
    def trailride_enquiries_process():
        enquiry_id = request.form.get("process_enquiry")

        if not enquiry_id:
            flash("No enquiry selected.", "warning")
            return redirect(url_for('trailride_enquiries'))

        # Safe int converter
        def to_int(val):
            try:
                return int(val)
            except:
                return None

        enquiry = TrailRideSubmission.query.get(enquiry_id)
        if not enquiry or enquiry.processed:
            flash("Enquiry not found or already processed.", "warning")
            return redirect(url_for('trailride_enquiries'))

        # ---------------------------------------------------------
        # IGNORE BRANCH — permanently hide this enquiry
        # ---------------------------------------------------------
        if request.form.get(f"ignore_{enquiry_id}") == "1":
            enquiry.ignored = True
            enquiry.processed = True
            enquiry.processed_at = datetime.utcnow()
            db.session.commit()
            flash("Enquiry ignored.", "info")
            return redirect(url_for('trailride_enquiries'))

        payload = enquiry.raw_payload
        riders = extract_riders_from_submission(payload)
        contact = get_main_contact_fields(payload)

        # ---------------------------------------------------------
        # NORMALISE PHONE + EMAIL
        # ---------------------------------------------------------
        phone = contact.get("phone")
        email = contact.get("email")

        if isinstance(phone, dict):
            phone = phone.get("full") or phone.get("value") or phone.get("text") or ""
        if isinstance(email, dict):
            email = email.get("value") or email.get("text") or email.get("full") or ""

        if phone:
            phone = str(phone).replace("-", "").replace(" ", "").strip()

        phone = phone or ""
        email = email or ""

        contact["phone"] = phone
        contact["email"] = email

        # ---------------------------------------------------------
        # SELECTED RIDERS FROM CHECKBOXES
        # ---------------------------------------------------------
        selected_riders = []
        for i in range(1, 15):
            if request.form.get(f"process_rider_{enquiry_id}_{i}"):
                selected_riders.append(i)

        if not selected_riders:
            flash("No riders selected.", "warning")
            return redirect(url_for('trailride_enquiries'))

        # ---------------------------------------------------------
        # BOOKING FIELDS FROM FORM (per enquiry)
        # ---------------------------------------------------------
        booking_date = request.form.get(f"booking_date_{enquiry_id}") or ""
        booking_time = request.form.get(f"booking_time_{enquiry_id}") or ""
        lesson_type = request.form.get(f"lesson_type_{enquiry_id}") or "Trail Ride"
        price_per_rider = request.form.get(f"price_per_rider_{enquiry_id}") or "0"
        payment_per_rider = request.form.get(f"payment_per_rider_{enquiry_id}") or "0"
        group_priv = request.form.get(f"group_priv_{enquiry_id}") or "P"

        # ---------------------------------------------------------
        # VALIDATE DATE + TIME (STAFF-PROOF)
        # ---------------------------------------------------------
        if not booking_date or not booking_time:
            flash("Please select BOTH a date and a time before processing.", "danger")
            return redirect(url_for('trailride_enquiries'))

        # ---------------------------------------------------------
        # CREATE LESSON ROWS + ENSURE CLIENT EXISTS
        # ---------------------------------------------------------
        for idx in selected_riders:
            r = riders[idx - 1] if idx - 1 < len(riders) else None
            if not r:
                continue

            rider_name = r.get("name") or ""
            rider_name = smart_proper_name(r.get("name") or "")
            rider_mobile = phone
            rider_email = email
            rider_age = r.get("age")
            rider_height = r.get("height_cm")
            rider_weight = r.get("weight_kg")


            # -----------------------------------------------------
            # ENSURE CLIENT EXISTS
            # -----------------------------------------------------
            existing_client = Client.query.filter(
                func.lower(Client.full_name) == rider_name.lower()
            ).first()

            if not existing_client:
                client = Client(
                    full_name=rider_name,
                    mobile=rider_mobile,
                    email_primary=rider_email,
                    age=rider_age,
                    height_cm=rider_height,
                    weight_kg=rider_weight
                )
                db.session.add(client)
                db.session.flush()
                is_new_client = True
            else:
                client = existing_client
                is_new_client = False

            # -----------------------------------------------------
            # CREATE LESSON (BALANCE = 0.0)
            # -----------------------------------------------------
            lesson = Lesson(
                lesson_date=booking_date,
                time_frame=booking_time,
                client=client.full_name,
                horse="",
                payment=float(payment_per_rider),
                price_pl=float(price_per_rider),
                attendance="",
                balance=0.0,
                lesson_type=lesson_type,
                group_priv=group_priv,
                block_key="",
            )

            db.session.add(lesson)

            # -----------------------------------------------------
            # CASCADE RECALC FOR EXISTING CLIENTS
            # -----------------------------------------------------
            if not is_new_client:
                recalc_client_cascade(client.full_name)

        # ---------------------------------------------------------
        # MARK ENQUIRY AS PROCESSED
        # ---------------------------------------------------------
        enquiry.processed = True
        enquiry.processed_at = datetime.utcnow()

        # ---------------------------------------------------------
        # COMMIT + REDIRECT
        # ---------------------------------------------------------
        db.session.commit()
        flash("Selected riders processed and lessons created.", "success")
        return redirect(url_for('trailride_enquiries'))



    @app.route('/fetch_trailride_submissions')
    def fetch_trailride_submissions():
        api_key = os.getenv("JOTFORM_API_KEY")
        if not api_key:
            flash("Missing JOTFORM_API_KEY", "danger")
            return redirect(url_for('trailride_enquiries'))

        cutoff_date = datetime.utcnow() - timedelta(days=30)

        latest = (TrailRideSubmission.query
                  .order_by(TrailRideSubmission.received_at.desc())
                  .first())

        latest_ts = latest.received_at if latest else cutoff_date

        url = (
            f"https://api.jotform.com/form/{TRAIL_FORM_ID}/submissions"
            f"?apiKey={api_key}"
            f"&limit=1000"
            f"&filter[created_at][gt]={latest_ts.strftime('%Y-%m-%d %H:%M:%S')}"
        )

        response = requests.get(url)
        data = response.json()

        submissions = data.get("content", [])

        imported = 0
        skipped = 0

        for sub in submissions:
            submission_id = sub.get("id")
            if not submission_id:
                continue

            existing = TrailRideSubmission.query.filter_by(submission_id=submission_id).first()
            if existing:
                skipped += 1
                continue

            created_at_str = sub.get("created_at")
            if created_at_str:
                created_at = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
                if created_at < cutoff_date:
                    skipped += 1
                    continue

            incoming = TrailRideSubmission(
                submission_id=submission_id,
                form_id=TRAIL_FORM_ID,
                raw_payload=sub,
                received_at=created_at,
                processed=False,
                jotform_id=submission_id
            )

            db.session.add(incoming)
            imported += 1

        db.session.commit()

        flash(f"Imported {imported} new submissions, skipped {skipped}.", "success")
        return redirect(url_for('trailride_enquiries'))


    @app.route('/trailride_enquiries/ignore/<int:id>')
    def trailride_enquiries_ignore(id):
        enquiry = TrailRideSubmission.query.get(id)
        if not enquiry:
            flash("Enquiry not found.", "danger")
            return redirect(url_for('trailride_enquiries'))

        enquiry.ignored = True
        db.session.commit()

        flash("Enquiry ignored.", "info")
        return redirect(url_for('trailride_enquiries'))


    @app.route("/fetch_general_enquiries")
    def fetch_general_enquiries():
        import requests
        from datetime import datetime
        from models import GeneralEnquirySubmission
        from helpers_jf import GENERAL_ENQUIRY_FORM_ID, parse_general_enquiry_payload

        api_key = current_app.config.get("JOTFORM_API_KEY")
        if not api_key:
            flash("Missing JotForm API key", "danger")
            return redirect(url_for("general_enquiries"))

        url = f"https://api.jotform.com/form/{GENERAL_ENQUIRY_FORM_ID}/submissions?apiKey={api_key}"
        response = requests.get(url)
        data = response.json()

        if data.get("responseCode") != 200:
            flash("Failed to fetch from JotForm", "danger")
            return redirect(url_for("general_enquiries"))

        submissions = data.get("content", [])
        new_count = 0

        for sub in submissions:
            submission_id = sub.get("id")
            created_at = sub.get("created_at")

            existing = GeneralEnquirySubmission.query.filter_by(submission_id=submission_id).first()
            if existing:
                continue

            parsed = parse_general_enquiry_payload(sub)

            entry = GeneralEnquirySubmission(
                submission_id=submission_id,
                created_at=datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S"),
                rider_name=parsed["rider_name"],
                rider_age=parsed["rider_age"],
                rider_height_cm=parsed["rider_height_cm"],
                rider_weight_kg=parsed["rider_weight_kg"],
                email_address=parsed["email_address"],
                mobile_phone=parsed["mobile_phone"],
                comments=parsed["comments"],
                ignored=False
            )

            db.session.add(entry)
            new_count += 1

        db.session.commit()

        flash(f"Imported {new_count} new enquiries.", "success")
        return redirect(url_for("general_enquiries"))


    @app.route("/general_enquiries")
    def general_enquiries():
        from models import GeneralEnquirySubmission, Time
        from datetime import date

        enquiries = (
            GeneralEnquirySubmission.query
            .filter_by(ignored=False, processed=False)
            .order_by(GeneralEnquirySubmission.created_at.desc())
            .all()
        )

        times = Time.query.order_by(Time.timerange).all()
        current_date = date.today().strftime("%Y-%m-%d")

        return render_template(
            "general_enquiries.html",
            enquiries=enquiries,
            times=times,
            current_date=current_date
        )


    @app.route("/ignore_general_enquiry/<int:enquiry_id>")
    def ignore_general_enquiry(enquiry_id):
        from models import GeneralEnquirySubmission

        entry = GeneralEnquirySubmission.query.get(enquiry_id)
        if not entry:
            return "Not found", 404

        entry.ignored = True
        db.session.commit()

        return redirect(url_for("general_enquiries"))


    @app.route("/general_enquiry/<int:enquiry_id>")
    def view_general_enquiry(enquiry_id):
        from models import GeneralEnquirySubmission

        entry = GeneralEnquirySubmission.query.get(enquiry_id)
        if not entry:
            return "Not found", 404

        return render_template("general_enquiry_view.html", e=entry)

    @app.route("/general_enquiries_process", methods=["POST"])
    def general_enquiries_process():
        enquiry_id = request.form.get("process_enquiry")

        if not enquiry_id:
            flash("No enquiry selected.", "warning")
            return redirect(url_for('general_enquiries'))

        from models import GeneralEnquirySubmission, Client, Lesson
        from datetime import datetime
        from sqlalchemy import func

        enquiry = GeneralEnquirySubmission.query.get(enquiry_id)
        if not enquiry or enquiry.ignored:
            flash("Enquiry not found or already processed.", "warning")
            return redirect(url_for('general_enquiries'))

        # -------------------------------
        # NAME NORMALISATION
        # -------------------------------
        def clean_name(name):
            if not name:
                return ""
            name = " ".join(name.split())      # collapse multiple spaces
            return name.title()                # proper case

        # IGNORE BRANCH
        if request.form.get(f"ignore_{enquiry_id}") == "1":
            enquiry.ignored = True
            enquiry.processed = True
            enquiry.processed_at = datetime.utcnow()
            db.session.commit()
            flash("Enquiry ignored.", "info")
            return redirect(url_for('general_enquiries'))

        # BOOKING FIELDS
        booking_date = request.form.get(f"booking_date_{enquiry_id}")
        booking_time = request.form.get(f"booking_time_{enquiry_id}")
        lesson_type = request.form.get(f"lesson_type_{enquiry_id}")
        group_priv = request.form.get(f"group_priv_{enquiry_id}")
        price_per_rider = float(request.form.get(f"price_per_rider_{enquiry_id}") or 0)
        payment_per_rider = float(request.form.get(f"payment_per_rider_{enquiry_id}") or 0)

        if not booking_date or not booking_time:
            flash("Please select BOTH a date and a time.", "danger")
            return redirect(url_for('general_enquiries'))

        # SINGLE RIDER CHECKBOX
        if not request.form.get(f"process_rider_{enquiry_id}_1"):
            flash("No rider selected.", "warning")
            return redirect(url_for('general_enquiries'))

        # CLEANED RIDER DETAILS
        rider_name = smart_proper_name(enquiry.rider_name)
        rider_mobile = enquiry.mobile_phone.strip() if enquiry.mobile_phone else ""
        rider_email = enquiry.email_address.strip() if enquiry.email_address else ""

        # ENSURE CLIENT EXISTS
        existing_client = Client.query.filter(
            func.lower(Client.full_name) == rider_name.lower()
        ).first()

        if not existing_client:
            client = Client(
                full_name=rider_name,
                mobile=rider_mobile,
                email_primary=rider_email
            )
            db.session.add(client)
            db.session.flush()
        else:
            client = existing_client

        # CREATE LESSON
        lesson = Lesson(
            lesson_date=booking_date,
            time_frame=booking_time,
            client=client.full_name,
            horse="",
            payment=payment_per_rider,
            price_pl=price_per_rider,
            attendance="",
            balance=0.0,
            lesson_type=lesson_type,
            group_priv=group_priv,
            block_key=""
        )

        db.session.add(lesson)

        # MARK ENQUIRY AS PROCESSED
        enquiry.processed = True
        enquiry.processed_at = datetime.utcnow()

        db.session.commit()

        flash("General enquiry processed and lesson created.", "success")
        return redirect(url_for('general_enquiries'))



    @app.route('/send_invoice')
    def send_invoice():
        import win32com.client
        from datetime import datetime

        lesson_id = request.args.get('lesson_id')
        if not lesson_id:
            return "Missing lesson ID", 400

        lesson = db.session.query(Lesson).get(int(lesson_id))
        if not lesson:
            return "Lesson not found", 404

        client = db.session.query(Client).filter(
                func.lower(func.trim(Client.full_name)) == func.lower(func.trim(lesson.client))
        ).first()
        if not client:
            return "Client not found", 404

        # Format email
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = client.email_primary or client.guardian_contact or "your@email.com"
        mail.Subject = f"Invoice for {lesson.lesson_date.strftime('%d %b %Y')}"
        mail.Body = f"""Hi {client.guardian_name},

Please find your invoice for the lesson on {lesson.lesson_date.strftime('%A, %d %B %Y')}.

Client: {client.full_name}
Horse: {lesson.horse}
Time: {lesson.time_frame}
Amount: ${lesson.price_pl:.2f}

Let us know if you have any questions.

Thanks,
Cherbon Waters Admin
"""
        mail.Display()  # Use mail.Send() to send immediately

        return "Invoice email opened in Outlook"




    @app.route('/update_invoice_flag', methods=['POST'])
    def update_invoice_flag():
        cid = int(request.form.get('client_id'))
        flag = 'invoice_required' in request.form
        client = db.session.query(Client).get(cid)
        if client:
            client.invoice_required = flag
            db.session.commit()
        return redirect(url_for('client_view', client=client.full_name))

    """
    @app.post("/recalculate_client")
    def recalculate_client():
        client = request.form.get("client")

        # get all lessons for this client
        lessons = Lesson.query.filter_by(client=client).all()

        for l in lessons:
            carry = l.carry_fwd or 0
            payment = l.payment or 0
            price = l.price_pl or 0
            att = (l.attendance or '').strip().upper()

            balance = carry + payment
            if att in ['Y', 'N']:
                balance -= price

            l.balance = balance

        db.session.commit()

        flash(f"Lessons recalculated for client: {client}", "success")
        return redirect(url_for('client_view', client_filter=client), code=303)
    """

    @app.route('/notifications/clear_processed')
    def clear_processed():
        db.session.query(IncomingSubmission).filter(
            (IncomingSubmission.processed == True) |
            (IncomingSubmission.processed.is_(None))
        ).delete(synchronize_session=False)
        db.session.commit()
        return redirect(url_for('notifications'))

    @app.route("/save_teacher_slots", methods=["POST"])
    def save_teacher_slots():
        data = request.get_json()
        date = data.get("date")
        slots = data.get("slots", {})

        for slot_num, teacher_name in slots.items():
            slot = TeacherSlot.query.filter_by(
                lesson_date=date,
                slot_number=int(slot_num)
            ).first()

            if not slot:
                slot = TeacherSlot(
                    lesson_date=date,
                    slot_number=int(slot_num)
                )
                db.session.add(slot)

            slot.teacher_name = teacher_name or ""

        db.session.commit()
        return jsonify({"status": "ok"})

    # ---------------------------------------------------------
    # INACTIVE ROUTE: SAVE T1–T5 TEACHER SLOT FLAGS
    # Reason: System now fast; unified main Save button handles
    #         all T1–T5 tag persistence. This route is parked
    #         for future use if needed.
    # ---------------------------------------------------------
    # @app.route("/save_t_slots", methods=["POST"])
    # def save_t_slots():
    #     data = request.get_json() or {}
    #
    #     # Expected payload structure:
    #     # {
    #     #     "date": "2026-03-24",
    #     #     "lessons": [
    #     #         { "lesson_id": 123, "T1": True,  "T2": False,
    #     #           "T3": False, "T4": False, "T5": False },
    #     #         { "lesson_id": 124, "T1": False, "T2": True,
    #     #           "T3": False, "T4": False, "T5": False }
    #     #     ]
    #     # }
    #
    #     print("Received T-slot payload:", data)
    #
    #     return jsonify({"status": "ok"}), 200

    @app.route("/save_teacher_blocks", methods=["POST"])
    def save_teacher_blocks():
        data = request.get_json()
        date = data.get("date")
        incoming = data.get("teacher_blocks", [])

        # Load existing rows for this date
        existing = TeacherBlock.query.filter_by(date=date).all()
        existing_by_id = {tb.id: tb for tb in existing}

        incoming_ids = []

        # Process incoming rows
        for item in incoming:
            tb_id = item.get("id")

            if tb_id and tb_id in existing_by_id:
                # Update existing
                tb = existing_by_id[tb_id]
                tb.block_key = item["block_key"]
                tb.horse = item["horse"]
                tb.teacher_name = item["teacher_name"]
                tb.notes = item["notes"]
                incoming_ids.append(tb_id)

            else:
                # Insert new
                tb = TeacherBlock(
                    date=date,
                    block_key=item["block_key"],
                    horse=item["horse"],
                    teacher_name=item["teacher_name"],
                    notes=item["notes"]
                )
                db.session.add(tb)
                db.session.flush()  # get ID
                incoming_ids.append(tb.id)

        # Delete rows removed in UI
        for tb in existing:
            if tb.id not in incoming_ids:
                db.session.delete(tb)

        db.session.commit()

        return jsonify({"status": "ok"})

    @app.route('/manage_teacher_times', methods=['GET'])
    def manage_teacher_times():
        rows = db.session.query(TeacherTime).order_by(
            TeacherTime.teacher_key,
            TeacherTime.weekday,
            TeacherTime.time
        ).all()

        grouped = {}
        for r in rows:
            grouped.setdefault(r.teacher_key, {}).setdefault(r.weekday, []).append(r)

        teacher_keys = [f"Teacher {i}" for i in range(1, 6)]
        for teacher in teacher_keys:
            grouped.setdefault(teacher, {})
            for weekday in range(7):
                grouped[teacher].setdefault(weekday, [])

        return render_template('manage_teacher_times_grouped.html',
                               grouped=grouped,
                               teacher_keys=teacher_keys)

    @app.route('/manage_teacher_times/add', methods=['POST'])
    def add_teacher_time():
        teacher_key = (request.form.get('teacher_key') or '').strip()
        weekday = request.form.get('weekday')
        time_input = (request.form.get('time') or '').strip()

        if not teacher_key or not weekday or not time_input:
            flash("Missing teacher, weekday, or time.", "danger")
            return redirect(url_for('manage_teacher_times'))

        # Normalize separators (e.g., "11.00" -> "11:00", "11 00" -> "11:00")
        raw_time = time_input
        time_norm = re.sub(r'[.\s]+', ':', raw_time)

        # Validate basic HH:MM shape
        m = re.match(r'^(\d{1,2}):(\d{1,2})$', time_norm)
        if not m:
            flash(f"Invalid time format: '{raw_time}'. Use HH:MM (e.g., 09:30).", "danger")
            return redirect(url_for('manage_teacher_times'))

        hour = int(m.group(1))
        minute = int(m.group(2))

        # Bounds check
        if not (0 <= hour <= 23 and 0 <= minute <= 59):
            flash(f"Out-of-range time: '{raw_time}'. Use 00:00 to 23:59.", "danger")
            return redirect(url_for('manage_teacher_times'))

        # Zero-pad to canonical HH:MM
        time_val = f"{hour:02d}:{minute:02d}"

        try:
            weekday_int = int(weekday)
        except Exception:
            weekday_int = 0

        t = TeacherTime(teacher_key=teacher_key, weekday=weekday_int, time=time_val)
        db.session.add(t)
        db.session.commit()
        flash(f"Added {teacher_key} - weekday {weekday_int} at {time_val}.", "success")
        return redirect(url_for('manage_teacher_times'))

    @app.route('/manage_teacher_times/delete/<int:row_id>', methods=['POST'])
    def delete_teacher_time(row_id):
        row = db.session.query(TeacherTime).get(row_id)
        if row:
            db.session.delete(row)
            db.session.commit()
        return redirect(url_for('manage_teacher_times'))







    @app.route("/debug_dates")
    def debug_dates():
        out = []
        for l in Lesson.query.order_by(Lesson.lesson_id.asc()).all():
            out.append(f"{l.lesson_id} | {repr(l.client)} | {repr(l.lesson_date)} | {type(l.lesson_date)}")
        return "<br>".join(out)




    @app.route('/admin/seed_jotform_ids')
    def seed_jotform_ids():
        rows = db.session.query(IncomingSubmission).all()
        updated = 0

        for row in rows:
            submission_id = str(row.form_id)
            riders = parse_jotform_payload(row.raw_payload)

            for rider in riders:
                # Use your existing fuzzy matching logic
                matches = rider.get("matches", [])
                existing = matches[0] if matches else None

                if existing:
                    # Seed the JotForm submission ID
                    existing.jotform_submission_id = submission_id
                    updated += 1

        db.session.commit()
        return f"Seed complete. Updated {updated} clients."

    def extract_start(t):
        if not t:
            return ""
        # Lessons always use "HH:MM - HH:MM"
        return t.split("-")[0].strip()


    def get_start_from_block_key(block_key):
        # 1. Extract the first segment before "_"
        raw = block_key.split("_")[0]

        # 2. Extract digits
        digits = "".join(ch for ch in raw if ch.isdigit())
        if len(digits) < 4:
            return ""

        hhmm = digits[:4]
        formatted = hhmm[:2] + ":" + hhmm[2:]

        # 3. Look up the real timerange in the times table
        all_times = Time.query.all()
        for t in all_times:
            t_digits = "".join(ch for ch in t.timerange if ch.isdigit())
            if t_digits.startswith(hhmm):
                return t.timerange.split("-")[0].strip()

        # fallback
        return formatted

    def rebuild_full_timerange(start_time):
        # start_time = "HH:MM"
        digits = "".join(ch for ch in start_time if ch.isdigit())
        if len(digits) < 4:
            return start_time

        hhmm = digits[:4]

        # Look up in Time table
        all_times = Time.query.all()
        for t in all_times:
            t_digits = "".join(ch for ch in t.timerange if ch.isdigit())
            if t_digits.startswith(hhmm):
                return t.timerange  # full "HH:MM - HH:MM"




    @app.route('/save_txt', methods=['POST'])
    def save_txt():
        try:
            selected_date = request.args.get("selected_date")
            if not selected_date:
                return {"error": "Missing selected_date"}, 400

            day = datetime.strptime(selected_date, "%Y-%m-%d").date()
            today_str = day.strftime("%d-%m-%y")
            selected_date_str = selected_date
            filename = f"lesson_schedule_{today_str}.txt"

            # --- Pull all horses ---
            horses = [
                (h.horse or "").strip()
                for h in db.session.query(Horse).order_by(Horse.orderpdk).all()
                if (h.horse or "").strip()
            ]

            # --- Pull lessons ---
            lessons = db.session.query(Lesson).filter(
                Lesson.lesson_date == day
            ).order_by(Lesson.time_frame.asc()).all()

            # --- Merge UI state into lessons ---
            ui_state = request.get_json(silent=True) or {}
            ui_lessons = {str(item["lesson_id"]): item for item in ui_state.get("lessons", [])}

            for l in lessons:
                lid = str(l.lesson_id)
                if lid in ui_lessons:
                    ui = ui_lessons[lid]

                    if ui.get("horse"):
                        l.horse = ui["horse"]

                    if ui.get("time"):
                        l.time_frame = rebuild_full_timerange(ui["time"])

                    if ui.get("attendance"):
                        l.attendance = ui["attendance"]

                    if ui.get("teacher"):
                        l.teacher = ui["teacher"]

            # --- Pull teacher blocks ---
            teacher_blocks = TeacherBlock.query.filter_by(date=selected_date_str).all()


            teacher_block_times = []
            for tb in teacher_blocks:
                start = get_start_from_block_key(tb.block_key)
                if start and tb.horse:
                    teacher_block_times.append((tb.horse.strip(), start + "*"))

            # --- Build time slots (include teacher block times) ---
            lesson_times = {
                extract_start(l.time_frame)
                for l in lessons
                if l.time_frame
            }

            block_times = {
                t.replace("*", "")
                for (_, t) in teacher_block_times
            }

            time_slots = sorted(lesson_times | block_times)

            # --- Build empty schedule matrix ---
            schedule = {h: {slot: "" for slot in time_slots} for h in horses}

            # --- Fill schedule with lessons ---
            for l in lessons:
                h = (l.horse or "").strip()
                if not h or h not in schedule:
                    continue

                if (l.attendance or "").upper() == "C":
                    continue

                slot = extract_start(l.time_frame)
                if slot not in time_slots:
                    continue

                if (l.lesson_type or "").strip() == "Trail Ride":
                    disp = f"{slot}T"
                else:
                    disp = slot

                schedule[h][slot] = disp

            # --- Insert teacher block times ---
            for horse, t in teacher_block_times:
                base = t.replace("*", "")
                if horse in schedule and base in schedule[horse]:
                    existing = schedule[horse][base]
                    if existing:
                        schedule[horse][base] = existing + "*"
                    else:
                        schedule[horse][base] = t

            # --- Build TXT output ---
            columns = ["Horse"] + time_slots

            col_widths = []
            for col in columns:
                if col == "Horse":
                    max_len = max(len(h) for h in horses)
                else:
                    max_len = max(len(schedule[h].get(col, "")) for h in horses)
                    max_len = max(max_len, len(col))
                col_widths.append(max(max_len, 5))

            def fmt_row(cells):
                return "| " + " | ".join(
                    str(c).ljust(col_widths[i]) for i, c in enumerate(cells)
                ) + " |"

            header = f"Lesson Schedule for {today_str}"
            header_line = fmt_row(columns)
            underline = "|" + "|".join("_" * (w + 2) for w in col_widths) + "|"

            lines = [header, "", header_line, underline]

            for h in horses:
                row = [h] + [schedule[h][slot] for slot in time_slots]
                lines.append(fmt_row(row))
                lines.append(underline)

            full_text = "\n".join(lines)

            return Response(
                full_text,
                mimetype="text/plain",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )

        except Exception as e:
            return {"error": str(e)}, 500

    @app.route('/save_xlsx', methods=['POST'])
    def save_xlsx():
        try:
            selected_date = request.args.get("selected_date")
            if not selected_date:
                return {"error": "Missing selected_date"}, 400

            day = datetime.strptime(selected_date, "%Y-%m-%d").date()
            selected_date_str = selected_date  # TeacherBlock stores date as string
            today_str = day.strftime("%d-%m-%y")
            filename = f"lesson_schedule_{today_str}.xlsx"

            # --- Pull all horses ---
            horses = [
                (h.horse or "").strip()
                for h in db.session.query(Horse).order_by(Horse.orderpdk).all()
                if (h.horse or "").strip()
            ]

            # --- Pull lessons ---
            lessons = db.session.query(Lesson).filter(
                Lesson.lesson_date == day
            ).order_by(Lesson.time_frame.asc()).all()

            # --- Merge UI state ---
            ui_state = request.get_json(silent=True) or {}
            ui_lessons = {str(item["lesson_id"]): item for item in ui_state.get("lessons", [])}

            for l in lessons:
                lid = str(l.lesson_id)
                if lid in ui_lessons:
                    ui = ui_lessons[lid]

                    if ui.get("horse"):
                        l.horse = ui["horse"]
                    if ui.get("time"):
                        l.time_frame = rebuild_full_timerange(ui["time"])
                    if ui.get("attendance"):
                        l.attendance = ui["attendance"]
                    if ui.get("teacher"):
                        l.teacher = ui["teacher"]

            # --- Pull teacher blocks ---
            teacher_blocks = TeacherBlock.query.filter_by(date=selected_date_str).all()

            teacher_block_times = []
            for tb in teacher_blocks:
                start = get_start_from_block_key(tb.block_key)
                if start and tb.horse:
                    teacher_block_times.append((tb.horse.strip(), start + "*"))

            # --- Build time slots ---
            lesson_times = {
                extract_start(l.time_frame)
                for l in lessons
                if l.time_frame
            }

            block_times = {
                t.replace("*", "")
                for (_, t) in teacher_block_times
            }

            time_slots = sorted(lesson_times | block_times)

            # --- Build schedule ---
            schedule = {h: {slot: "" for slot in time_slots} for h in horses}

            # --- Fill lessons ---
            for l in lessons:
                h = (l.horse or "").strip()
                if not h or h not in schedule:
                    continue
                if (l.attendance or "").upper() == "C":
                    continue

                slot = extract_start(l.time_frame)
                if slot not in time_slots:
                    continue

                disp = f"{slot}T" if (l.lesson_type or "").strip() == "Trail Ride" else slot
                schedule[h][slot] = disp

            # --- Insert teacher blocks ---
            for horse, t in teacher_block_times:
                base = t.replace("*", "")
                if horse in schedule and base in schedule[horse]:
                    existing = schedule[horse][base]
                    schedule[horse][base] = existing + "*" if existing else t

            # --- Excel output ---
            out_dir = "/home/schedule_exports"
            os.makedirs(out_dir, exist_ok=True)
            excel_path = os.path.join(out_dir, filename)

            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Lesson Schedule"

            columns = ["Horse"] + time_slots
            ws.append(columns)

            for h in horses:
                row = [h] + [schedule[h][slot] for slot in time_slots]
                ws.append(row)

            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.border = thin_border

            ws.page_margins.left   = 0.2
            ws.page_margins.right  = 0.2
            ws.page_margins.top    = 0.1
            ws.page_margins.bottom = 0.1
            ws.page_margins.header = 0.0
            ws.page_margins.footer = 0.0
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.paperSize = ws.PAPERSIZE_A4

            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                adjusted_width = max(max_length + 2, 10)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            wb.save(excel_path)

            return send_file(excel_path, as_attachment=True)

        except Exception as e:
            return {"error": str(e)}, 500

    @app.route('/save_grid_override', methods=['POST'])
    def save_grid_override():
        data = request.get_json()

        override_date = data.get('date')
        time_label = data.get('time')
        teacher_index = data.get('teacher')
        state = data.get('state')

        if not (override_date and time_label and teacher_index is not None):
            return jsonify(success=False, error="Missing fields"), 400

        try:
            conn = get_db()
            cur = conn.cursor()

            cur.execute("""
                INSERT INTO teacher_grid_overrides (override_date, time_label, teacher_index, state)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (override_date, time_label, teacher_index)
                DO UPDATE SET state = EXCLUDED.state, updated_at = NOW();
            """, (override_date, time_label, teacher_index, state))

            conn.commit()
            cur.close()

            return jsonify(success=True)

        except Exception as e:
            conn.rollback()
            return jsonify(success=False, error=str(e)), 500

    @app.route('/reset_grid_overrides', methods=['POST'])
    def reset_grid_overrides():
        data = request.get_json()
        override_date = data.get('date')

        if not override_date:
            return jsonify({'status': 'error', 'message': 'Missing date'}), 400

        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            DELETE FROM teacher_grid_overrides
            WHERE override_date = %s
        """, (override_date,))

        conn.commit()
        cur.close()

        return jsonify({'status': 'ok'})


    @app.route("/debug_order/<client>")
    def debug_order(client):
        lessons = Lesson.query.filter_by(client=client).order_by(
            Lesson.lesson_date.asc(),
            Lesson.lesson_id.asc()
        ).all()

        out = []
        for l in lessons:
            out.append(f"{l.lesson_id} | {l.lesson_date} | {l.time_frame}")
        return "<br>".join(out)

    @app.route("/save_weekly_events", methods=["POST"])
    def save_weekly_events():
        fy = request.form.get("fy")

        for key, value in request.form.items():
            if key.startswith("event1_") or key.startswith("event2_") or key.startswith("notes_"):
                field, week_start_str = key.split("_")
                week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()

                row = (
                    WeeklyEvent.query
                    .filter_by(week_start=week_start, fy=fy)
                    .first()
                )

                if not row:
                    row = WeeklyEvent(
                        week_start=week_start,
                        fy=fy
                    )
                    db.session.add(row)

                setattr(row, field, value.strip())

        db.session.commit()
        return redirect(f"/summaries?fy={fy}")


    @app.route("/recalc_all_lessons", methods=["POST"])
    def recalc_all_lessons_route():
        try:
            recalc_all_lessons()   # call your function
            return "OK", 200
        except Exception as e:
            print("Recalc error:", e)
            return "ERROR", 500



    @app.route('/debug_payload/<int:id>')
    def debug_payload(id):
        row = db.session.query(IncomingSubmission).get(id)
        print(row.raw_payload)
        return "Printed to console"

    @app.route("/client/<int:client_id>/change_price", methods=["POST"])
    def change_client_price(client_id):
        mode = request.form.get("mode")
        group_priv = request.form.get("group_priv")
        cutoff_date = request.form.get("cutoff_date")
        new_price = request.form.get("new_price")

        if not group_priv:
            flash("Group Priv is required.", "danger")
            return redirect(url_for("client_view", client_id=client_id))

        if not new_price:
            flash("New price is required.", "danger")
            return redirect(url_for("client_view", client_id=client_id))

        # Get client full name (lessons.client stores the name, not ID)
        client_obj = Client.query.get(client_id)
        client_name = client_obj.full_name

        db.session.execute(
            text("""
                UPDATE lessons
                SET price_pl = :new_price
                WHERE client = :client_name
                  AND group_priv = :group_priv
                  AND lesson_date >= :cutoff_date
            """),
            {
                "new_price": new_price,
                "client_name": client_name,
                "group_priv": group_priv,
                "cutoff_date": cutoff_date
            }
        )

        db.session.commit()
        return redirect(url_for("client_view", client_id=client_id))


    @app.route("/client/<int:client_id>/change_time", methods=["POST"])
    def change_client_time(client_id):
        mode = request.form.get("mode")
        cutoff = request.form.get("cutoff_date")
        gp = request.form.get("group_priv")
        new_time = request.form.get("new_time")

        if not new_time or not cutoff or not gp:
            return {"status": "error", "message": "Missing data"}, 400

        cutoff_date = datetime.strptime(cutoff, "%Y-%m-%d").date()

        client = Client.query.get(client_id)
        if not client:
            return {"status": "error", "message": "Client not found"}, 404

        lessons = (
            Lesson.query
            .filter(
                Lesson.client == client.full_name,
                Lesson.lesson_date == cutoff_date,   # single date only
                Lesson.group_priv == gp
            )
            .all()
        )

        for l in lessons:
            l.time_frame = new_time

        db.session.commit()
        recalc_all_lessons()

        return redirect(url_for("client_view", client_id=client_id))


    @app.route("/client/<int:client_id>/change_horse", methods=["POST"])
    def change_client_horse(client_id):
        mode = request.form.get("mode")
        group_priv = request.form.get("group_priv")
        cutoff_date = request.form.get("cutoff_date")
        new_horse_id = request.form.get("new_horse")

        if not group_priv:
            flash("Group Priv is required.", "danger")
            return redirect(url_for("client_view", client_id=client_id))

        # Get client full name (lessons.client stores the name)
        client_obj = Client.query.get(client_id)
        client_name = client_obj.full_name

        # Get horse name (lessons.horse stores the name)
        horse_obj = Horse.query.get(new_horse_id) if new_horse_id else None
        horse_name = horse_obj.horse if horse_obj else None

        # ---------------------------------------------------------
        # CHANGE ALL FUTURE LESSONS TO A NEW HORSE
        # ---------------------------------------------------------
        if mode == "change_horse":
            db.session.execute(
                text("""
                    UPDATE lessons
                    SET horse = :horse_name
                    WHERE client = :client_name
                      AND group_priv = :group_priv
                      AND lesson_date >= :cutoff_date
                """),
                {
                    "horse_name": horse_name,
                    "client_name": client_name,
                    "group_priv": group_priv,
                    "cutoff_date": cutoff_date
                }
            )

        # ---------------------------------------------------------
        # ASSIGN HORSE ONLY IF EMPTY
        # ---------------------------------------------------------
        elif mode == "assign_if_empty":
            db.session.execute(
                text("""
                    UPDATE lessons
                    SET horse = :horse_name
                    WHERE client = :client_name
                      AND group_priv = :group_priv
                      AND lesson_date >= :cutoff_date
                      AND (horse IS NULL OR horse = '')
                """),
                {
                    "horse_name": horse_name,
                    "client_name": client_name,
                    "group_priv": group_priv,
                    "cutoff_date": cutoff_date
                }
            )

        # ---------------------------------------------------------
        # NEW MODE: CHANGE HORSE X → Y FROM CUTOFF
        # ---------------------------------------------------------
        elif mode == "change_horse_specific":
            old_horse = request.form.get("old_horse")

            db.session.execute(
                text("""
                    UPDATE lessons
                    SET horse = :new_horse
                    WHERE client = :client_name
                      AND group_priv = :group_priv
                      AND lesson_date >= :cutoff_date
                      AND horse = :old_horse
                """),
                {
                    "new_horse": horse_name,
                    "client_name": client_name,
                    "group_priv": group_priv,
                    "cutoff_date": cutoff_date,
                    "old_horse": old_horse
                }
            )

        db.session.commit()
        return redirect(url_for("client_view", client_id=client_id))


    @app.route("/phone_lookup", methods=["GET", "POST"])
    def phone_lookup():
        client = None
        number = ""

        if request.method == "POST":
            number = (request.form.get("phone") or "").strip()
            clean = "".join(filter(str.isdigit, number))

            client = Client.query.filter(
                Client.mobile == clean
            ).first()

        return render_template("phone_lookup.html", client=client, number=number)


    @app.route('/import_lessons_xlsx', methods=['GET', 'POST'])
    def import_lessons_xlsx():

        # ⭐ GET → show upload page
        if request.method == 'GET':
            return render_template('import_lessons_xlsx.html')

        # ⭐ POST → process file
        file = request.files.get('file')
        if not file:
            return jsonify({"status": "error", "message": "No file uploaded"}), 400

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # --- Extract header row ---
        headers = [(cell.value.strip() if cell.value else None) for cell in ws[1]]

        # --- Valid Lesson fields ---
        lesson_fields = [col.name for col in Lesson.__table__.columns]

        # --- Map XLSX column index -> Lesson field name ---
        field_map = {}
        for idx, header in enumerate(headers):
            if header in lesson_fields:
                field_map[idx] = header

        inserted = 0
        skipped = 0

        # --- Process each row ---
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}

            # Map known Lesson fields + normalise blanks
            for idx, value in enumerate(row):
                if idx in field_map:
                    if isinstance(value, str) and value.strip() == "":
                        value = None
                    row_data[field_map[idx]] = value

            # Skip fully empty rows
            if not any(row_data.values()):
                skipped += 1
                continue

            # --- TEXT NORMALISATION ---
            for key, val in row_data.items():
                if isinstance(val, str):
                    row_data[key] = " ".join(val.split())

            # --- REQUIRED FIELDS ---
            required_fields = ["client", "lesson_date"]
            missing = [f for f in required_fields if not row_data.get(f)]
            if missing:
                print(f"SKIPPED ROW — missing required fields {missing} | {row_data}")
                skipped += 1
                continue

            # --- FETCH OR CREATE CLIENT ---
            client_name = row_data["client"]
            client = Client.query.filter_by(full_name=client_name).first()

            if client is None:
                client = Client(full_name=client_name)
                db.session.add(client)
                db.session.flush()  # get client_id without committing

            # --- UPDATE CLIENT AGE/WEIGHT/HEIGHT ---
            client_fields = ["age", "weight_kg", "height_cm"]

            for f in client_fields:
                if f in headers:
                    col_index = headers.index(f)
                    raw_val = row[col_index]

                    # Normalise blanks
                    if isinstance(raw_val, str):
                        raw_val = raw_val.strip()
                        if raw_val == "":
                            raw_val = None

                    # Convert to int if possible
                    if raw_val not in [None, ""]:
                        try:
                            setattr(client, f, int(raw_val))
                        except:
                            print(f"BAD CLIENT VALUE — {f}={raw_val} for {client_name}")

            # --- DATE NORMALISATION ---
            if "lesson_date" in row_data and row_data["lesson_date"]:
                val = row_data["lesson_date"]

                if isinstance(val, (datetime, date)):
                    row_data["lesson_date"] = val.date() if isinstance(val, datetime) else val

                elif isinstance(val, str):
                    try:
                        row_data["lesson_date"] = datetime.strptime(val.strip(), "%d/%m/%Y").date()
                    except ValueError:
                        print(f"BAD DATE FORMAT — {val} | {row_data}")
                        skipped += 1
                        continue

            # --- NUMERIC NORMALISATION ---
            numeric_fields = ["payment", "price_pl", "adjust", "carry_fwd", "balance"]

            for key in numeric_fields:
                if key in row_data:
                    val = row_data[key]

                    if val in [None, "", " ", "None"]:
                        row_data[key] = None
                        continue

                    try:
                        row_data[key] = float(val)
                    except:
                        print(f"BAD NUMERIC VALUE — {key}={val} | forcing to None")
                        row_data[key] = None

            # --- CREATE LESSON ---
            try:
                lesson = Lesson(**row_data)
                db.session.add(lesson)
                inserted += 1
            except Exception as e:
                print(f"ROW FAILED — {e} | {row_data}")
                skipped += 1
                continue

        # --- COMMIT ALL CHANGES ---
        db.session.commit()

        # --- RECALC ---
        try:
            recalc_all_lessons()
        except Exception as e:
            flash(f"Import succeeded but recalc failed: {e}", "error")

        flash(f"Imported {inserted} lessons. Skipped {skipped} rows.", "success")
        return redirect(url_for('other_tools'))


    @app.route("/save_teacher_tags", methods=["POST"])
    def save_teacher_tags():
        data = request.get_json(force=True)
        print("SAVE ROUTE HIT")

        # --- Extract date ---
        date_str = data.get("date")
        if not date_str:
            return jsonify({"status": "error", "message": "Missing date"}), 400

        try:
            lesson_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return jsonify({"status": "error", "message": "Invalid date"}), 400

        # --- Extract overrides ---
        overrides = data.get("overrides", {})
        print("RAW OVERRIDES:", overrides)   # 🔍 DEBUG #1
        if not isinstance(overrides, dict):
            return jsonify({"status": "error", "message": "Invalid overrides"}), 400

        # --- Build block-level structure ---
        block_map = {}
        block_meta = {}

        for lesson_id_str, tags in overrides.items():

            # Normalise tags so None/undefined doesn't crash Python
            tags = tags or []   # 🔧 prevents 502 on untick

            try:
                lid = int(lesson_id_str)
            except ValueError:
                continue

            lesson = Lesson.query.get(lid)
            if not lesson:
                continue

            block_key = lesson.time_frame or ""

            block_map.setdefault(block_key, set()).update(tags)

            block_meta[block_key] = {
                "lesson_type": lesson.lesson_type or "",
                "group_priv": lesson.group_priv or ""
            }

        # --- Debug block_map BEFORE saving ---
        print("FINAL BLOCK MAP:", block_map)   # 🔍 DEBUG #2

        # --- Delete existing block tags for this date ---
        LessonBlockTag.query.filter_by(lesson_date=lesson_date).delete()

        # --- Insert new block-level tags ---
        for block_key, tagset in block_map.items():
            meta = block_meta.get(block_key, {})

            row = LessonBlockTag(
                lesson_date=lesson_date,
                time_range=block_key or "",
                lesson_type=meta.get("lesson_type") or "",
                group_priv=meta.get("group_priv") or "",
                teacher_tags=",".join(sorted(tagset))
            )
            db.session.add(row)

        db.session.commit()

        return jsonify({"status": "ok"})

    @app.route('/weddings/staffing')
    def wedding_staffing():
        weddings = Wedding.query.order_by(Wedding.date.asc()).all()

        wedding_rows = []
        for w in weddings:
            staff_names = [a.staff.name for a in w.assignments]

            wedding_rows.append({
                "id": w.id,
                "date": w.date.strftime('%d %b %Y'),
                "pax": w.pax,
                "time": w.time,
                "service1": w.service1,
                "notes": w.notes,
                "category": w.category,
                "staff": staff_names,
                "staff_count": len(staff_names)
            })

        return render_template(
            'wedding_staffing.html',
            weddings=wedding_rows
        )


    @app.route('/staff/unavailability', methods=['GET', 'POST'])
    def manage_unavailability():
        staff_list = WeddingStaff.query.order_by(WeddingStaff.name.asc()).all()

        if request.method == 'POST':
            staff_id = int(request.form['staff_id'])
            date = request.form['date']
            reason = request.form.get('reason', '')

            entry = WeddingStaffUnavailability(
                staff_id=staff_id,
                date=date,
                reason=reason
            )
            db.session.add(entry)
            db.session.commit()

            return redirect(url_for('manage_unavailability'))

        all_unavailability = WeddingStaffUnavailability.query.order_by(
            WeddingStaffUnavailability.date.asc()
        ).all()

        return render_template(
            'manage_unavailability.html',
            staff_list=staff_list,
            unavailability=all_unavailability
        )

    @app.route('/staff/unavailability/delete/<int:id>')
    def delete_unavailability(id):
        entry = WeddingStaffUnavailability.query.get_or_404(id)
        db.session.delete(entry)
        db.session.commit()
        return redirect(url_for('manage_unavailability'))



    @app.route('/weddings/staffing/<int:wedding_id>', methods=['GET', 'POST'])
    def edit_wedding_staffing(wedding_id):
        wedding = Wedding.query.get_or_404(wedding_id)
        staff_list = WeddingStaff.query.order_by(WeddingStaff.name.asc()).all()

        if request.method == 'POST':
            raw_ids = request.form.get('staff_ids', '')
            selected_ids = [int(x) for x in raw_ids.split(',') if x.strip()]

            WeddingAssignment.query.filter_by(wedding_id=wedding.id).delete()

            for sid in selected_ids:
                db.session.add(WeddingAssignment(
                    wedding_id=wedding.id,
                    staff_id=sid
                ))

            db.session.commit()
            return redirect(url_for('wedding_staffing'))

        assigned_ids = [a.staff_id for a in wedding.assignments]

        # NEW: load unavailable staff for this wedding date
        unavailable_ids = {
            u.staff_id
            for u in WeddingStaffUnavailability.query.filter_by(date=wedding.date).all()
        }

        return render_template(
            'edit_wedding_staffing.html',
            wedding=wedding,
            staff_list=staff_list,
            assigned_ids=assigned_ids,
            unavailable_ids=unavailable_ids
        )

    @app.route('/weddings/staff/manage', methods=['GET', 'POST'])
    def manage_wedding_staff():
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if name:
                db.session.add(WeddingStaff(name=name))
                db.session.commit()
            return redirect(url_for('manage_wedding_staff'))

        staff = WeddingStaff.query.order_by(WeddingStaff.name.asc()).all()

        return render_template(
            'manage_wedding_staff.html',
            staff=staff
        )
    @app.route('/weddings/staff/delete/<int:staff_id>')
    def delete_wedding_staff(staff_id):
        staff = WeddingStaff.query.get_or_404(staff_id)

        db.session.delete(staff)
        db.session.commit()

        return redirect(url_for('manage_wedding_staff'))

    @app.route('/weddings/edit/<int:wedding_id>', methods=['GET', 'POST'])
    def edit_wedding(wedding_id):
        wedding = Wedding.query.get_or_404(wedding_id)

        if request.method == 'POST':
            date_str = request.form.get('date')
            wedding.notes = request.form.get('notes', '').strip()

            wedding.pax = request.form.get('pax') or None
            wedding.time = request.form.get('time') or None
            wedding.service1 = request.form.get('service1') or None
            wedding.category = request.form.get('category')  # NEW: allow editing type

            if date_str:
                wedding.date = datetime.strptime(date_str, "%Y-%m-%d").date()

            db.session.commit()
            return redirect(url_for('wedding_staffing'))

        return render_template('edit_wedding.html', wedding=wedding)

    @app.route('/weddings/delete/<int:wedding_id>')
    def delete_wedding(wedding_id):
        wedding = Wedding.query.get_or_404(wedding_id)

        db.session.delete(wedding)
        db.session.commit()

        return redirect(url_for('wedding_staffing'))

    @app.route('/weddings/add', methods=['GET', 'POST'])
    def add_wedding():
        if request.method == 'POST':
            date_str = request.form.get('date')
            notes = request.form.get('notes', '').strip()
            category = request.form.get('category')  # CO / IND / WR

            if date_str:
                new_wedding = Wedding(
                    date=datetime.strptime(date_str, "%Y-%m-%d").date(),
                    notes=notes,
                    category=category
                )
                db.session.add(new_wedding)
                db.session.commit()

                return redirect(url_for('wedding_staffing'))

        return render_template('add_wedding.html')


    @app.get('/add_client')
    def add_client_page():
        return render_template('add_client_page.html')


    @app.route("/db-health")
    def db_health():
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("SELECT 1"))
            return {"status": "ok"}
        except Exception:
            return {"status": "restarting"}, 503


    @app.route("/messages_menu")
    def messages_menu():
        return render_template("messages_menu.html")

    @app.route('/upgrade-items')
    def upgrade_items_list():
        items = UpgradeItem.query.order_by(UpgradeItem.created_at.desc()).all()
        return render_template('upgrade_items_list.html', items=items)


    @app.route('/upgrade-items/add', methods=['POST'])
    def upgrade_items_add():
        title = request.form.get('title')
        notes = request.form.get('notes')

        if not title:
            flash("Title is required.", "danger")
            return redirect(url_for('upgrade_items_list'))

        item = UpgradeItem(title=title, notes=notes)
        db.session.add(item)
        db.session.commit()

        flash("Upgrade item added.", "success")
        return redirect(url_for('upgrade_items_list'))


    @app.route('/upgrade-items/delete/<int:item_id>', methods=['POST'])
    def upgrade_items_delete(item_id):
        item = UpgradeItem.query.get_or_404(item_id)
        db.session.delete(item)
        db.session.commit()

        flash("Upgrade item deleted.", "success")
        return redirect(url_for('upgrade_items_list'))


    @app.route("/employeehours/lastweek")
    def employee_last_week():
        emp_id = session.get("employee_id")
        if not emp_id:
            return redirect("/employeehours")

        emp = Employee.query.get(emp_id)

        # ⭐ Use the SAME timezone as main week route
        today = datetime.now(ZoneInfo("Australia/Brisbane")).date()

        # ⭐ This week's Monday
        this_monday = today - timedelta(days=today.weekday())

        # ⭐ Last week's Monday
        start_of_week = this_monday - timedelta(days=7)
        end_of_week = start_of_week + timedelta(days=6)

        days = []

        for i in range(7):
            d = start_of_week + timedelta(days=i)

            row = EmployeeHours.query.filter_by(
                employee_id=emp_id,
                date=d
            ).first()

            # ⭐ IDENTICAL STATUS LOGIC
            if d > today:
                status = "future"
            elif row and row.sign_in and row.sign_out:
                status = "complete"
            elif row and row.sign_in and not row.sign_out:
                status = "incomplete"
            elif d == today:
                status = "today"
            else:
                status = "incomplete"

            days.append({
                "date": d,
                "status": status
            })

        return render_template(
            "employee_week_view.html",
            days=days,
            today=today,
            emp=emp,
            is_last_week=True
        )


    @app.route("/admin/employees/new", methods=["POST"])
    def create_employee():
        name = request.form.get("full_name")
        if not name:
            return "Missing name", 400

        import secrets
        setup_code = "CW-" + secrets.token_hex(3).upper()

        emp = Employee(full_name=name, setup_code=setup_code)
        db.session.add(emp)
        db.session.commit()

        return {"status": "ok", "setup_code": setup_code}


    @app.route("/admin/employees/<int:emp_id>/hours")
    def admin_employee_hours_list(emp_id):
        emp = Employee.query.get(emp_id)
        if not emp:
            return "Employee not found", 404

        rows = EmployeeHours.query.filter_by(employee_id=emp_id).order_by(
            EmployeeHours.date.desc()
        ).all()

        return render_template("admin_employee_hours_list.html", emp=emp, rows=rows)

    @app.route("/admin/employees/hours/<int:row_id>/edit")
    def admin_edit_hours(row_id):
        row = EmployeeHours.query.get(row_id)
        if not row:
            return "Not found", 404

        emp = Employee.query.get(row.employee_id)

        return render_template("admin_edit_hours.html", emp=emp, row=row)

    @app.route("/admin/employees/<int:emp_id>/reset_pin", methods=["POST"])
    def admin_reset_pin(emp_id):
        emp = Employee.query.get_or_404(emp_id)

        import secrets
        hex_code = secrets.token_hex(3).upper()
        setup_code = f"CW-{hex_code}"

        emp.setup_code = setup_code
        db.session.commit()

        reset_link = f"https://cherbonapp.click/employeehours?code={setup_code}"

        send_sms_clicksend(
            emp.phone,
            f"Your Cherbon Waters login reset code is {setup_code}\n"
            f"Tap to set your PIN:\n{reset_link}",
            app.config["EQUESTRIAN_SENDER"]
        )

        flash(f"SMS sent to {emp.phone}", "success")
        return redirect(url_for("admin_employees"))

    @app.route("/employeehours/day", methods=["GET", "POST"])
    def employeehours_day_view():

        def make_aware(x):
            if x is None:
                return None
            if x.tzinfo is None:
                return datetime(
                    x.year, x.month, x.day,
                    x.hour, x.minute, x.second,
                    tzinfo=ZoneInfo("Australia/Brisbane")
                )
            return x

        emp_id = session.get("employee_id")
        if not emp_id:
            return redirect("/employeehours")

        date_str = request.args.get("date")
        if not date_str:
            return "Missing date", 400

        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            return "Invalid date", 400

        # Load existing row if any
        row = EmployeeHours.query.filter_by(employee_id=emp_id, date=d).first()

        # Make stored datetimes timezone-aware
        if row:
            row.sign_in = make_aware(row.sign_in)
            row.break_start = make_aware(row.break_start)
            row.break_end = make_aware(row.break_end)
            row.sign_out = make_aware(row.sign_out)

        # Determine editability rules
        today = datetime.now(ZoneInfo("Australia/Brisbane")).date()
        is_today = (d == today)

        # Incomplete if missing sign_in OR sign_out OR row doesn't exist
        is_incomplete = (not row) or (not row.sign_in) or (not row.sign_out)

        # Editable if today OR past incomplete
        is_future = d > today
        editable = (not is_future) and (is_today or is_incomplete)

        if request.method == "POST":

            # Backend safety: block editing completed past days
            if not editable:
                return "This day is complete and cannot be edited", 403

            # If user already confirmed, skip checks and save
            if request.form.get("confirmed") == "1":
                action = request.form.get("action")
                time_str = request.form.get("corrected_time") or request.form.get("time")
                notes = request.form.get("notes", "")

                t = datetime.strptime(time_str, "%H:%M").time()
                dt = datetime.combine(d, t, tzinfo=ZoneInfo("Australia/Brisbane"))

                if not row:
                    row = EmployeeHours(employee_id=emp_id, date=d)
                    db.session.add(row)

                field_map = {
                    "start": "sign_in",
                    "break_start": "break_start",
                    "break_end": "break_end",
                    "finish": "sign_out"
                }

                field = field_map.get(action)
                if field:
                    setattr(row, field, dt)

                # ⭐ RECORD SUBMISSION TIME
                if action == "finish":
                    row.submitted_at = datetime.now(ZoneInfo("Australia/Brisbane"))

                if notes:
                    row.notes = notes

                db.session.commit()
                return redirect(f"/employeehours/day?date={date_str}")

            # Normal POST (validation required)
            action = request.form.get("action")
            time_str = request.form.get("time")
            notes = request.form.get("notes", "")

            # Convert time string to datetime
            t = datetime.strptime(time_str, "%H:%M").time()
            dt = datetime.combine(d, t, tzinfo=ZoneInfo("Australia/Brisbane"))

            # Create row if missing
            if not row:
                row = EmployeeHours(employee_id=emp_id, date=d)
                db.session.add(row)

            now = datetime.now(ZoneInfo("Australia/Brisbane"))

            # ============================================================
            #  ACE'S FULL COMBINED VALIDATION LOGIC
            # ============================================================

            # 1) SUSPICIOUS AM/PM CHECK (SIGN-IN ONLY)
            if action == "start":

                # If selected time is more than 6 hours in the future → AM/PM mistake
                if dt > now + timedelta(hours=6):
                    return f"Suspicious time: {t.strftime('%I:%M %p')}. Check AM/PM.", 400

                # WRONG-DAY CHECK (signing in for a past day)
                if d < now.date():
                    if dt.time() > time(18, 0):
                        return f"You're signing in for {d.strftime('%A')}. {t.strftime('%I:%M %p')} looks incorrect.", 400

                # AUTO-CORRECT SUGGESTION (NEW FRONTEND FORMAT)
                if t.hour >= 18:  # 6 PM or later
                    alt_hour = (t.hour - 12) if t.hour > 12 else t.hour
                    alt_time = time(alt_hour, t.minute)
                    alt_str = alt_time.strftime("%I:%M %p")
                    alt_24 = alt_time.strftime("%H:%M")  # for hidden corrected_time field

                    message = f"Did you mean {alt_str} instead of {t.strftime('%I:%M %p')}?"
                    return f"{message}::{alt_24}", 200

            # 2) BREAK START VALIDATION
            if action == "break_start":
                if not row.sign_in:
                    return "You must start work before starting a break.", 400
                if dt <= row.sign_in:
                    return "Break start must be after your start time.", 400

            # 3) BREAK END VALIDATION
            if action == "break_end":
                if not row.break_start:
                    return "You must start your break before ending it.", 400
                if dt <= row.break_start:
                    return "Break end must be after break start.", 400

            # 4) FINISH VALIDATION (WITH OVERNIGHT SHIFT SUPPORT)
            if action == "finish":
                if not row.sign_in:
                    return "You must start work before finishing.", 400

                # SAME-DAY FINISH
                if dt >= row.sign_in:
                    end_dt = dt
                else:
                    # POSSIBLE OVERNIGHT SHIFT
                    overnight_dt = dt + timedelta(days=1)
                    if overnight_dt <= row.sign_in + timedelta(hours=16):
                        end_dt = overnight_dt
                    else:
                        return "Finish time cannot be before start time.", 400

                # ---------------------------------------------------
                #   BREAK‑DEDUCTED SHIFT HOURS (THE REAL FIX)
                # ---------------------------------------------------
                shift_seconds = (end_dt - row.sign_in).total_seconds()

                break_seconds = 0
                if row.break_start and row.break_end:
                    break_seconds = (row.break_end - row.break_start).total_seconds()

                paid_seconds = shift_seconds - break_seconds

                hours = paid_seconds / 3600
                hours_str = f"{int(hours)}h {int((hours % 1) * 60)}m"

                return f"CONFIRM_SHIFT::{t.strftime('%I:%M %p')}::{hours_str}", 200

            # ============================================================
            #  END VALIDATION
            # ============================================================

            # Map actions to fields
            field_map = {
                "start": "sign_in",
                "break_start": "break_start",
                "break_end": "break_end",
                "finish": "sign_out"
            }

            field = field_map.get(action)
            if field:
                setattr(row, field, dt)

            # ⭐ RECORD SUBMISSION TIME
            if action == "finish":
                row.submitted_at = datetime.now(ZoneInfo("Australia/Brisbane"))

            if notes:
                row.notes = notes

            db.session.commit()
            return redirect(f"/employeehours/day?date={date_str}")

        return render_template(
            "employee_day_view.html",
            date=d,
            row=row,
            editable=editable,
            is_today=is_today,
            is_incomplete=is_incomplete,
            is_future=is_future
        )

    @app.route("/employeehours/login", methods=["POST"])
    def employeehours_login():
        pin = request.form.get("pin", "").strip()

        if not pin or not pin.isdigit() or len(pin) != 6:
            return {"error": "Invalid PIN format"}, 400

        from werkzeug.security import check_password_hash

        # Find employee by PIN hash
        emp = None
        for e in Employee.query.all():
            if e.pin_hash and check_password_hash(e.pin_hash, pin):
                emp = e
                break

        # If no employee matches this PIN → failure
        if not emp:
            # No idea who attempted → cannot increment per-employee failures
            return {"error": "Incorrect PIN"}, 400

        # Check lockout
        now = datetime.now()
        if emp.locked_until and emp.locked_until > now:
            remaining = int((emp.locked_until - now).total_seconds() // 60)
            return {"error": f"Account locked. Try again in {remaining} minutes."}, 403

        # If PIN matches, reset failures
        emp.pin_failures = 0
        emp.locked_until = None
        db.session.commit()

        # Start session
        session["employee_id"] = emp.id

        return {"status": "ok"}

    @app.route("/admin/employees/<int:emp_id>/unlock", methods=["POST"])
    def admin_unlock_employee(emp_id):
        emp = Employee.query.get_or_404(emp_id)

        emp.pin_failures = 0
        emp.locked_until = None

        db.session.commit()

        return {"status": "ok"}

    @app.route("/admin/employees/<int:emp_id>/force_logout", methods=["POST"])
    def admin_force_logout(emp_id):
        # If the employee is currently logged in, remove their session
        if session.get("employee_id") == emp_id:
            session.pop("employee_id", None)

        return {"status": "ok"}

    @app.route("/admin/lockouts/clear", methods=["POST"])
    def admin_clear_all_lockouts():
        employees = Employee.query.all()
        for emp in employees:
            emp.pin_failures = 0
            emp.locked_until = None

        db.session.commit()
        return {"status": "ok"}


    @app.route("/admin/lockouts")
    def admin_lockouts():
        locked = Employee.query.filter(Employee.locked_until != None).all()
        return render_template("admin_lockouts.html", locked=locked)



    @app.route("/admin/weekly_summary")
    def admin_weekly_summary():
        today = date.today()

        # AU FY default: FY starts 1 July → 30 June
        default_fy = today.year if today.month >= 7 else today.year - 1
        fy = int(request.args.get("fy", default_fy))

        week_num = int(request.args.get("week", 0))

        # Build FY weeks
        weeks = build_fy_weeks(fy)

        # Default to current FY + correct week (Monday shows previous week)
        if week_num == 0:
            today = date.today()

            # Monday should show the week just completed
            effective_day = today - timedelta(days=1) if today.weekday() == 0 else today

            first_start = weeks[0]["start"]
            last_end = weeks[-1]["end"]

            if effective_day < first_start:
                week_num = weeks[0]["week_number"]
            elif effective_day > last_end:
                week_num = weeks[-1]["week_number"]
            else:
                for w in weeks:
                    if w["start"] <= effective_day <= w["end"]:
                        week_num = w["week_number"]
                        break

        # Clamp week_num inside valid range (AFTER default logic)
        if week_num < 1:
            week_num = 1
        if week_num > len(weeks):
            week_num = len(weeks)

        selected = weeks[week_num - 1]
        start_of_week = selected["start"]
        end_of_week = selected["end"]

        employees = Employee.query.order_by(Employee.full_name.asc()).all()

        summary = []

        for emp in employees:
            week_rows = []
            running_total = timedelta()

            for i in range(7):
                day = start_of_week + timedelta(days=i)

                row = EmployeeHours.query.filter_by(
                    employee_id=emp.id,
                    date=day
                ).first()

                if row:
                    work = timedelta()
                    if row.sign_in and row.sign_out:
                        work = row.sign_out - row.sign_in

                    brk = timedelta()
                    if row.break_start and row.break_end:
                        brk = row.break_end - row.break_start

                    net = work - brk
                    running_total += net
                else:
                    work = brk = net = None

                week_rows.append({
                    "date": day,
                    "row": row,
                    "work": work,
                    "break": brk,
                    "net": net,
                    "running_total": running_total
                })

            summary.append({
                "emp": emp,
                "week_rows": week_rows,
                "week_total": running_total
            })

        # -----------------------------
        # DYNAMIC FY LIST (2025 → current FY)
        # -----------------------------
        start_fy = 2025
        current_fy = today.year if today.month >= 7 else today.year - 1
        fy_years = list(range(start_fy, current_fy + 1))

        return render_template(
            "admin_weekly_summary.html",
            summary=summary,
            weeks=weeks,
            fy=fy,
            fy_years=fy_years,
            selected_week=week_num,
            start_of_week=start_of_week,
            end_of_week=end_of_week
        )

    @app.route("/employeehours/summary")
    def employee_weekly_summary():
        emp_id = session.get("employee_id")
        if not emp_id:
            return redirect("/employeehours")

        emp = Employee.query.get(emp_id)

        # Determine current week (Mon–Sun)
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        rows = EmployeeHours.query.filter(
            EmployeeHours.employee_id == emp.id,
            EmployeeHours.date >= start_of_week,
            EmployeeHours.date <= end_of_week
        ).order_by(EmployeeHours.date.asc()).all()

        # Calculate totals
        total_work = timedelta()
        total_break = timedelta()

        for r in rows:
            if r.sign_in and r.sign_out:
                total_work += (r.sign_out - r.sign_in)

            if r.break_start and r.break_end:
                total_break += (r.break_end - r.break_start)

        net_hours = total_work - total_break

        return render_template(
            "employee_weekly_summary.html",
            emp=emp,
            rows=rows,
            start_of_week=start_of_week,
            end_of_week=end_of_week,
            total_work=total_work,
            total_break=total_break,
            net_hours=net_hours
        )


    @app.route("/admin/employees/hours/<int:row_id>/edit", methods=["POST"])
    def admin_edit_hours_save(row_id):
        row = EmployeeHours.query.get(row_id)
        if not row:
            return "Not found", 404

        from datetime import datetime

        def parse_or_none(val):
            return datetime.fromisoformat(val) if val else None

        row.sign_in = parse_or_none(request.form.get("sign_in"))
        row.break_start = parse_or_none(request.form.get("break_start"))
        row.break_end = parse_or_none(request.form.get("break_end"))
        row.sign_out = parse_or_none(request.form.get("sign_out"))

        # Mark as corrected by admin
        row.corrected = True
        row.corrected_at = datetime.now()
        row.corrected_by = "admin"

        db.session.commit()

        return redirect(f"/admin/employees/{row.employee_id}/hours")

    @app.route("/employeehours/action/<action>/<date>", methods=["POST"])
    def employee_action(action, date):
        if "emp_id" not in session:
            return redirect("/employeehours")

        emp_id = session["emp_id"]
        d = datetime.strptime(date, "%Y-%m-%d").date()

        row = Hours.query.filter_by(employee_id=emp_id, date=d).first()
        if not row:
            row = Hours(employee_id=emp_id, date=d)
            db.session.add(row)

        time_str = request.form.get("time")
        notes = request.form.get("notes", "")

        t = datetime.strptime(time_str, "%H:%M").time()
        dt = datetime.combine(d, t)

        field_map = {
            "start": "sign_in",
            "break_start": "break_start",
            "break_end": "break_end",
            "finish": "sign_out"
        }

        field = field_map.get(action)
        if field:
            setattr(row, field, dt)

        if notes:
            row.notes = notes

        db.session.commit()
        return redirect("/employeehours/week")






    @app.route("/admin/employeehours")
    def admin_employee_hours():
        rows = EmployeeHours.query.order_by(EmployeeHours.date.desc()).all()
        return render_template("admin_employee_hours.html", rows=rows)


    @app.route("/admin/employees/deactivate/<int:emp_id>", methods=["POST"])
    def admin_deactivate_employee(emp_id):
        emp = Employee.query.get(emp_id)
        if emp:
            emp.active = False
            db.session.commit()
        return {"status": "ok"}


    # -------------------------------
    # ADMIN: EMPLOYEE MANAGEMENT
    # -------------------------------

    @app.route("/admin/employees")
    def admin_employees():
        employees = Employee.query.order_by(Employee.full_name).all()
        return render_template("admin_employees.html", employees=employees)


    @app.route("/admin/employees/add", methods=["GET", "POST"])
    def admin_add_employee():
        if request.method == "POST":
            full_name = request.form["full_name"].strip()
            phone = request.form["phone"].strip()

            emp = Employee(
                full_name=full_name,
                phone=phone,
                active=True
            )

            db.session.add(emp)
            db.session.commit()
            flash("Employee added.", "success")
            return redirect(url_for("admin_employees"))

        return render_template("admin_employee_add.html")


    @app.route("/admin/employees/<int:emp_id>/edit", methods=["GET", "POST"])
    def admin_edit_employee(emp_id):
        emp = Employee.query.get_or_404(emp_id)

        if request.method == "POST":
            emp.full_name = request.form.get("full_name", "").strip()
            emp.phone = request.form.get("phone", "").strip()
            emp.role = request.form.get("role", "").strip()
            emp.active = True if request.form.get("active") == "on" else False

            db.session.commit()
            flash("Employee updated.", "success")
            return redirect(url_for("admin_employees"))

        return render_template("admin_employee_edit.html", emp=emp)

    @app.route("/admin/employees/<int:emp_id>/delete", methods=["POST"])
    def admin_delete_employee(emp_id):
        emp = Employee.query.get_or_404(emp_id)

        # Soft delete — keep hours history
        emp.active = False
        emp.setup_code = None
        emp.pin_hash = None

        db.session.commit()

        flash("Employee deactivated.", "info")
        return redirect(url_for("admin_employees"))


    @app.route("/employee/setup")
    def employee_setup_page():
        return render_template("employee_setup.html")

    @app.route("/employeehours/week")
    def employee_week_view():
        emp_id = session.get("employee_id")
        if not emp_id:
            return redirect("/employeehours")

        emp = Employee.query.get(emp_id)  # ⭐ load employee for welcome message

        today = datetime.now(ZoneInfo("Australia/Brisbane")).date()
        start_of_week = today - timedelta(days=today.weekday())
        days = []

        for i in range(7):
            d = start_of_week + timedelta(days=i)
            row = EmployeeHours.query.filter_by(employee_id=emp_id, date=d).first()

            if d > today:
                status = "future"
            elif row and row.sign_in and row.sign_out:
                status = "complete"
            elif row and row.sign_in and not row.sign_out:
                status = "incomplete"
            elif d == today:
                status = "today"
            else:
                status = "incomplete"

            days.append({
                "date": d,
                "status": status
            })

        return render_template(
            "employee_week_view.html",
            days=days,
            today=today,
            emp=emp,
            is_last_week=False
        )





    @app.route("/employeehours/action/start", methods=["POST"])
    def action_start_work():
        emp_id = session.get("employee_id")
        if not emp_id:
            return jsonify({"error": "Not logged in"}), 401

        date_raw = request.form.get("date")
        time_raw = request.form.get("time")

        selected_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
        selected_dt = parse_dt(time_raw)

        row = EmployeeHours.query.filter_by(employee_id=emp_id, date=selected_date).first()

        # 🔒 BLOCK IF DAY COMPLETE
        if row and row.sign_out:
            return jsonify({"error": "Day is complete and cannot be edited"}), 403

        if selected_dt.date() != selected_date:
            return jsonify({"error": "Start Work must be on the same day"}), 400

        if selected_dt > datetime.now():
            return jsonify({"error": "Cannot start work in the future"}), 400

        if not row:
            row = EmployeeHours(employee_id=emp_id, date=selected_date)
            db.session.add(row)

        if row.sign_in:
            return jsonify({"error": "Start Work already recorded"}), 400

        row.sign_in = selected_dt
        db.session.commit()

        return jsonify({"status": "ok"})


    @app.route("/employeehours/action/break_start", methods=["POST"])
    def action_break_start():
        emp_id = session.get("employee_id")
        if not emp_id:
            return jsonify({"error": "Not logged in"}), 401

        date_raw = request.form.get("date")
        time_raw = request.form.get("time")

        selected_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
        selected_dt = parse_dt(time_raw)

        row = EmployeeHours.query.filter_by(employee_id=emp_id, date=selected_date).first()

        # 🔒 BLOCK IF DAY COMPLETE
        if row and row.sign_out:
            return jsonify({"error": "Day is complete and cannot be edited"}), 403

        if not row or not row.sign_in:
            return jsonify({"error": "Start Work required first"}), 400

        if row.break_start:
            return jsonify({"error": "Break Start already recorded"}), 400

        if selected_dt <= row.sign_in:
            return jsonify({"error": "Break Start must be after Start Work"}), 400

        if selected_dt > datetime.now():
            return jsonify({"error": "Cannot start break in the future"}), 400

        row.break_start = selected_dt
        db.session.commit()

        return jsonify({"status": "ok"})


    @app.route("/employeehours/action/break_end", methods=["POST"])
    def action_break_end():
        emp_id = session.get("employee_id")
        if not emp_id:
            return jsonify({"error": "Not logged in"}), 401

        date_raw = request.form.get("date")
        time_raw = request.form.get("time")

        selected_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
        selected_dt = parse_dt(time_raw)

        row = EmployeeHours.query.filter_by(employee_id=emp_id, date=selected_date).first()

        # 🔒 BLOCK IF DAY COMPLETE
        if row and row.sign_out:
            return jsonify({"error": "Day is complete and cannot be edited"}), 403

        if not row or not row.break_start:
            return jsonify({"error": "Break Start required first"}), 400

        if row.break_end:
            return jsonify({"error": "Break End already recorded"}), 400

        if selected_dt <= row.break_start:
            return jsonify({"error": "Break End must be after Break Start"}), 400

        if selected_dt > datetime.now():
            return jsonify({"error": "Cannot end break in the future"}), 400

        row.break_end = selected_dt
        db.session.commit()

        return jsonify({"status": "ok"})


    @app.route("/employeehours/action/finish", methods=["POST"])
    def action_finish_work():
        emp_id = session.get("employee_id")
        if not emp_id:
            return jsonify({"error": "Not logged in"}), 401

        date_raw = request.form.get("date")
        time_raw = request.form.get("time")

        selected_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
        selected_dt = parse_dt(time_raw)

        row = EmployeeHours.query.filter_by(employee_id=emp_id, date=selected_date).first()

        if not row or not row.sign_in:
            return jsonify({"error": "Start Work required first"}), 400

        if row.sign_out:
            return jsonify({"error": "Finish Work already recorded"}), 400

        if selected_dt <= row.sign_in:
            return jsonify({"error": "Finish Work must be after Start Work"}), 400

        if row.break_end and selected_dt <= row.break_end:
            return jsonify({"error": "Finish Work must be after Break End"}), 400

        if selected_dt > datetime.now():
            return jsonify({"error": "Cannot finish work in the future"}), 400

        # ---------------------------------------------------------
        #   CALCULATE PAID HOURS (BREAK DEDUCTED)
        # ---------------------------------------------------------
        shift_seconds = (selected_dt - row.sign_in).total_seconds()

        break_seconds = 0
        if row.break_start and row.break_end:
            break_seconds = (row.break_end - row.break_start).total_seconds()

        paid_seconds = shift_seconds - break_seconds
        hours = round(paid_seconds / 3600, 2)

        # ---------------------------------------------------------
        #   SEND CONFIRMATION BACK TO FRONTEND
        # ---------------------------------------------------------
        return f"CONFIRM_SHIFT::{selected_dt.strftime('%-I:%M %p')}::{hours}"



    @app.route("/admin/employeehours/day/<string:day>/<int:emp_id>")
    def admin_override_day(day, emp_id):
        try:
            selected_date = datetime.strptime(day, "%Y-%m-%d").date()
        except:
            return "Invalid date", 400

        row = EmployeeHours.query.filter_by(employee_id=emp_id, date=selected_date).first()

        return render_template(
            "admin_override_day.html",
            date=selected_date,
            emp_id=emp_id,
            row=row
        )

    @app.route("/admin/employees/<int:emp_id>/setup_qr")
    def admin_employee_setup_qr(emp_id):
        emp = Employee.query.get_or_404(emp_id)

        if not emp.setup_code:
            return "No setup code. Reset PIN first.", 400

        setup_url = f"https://cherbonapp.click/employeehours/setup?code={emp.setup_code}"

        import qrcode
        import base64
        from io import BytesIO

        img = qrcode.make(setup_url)
        buf = BytesIO()
        img.save(buf, format="PNG")
        qr_b64 = base64.b64encode(buf.getvalue()).decode()

        return render_template(
            "admin_employee_setup_qr.html",
            emp=emp,
            setup_url=setup_url,
            qr_b64=qr_b64
        )
    @app.route("/employeehours/setup", methods=["GET", "POST"])
    def employeehours_setup():
        code = request.args.get("code") if request.method == "GET" else request.form.get("code")
        if not code:
            return "Missing setup code", 400

        emp = Employee.query.filter_by(setup_code=code).first()
        if not emp:
            return "Invalid or expired setup code", 400

        if request.method == "POST":
            pin = request.form.get("pin", "")
            confirm = request.form.get("confirm", "")

            # PIN rules
            if len(pin) != 6 or not pin.isdigit():
                return "PIN must be exactly 6 digits", 400

            if pin != confirm:
                return "PINs do not match", 400

            # No sequences
            if pin in ("012345", "123456", "234567", "345678", "456789"):
                return "PIN cannot be a sequence", 400

            # No repeats
            if len(set(pin)) == 1:
                return "PIN cannot be repeating digits", 400

            # Hash + save
            from werkzeug.security import generate_password_hash
            emp.pin_hash = generate_password_hash(pin)
            emp.setup_code = None  # one-time use
            emp.pin_failures = 0
            emp.locked_until = None

            db.session.commit()

            return redirect("/employeehours")

        return render_template("employee_setup.html", emp=emp, code=code)



    @app.route("/employeehours")
    def employeehours_login_page():
        # Already logged in → go to week view
        if "employee_id" in session:
            return redirect("/employeehours/week")

        # If a setup code is in the URL → go to setup page
        code = request.args.get("code")
        if code:
            return redirect(url_for("employeehours_setup", code=code))

        # Otherwise show PIN login page
        return render_template("employee_pin_login.html")



    @app.route("/admin/corrections")
    def admin_corrections():
        rows = EmployeeHours.query.filter(
            EmployeeHours.corrected == True
        ).order_by(EmployeeHours.corrected_at.desc()).all()

        return render_template("admin_corrections.html", rows=rows)


    @app.route("/employees")
    def employees_home():
        return render_template("employees_home.html")




    @app.route("/other_tools")
    def other_tools():
        return render_template("other_tools.html")

    @app.route('/nuke_aws_backups', methods=['POST'])
    def nuke_aws_backups():
        import boto3
        from flask import request, jsonify, current_app, flash
        from werkzeug.security import check_password_hash

        data = request.get_json() or {}
        confirm_phrase = data.get('confirm_phrase', '')
        admin_password = data.get('admin_password', '')

        # 1) Validate phrase
        if confirm_phrase != "DELETE EVERYTHING FOREVER":
            return jsonify({"status": "error", "message": "Invalid confirmation phrase"}), 400

        # 2) Validate admin password
        admin_user = Users.query.filter_by(username='steve').first()
        if not admin_user or not check_password_hash(admin_user.password_hash, admin_password):
            return jsonify({"status": "error", "message": "Invalid admin credentials"}), 403

        # 3) Log the event
        current_app.logger.warning("NUKE AWS BACKUPS triggered by %s", admin_user.username)

        # 4) AWS RDS client
        rds = boto3.client('rds', region_name='ap-southeast-2')

        # *** IMPORTANT ***
        db_identifier = "cherboneq"

        # 5) Set retention to 0 (kills PITR + future automated backups)
        try:
            rds.modify_db_instance(
                DBInstanceIdentifier=db_identifier,
                BackupRetentionPeriod=0,
                ApplyImmediately=True
            )
            current_app.logger.warning("Backup retention set to 0 for %s", db_identifier)
        except Exception as e:
            current_app.logger.error("Error setting retention to 0: %s", e)

        # 6) Delete ALL manual snapshots
        try:
            manual_snaps = rds.describe_db_snapshots(
                DBInstanceIdentifier=db_identifier,
                SnapshotType='manual'
            )['DBSnapshots']

            for snap in manual_snaps:
                snap_id = snap['DBSnapshotIdentifier']
                current_app.logger.warning("Deleting manual snapshot: %s", snap_id)
                rds.delete_db_snapshot(DBSnapshotIdentifier=snap_id)

        except Exception as e:
            current_app.logger.error("Error deleting manual snapshots: %s", e)

        # 7) Delete ALL automated snapshots
        try:
            auto_snaps = rds.describe_db_snapshots(
                DBInstanceIdentifier=db_identifier,
                SnapshotType='automated'
            )['DBSnapshots']

            for snap in auto_snaps:
                snap_id = snap['DBSnapshotIdentifier']
                current_app.logger.warning("Deleting automated snapshot: %s", snap_id)
                rds.delete_db_snapshot(DBSnapshotIdentifier=snap_id)

        except Exception as e:
            current_app.logger.error("Error deleting automated snapshots: %s", e)

        flash("AWS backups, snapshots, and PITR logs scheduled for deletion.", "warning")
        return jsonify({"status": "ok"})


    @app.route("/bulk_update_lessons", methods=["POST"])
    def bulk_update_lessons():
        data = request.get_json() or {}
        updates = data.get("updates", [])

        if not isinstance(updates, list):
            return jsonify({"status": "error", "message": "Invalid payload"}), 400

        try:
            for item in updates:
                lesson_id = item.get("lesson_id")
                if not lesson_id:
                    continue

                lesson = Lesson.query.get(int(lesson_id))
                if not lesson:
                    continue

                # Update lesson fields
                lesson.attendance = item.get("attendance", "")
                lesson.payment    = parse_money(item.get("payment", "0"))
                lesson.price_pl   = parse_money(item.get("price_pl", "0"))
                lesson.adjust     = parse_money(item.get("adjust", "0"))

                # Update CLIENT notes
                client = Client.query.filter_by(full_name=lesson.client).first()
                if client:
                    client.notes = item.get("notes", "")

            import time
            start = time.time()
            db.session.commit()
            print("COMMIT TIME:", time.time() - start)

            return jsonify({"status": "ok"}), 200

        except Exception as e:
            db.session.rollback()
            print("Bulk update error:", e)
            return jsonify({"status": "error", "message": str(e)}), 500


    @app.post("/bulk_delete_day")
    def bulk_delete_day():
        # Reset poisoned session
        db.session.rollback()

        cutoff = request.form.get("cutoff_date")
        gp = request.form.get("group_priv")
        client_name = request.args.get("client")

        if not cutoff:
            return {"status": "error", "message": "Missing cutoff_date"}

        from datetime import datetime
        cutoff_date = datetime.strptime(cutoff, "%Y-%m-%d").date()

        db.session.execute(text("""
            DELETE FROM lessons
            WHERE client = :cname
            AND lesson_date = :cut
            AND group_priv = :gp
        """), {"cname": client_name, "cut": cutoff_date, "gp": gp})

        db.session.commit()

        return redirect(url_for("client_view", client=client_name))


    @app.route('/enquiries')
    def enquiries_home():
        return render_template('enquiries_home.html')


    @app.route('/client_view')
    def client_view():
        client_id = request.args.get('client', type=int) or request.args.get('client_filter', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = int(request.args.get('page', 1))
        per_page = 50

        lessons = []
        total = 0
        client_obj = None

        # ---------------------------------------------------------
        # LOAD ALL CLIENTS
        # ---------------------------------------------------------
        clients = db.session.query(Client).order_by(Client.full_name.asc()).all()
        client_names = [(c.client_id, c.full_name) for c in clients]

        # ---------------------------------------------------------
        # LOAD ALL HORSES
        # ---------------------------------------------------------
        horses = db.session.query(Horse).order_by(Horse.horse.asc()).all()

        # ---------------------------------------------------------
        # LOAD ALL TIMES  <-- REQUIRED FOR CHANGE TIME TOOL
        # ---------------------------------------------------------
        times = db.session.query(Time).order_by(Time.timerange.asc()).all()


        # ---------------------------------------------------------
        # LOOK UP SELECTED CLIENT
        # ---------------------------------------------------------
        if client_id:
            client_obj = Client.query.get(client_id)

            if client_obj:
                query = db.session.query(Lesson).filter(
                    Lesson.client == client_obj.full_name
                )

                if start_date:
                    query = query.filter(Lesson.lesson_date >= start_date)
                if end_date:
                    query = query.filter(Lesson.lesson_date <= end_date)

                total = query.count()

                lessons = (
                    query.order_by(
                        db.func.date(Lesson.lesson_date).desc(),
                        Lesson.lesson_id.desc()
                    )
                    .offset((page - 1) * per_page)
                    .limit(per_page)
                    .all()
                )

        return render_template(
            'client_view.html',
            lessons=lessons,
            client_filter=client_id,
            start_date=start_date,
            end_date=end_date,
            page=page,
            total=total,
            per_page=per_page,
            client_names=client_names,
            client_obj=client_obj,
            horses=horses,
            times=times,                     # <-- NOW INCLUDED
            today=date.today().isoformat()
        )

    return app


app = create_app()


if __name__ == '__main__':
    app = create_app()
    app.run(host='0.0.0.0', port=5000, debug=True)